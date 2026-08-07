#!/usr/bin/env python3
"""抖音无水印下载器 · Web 服务版（含代理池 + 管理后台）

无需登录抖音账号或客户端签名。后端负责：短链解析 → 分享页元数据提取 → 无水印地址还原，
并以流式代理（支持 Range）转发视频/图片，绕过抖音 CDN 的 UA / 防盗链限制，
让浏览器可以直接在线播放与下载。

反封锁能力：
  · 代理 IP 池（http/https/socks5），所有出站请求轮换走代理
  · 失败自动转移到下一个代理 + 失败计数退避
  · 移动端 UA 池轮换 + Referer 伪装
  · 管理后台（密码鉴权）增删/启停/测试代理、查看出口 IP 与统计

启动:  uvicorn server:app --host 0.0.0.0 --port 8000 --no-access-log
环境变量:  ADMIN_PASSWORD  管理后台密码（默认 douyin-admin，生产务必修改）
"""

import gzip
import hashlib
import hmac
import io
import json
import os
import platform
import queue
import random
import re
import secrets
import shutil
import sqlite3
import subprocess
import threading
import time
from pathlib import Path
from typing import Optional
from urllib import error as urlerr
from urllib import parse as urlparse
from urllib import request as urlreq

from fastapi import FastAPI, Request
from fastapi.responses import (FileResponse, HTMLResponse, JSONResponse,
                               PlainTextResponse, Response, StreamingResponse)
from pydantic import BaseModel

# ---------------------------------------------------------------- 常量与存储

# 版本号（语义化：修 bug +patch，新功能 +minor，不兼容改动 +major）。
# 每次改动必须同步更新 README.md 顶部版本号与「更新日志」，规则见 CLAUDE.md。
APP_VERSION = "1.13.0"
_BUILD_DATE = time.strftime("%Y-%m-%d", time.gmtime())  # 进程启动日期，供 sitemap lastmod

DATA_DIR = Path(os.environ.get("DATA_DIR", "data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
try:
    DATA_DIR.chmod(0o700)
except OSError:
    pass
STORE_FILE = DATA_DIR / "config.json"


def _load_app_secret() -> bytes:
    """读取稳定密钥；未配置时用完整临时文件 + hard-link 原子持久化。"""
    configured = (os.environ.get("APP_SECRET")
                  or os.environ.get("CAPTCHA_SECRET") or "").strip()
    if configured:
        if len(configured.encode()) < 32:
            raise RuntimeError("APP_SECRET/CAPTCHA_SECRET 至少需要 32 字节")
        return configured.encode()
    path = DATA_DIR / ".app-secret"
    try:
        saved = path.read_text("utf-8").strip()
        if saved:
            path.chmod(0o600)
            return saved.encode()
    except FileNotFoundError:
        pass

    candidate = secrets.token_urlsafe(48)
    tmp = DATA_DIR / f".app-secret.{os.getpid()}.{secrets.token_hex(6)}.tmp"
    fd = os.open(tmp, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(candidate)
            f.flush()
            os.fsync(f.fileno())
        try:
            os.link(tmp, path)
        except FileExistsError:
            pass
    finally:
        try:
            tmp.unlink()
        except FileNotFoundError:
            pass
    saved = path.read_text("utf-8").strip()
    if not saved:
        raise RuntimeError(f"应用密钥文件为空，请删除后重启或设置 APP_SECRET：{path}")
    path.chmod(0o600)
    return saved.encode()


APP_SECRET = _load_app_secret()

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "douyin-admin")
if ADMIN_PASSWORD == "douyin-admin":
    import sys as _sys
    print("⚠️  警告：正在使用默认管理员密码，请设置环境变量 ADMIN_PASSWORD 后再对外部署！",
          file=_sys.stderr)

# 免费使用配额（防薅羊毛）
FREE_ANON_DAILY = int(os.environ.get("FREE_ANON_DAILY", "3"))    # 匿名：每天 3 次
FREE_USER_DAILY = int(os.environ.get("FREE_USER_DAILY", "10"))   # 登录用户：每天 10 次


def _clamped_env_int(name: str, default: int,
                     minimum: int, maximum: int) -> int:
    try:
        value = int(os.environ.get(name, str(default)))
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(maximum, value))


MEDIA_TOKEN_TTL = max(300, min(86400, int(os.environ.get("MEDIA_TOKEN_TTL", "43200"))))
MEDIA_REQUESTS_PER_MIN = max(10, int(os.environ.get("MEDIA_REQUESTS_PER_MIN", "120")))
MEDIA_MAX_CONCURRENT = max(1, int(os.environ.get("MEDIA_MAX_CONCURRENT", "6")))
MEDIA_RESUME_MAX_ATTEMPTS = _clamped_env_int(
    "MEDIA_RESUME_MAX_ATTEMPTS", 64, 1, 256)
MEDIA_RESUME_MAX_SECONDS = _clamped_env_int(
    "MEDIA_RESUME_MAX_SECONDS", 3600, 30, 7200)
MEDIA_RESUME_MAX_FAILURES = _clamped_env_int(
    "MEDIA_RESUME_MAX_FAILURES", 8, 2, 16)
DATA_RETENTION_DAYS = max(
    1, min(30, int(os.environ.get("DATA_RETENTION_DAYS", "30")))
)
API_JOB_WORKERS = max(1, min(8, int(os.environ.get("API_JOB_WORKERS", "2"))))
QUOTA_RESERVATION_TTL = max(300, int(os.environ.get("QUOTA_RESERVATION_TTL", "3600")))

# ---------------------------------------------------------------- SQLite 数据层

DB_FILE = DATA_DIR / "app.db"
_db_lock = threading.Lock()

_SCHEMA = """
CREATE TABLE IF NOT EXISTS usage_daily(
  day INTEGER, subject TEXT, count INTEGER DEFAULT 0,
  PRIMARY KEY(day, subject)
);
CREATE TABLE IF NOT EXISTS quota_reservations(
  id TEXT PRIMARY KEY, day INTEGER, subjects TEXT, units INTEGER,
  committed_units INTEGER DEFAULT 0, status TEXT,
  endpoint TEXT, created INTEGER, settled INTEGER, lease_until INTEGER
);
CREATE INDEX IF NOT EXISTS idx_quota_reservation_lease
  ON quota_reservations(status, lease_until);
CREATE TABLE IF NOT EXISTS request_logs(
  id INTEGER PRIMARY KEY AUTOINCREMENT, ts INTEGER, kind TEXT, subject TEXT,
  ip TEXT, ua TEXT, link TEXT, ok INTEGER, path TEXT, user_id INTEGER
);
CREATE INDEX IF NOT EXISTS idx_reqlog_ts ON request_logs(ts);
CREATE TABLE IF NOT EXISTS page_views(
  id INTEGER PRIMARY KEY AUTOINCREMENT, ts INTEGER, ip TEXT, ua TEXT, path TEXT, fp TEXT
);
CREATE INDEX IF NOT EXISTS idx_pv_ts ON page_views(ts);
CREATE TABLE IF NOT EXISTS users(
  id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT UNIQUE, pw_salt TEXT, pw_hash TEXT,
  created_at INTEGER, last_login INTEGER, disabled INTEGER DEFAULT 0, reg_ip TEXT
);
CREATE TABLE IF NOT EXISTS api_keys(
  key TEXT PRIMARY KEY, user_id INTEGER, name TEXT, created INTEGER, enabled INTEGER DEFAULT 1,
  balance_cents INTEGER DEFAULT 100, spent_cents INTEGER DEFAULT 0, calls INTEGER DEFAULT 0,
  last_used INTEGER, reserved_cents INTEGER DEFAULT 0, deleted_at INTEGER
);
CREATE TABLE IF NOT EXISTS jobs(
  id TEXT PRIMARY KEY, key TEXT, user_id INTEGER, status TEXT, total INTEGER, done INTEGER DEFAULT 0,
  ok INTEGER DEFAULT 0, cost_cents INTEGER DEFAULT 0, links TEXT, results TEXT,
  created INTEGER, finished INTEGER, price_cents INTEGER DEFAULT 0,
  updated INTEGER, request_id TEXT
);
CREATE INDEX IF NOT EXISTS idx_jobs_key ON jobs(key);
CREATE TABLE IF NOT EXISTS job_items(
  job_id TEXT, idx INTEGER, link TEXT, status TEXT DEFAULT 'pending',
  price_cents INTEGER DEFAULT 0, reserved INTEGER DEFAULT 0,
  result TEXT, error TEXT, attempts INTEGER DEFAULT 0,
  lease_owner TEXT, lease_until INTEGER,
  started INTEGER, finished INTEGER,
  PRIMARY KEY(job_id, idx)
);
CREATE INDEX IF NOT EXISTS idx_job_items_status ON job_items(status, job_id);
CREATE INDEX IF NOT EXISTS idx_job_items_claim
  ON job_items(status, lease_until, job_id, idx);
CREATE TABLE IF NOT EXISTS api_logs(
  id INTEGER PRIMARY KEY AUTOINCREMENT, ts INTEGER, key TEXT, user_id INTEGER,
  link TEXT, ok INTEGER, cost_cents INTEGER, job_id TEXT, item_idx INTEGER
);
CREATE INDEX IF NOT EXISTS idx_apilog_ts ON api_logs(ts);
CREATE TABLE IF NOT EXISTS api_ledger(
  id INTEGER PRIMARY KEY AUTOINCREMENT, ts INTEGER, key TEXT,
  job_id TEXT, item_idx INTEGER, event TEXT,
  balance_delta INTEGER DEFAULT 0, reserved_delta INTEGER DEFAULT 0,
  spent_delta INTEGER DEFAULT 0, calls_delta INTEGER DEFAULT 0, reason TEXT
);
CREATE INDEX IF NOT EXISTS idx_ledger_job ON api_ledger(job_id, item_idx);
CREATE TABLE IF NOT EXISTS app_settings(k TEXT PRIMARY KEY, v TEXT);
CREATE TABLE IF NOT EXISTS shares(
  id TEXT PRIMARY KEY, item_id TEXT, kind TEXT, vid TEXT,
  owner_user_id INTEGER, owner_fp TEXT, owner_ip TEXT,
  title TEXT, author TEXT, avatar TEXT, cover TEXT,
  payload TEXT, custom_title TEXT,
  visibility TEXT DEFAULT 'link', pw_salt TEXT, pw_hash TEXT,
  expires_at INTEGER DEFAULT 0, refreshed_at INTEGER,
  status TEXT DEFAULT 'ok',
  views INTEGER DEFAULT 0, plays INTEGER DEFAULT 0,
  downloads INTEGER DEFAULT 0, cta_clicks INTEGER DEFAULT 0,
  created INTEGER
);
CREATE INDEX IF NOT EXISTS idx_shares_owner ON shares(owner_user_id);
CREATE INDEX IF NOT EXISTS idx_shares_item ON shares(item_id);
CREATE TABLE IF NOT EXISTS share_events(
  id INTEGER PRIMARY KEY AUTOINCREMENT, ts INTEGER, sid TEXT, kind TEXT,
  ip TEXT, ua TEXT, referer TEXT, wechat INTEGER, fp TEXT,
  source TEXT, stage TEXT, detail TEXT, ms INTEGER
);
CREATE INDEX IF NOT EXISTS idx_share_ev ON share_events(ts, sid);
CREATE INDEX IF NOT EXISTS idx_share_ev_kind ON share_events(kind, ts);
CREATE TABLE IF NOT EXISTS reports(
  id INTEGER PRIMARY KEY AUTOINCREMENT, ts INTEGER, sid TEXT,
  reason TEXT, contact TEXT, ip TEXT, handled INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS media_traffic(
  day INTEGER, scope TEXT,
  requests INTEGER DEFAULT 0, bytes INTEGER DEFAULT 0,
  PRIMARY KEY(day, scope)
);
-- AnyToCopy 增强线路：结果缓存（按 item_id 全站共享，热门视频只调一次 API）
CREATE TABLE IF NOT EXISTS atc_cache(
  item_id TEXT PRIMARY KEY,
  video_url TEXT, url_fetched_at INTEGER,
  content TEXT, text_content TEXT,
  audio_url TEXT, duration REAL,
  created INTEGER, updated INTEGER
);
-- ATC 内部任务队列（对方并发上限 5，串行化提交；重启后按 task_id 续查）
CREATE TABLE IF NOT EXISTS atc_jobs(
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  item_id TEXT, work_url TEXT,
  purpose TEXT,
  task_id TEXT,
  status TEXT DEFAULT 'pending',
  error TEXT, created INTEGER, updated INTEGER
);
CREATE INDEX IF NOT EXISTS idx_atc_jobs_status ON atc_jobs(status, id);
"""


def _db():
    conn = sqlite3.connect(DB_FILE, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout=10000")
    return conn


def db_exec(sql: str, params=(), fetch: Optional[str] = None):
    with _db_lock:
        conn = _db()
        try:
            cur = conn.execute(sql, params)
            out = (cur.fetchone() if fetch == "one"
                   else cur.fetchall() if fetch == "all"
                   else cur.rowcount if fetch == "rowcount"
                   else cur.lastrowid)
            conn.commit()
            return out
        finally:
            conn.close()


def _privacy_hash(kind: str, value: str, scope: str = "") -> str:
    """把网络标识转为本站不可逆 HMAC，避免在数据库中保存原始 IP/指纹。"""
    value = str(value or "").strip()
    if not value:
        return ""
    digest = hmac.new(APP_SECRET, f"{kind}:{scope}:{value}".encode(),
                      hashlib.sha256).hexdigest()[:24]
    return f"h:{digest}"


def _safe_referer(value: str) -> str:
    """埋点只保留来源的 scheme/host/path，丢弃可能含个人信息的 query/fragment。"""
    try:
        p = urlparse.urlsplit(value or "")
        return urlparse.urlunsplit((p.scheme, p.netloc, p.path, "", ""))[:200]
    except Exception:
        return ""


def _migrate_privacy_data(conn) -> None:
    """一次性把老库中的原始网络标识就地改成带 h: 前缀的不可逆 HMAC。"""
    columns = (("request_logs", "ip", "request-ip"),
               ("page_views", "ip", "analytics-visitor"))
    for table, col, kind in columns:
        rows = conn.execute(
            f"SELECT rowid,{col} FROM {table} "
            f"WHERE COALESCE({col},'')<>'' AND {col} NOT LIKE 'h:%'"
        ).fetchall()
        for rowid, value in rows:
            conn.execute(f"UPDATE {table} SET {col}=? WHERE rowid=?",
                         (_privacy_hash(kind, value), rowid))

    rows = conn.execute(
        "SELECT rowid,subject FROM request_logs "
        "WHERE COALESCE(subject,'')<>'' AND subject NOT LIKE 'h:%' "
        "AND subject NOT LIKE 'user:%'"
    ).fetchall()
    for rowid, value in rows:
        conn.execute("UPDATE request_logs SET subject=? WHERE rowid=?",
                     (_privacy_hash("request-subject", value), rowid))
    rows = conn.execute(
        "SELECT rowid,link FROM request_logs WHERE COALESCE(link,'')<>'' "
        "AND link NOT LIKE 'h:%'"
    ).fetchall()
    for rowid, value in rows:
        conn.execute("UPDATE request_logs SET link=? WHERE rowid=?",
                     (_privacy_hash("submitted-link", value)[:26], rowid))

    # 这些旧字段没有业务读取用途，直接清空比继续保留可关联摘要更符合最小化原则。
    conn.execute("UPDATE users SET reg_ip='' WHERE COALESCE(reg_ip,'')<>''")
    conn.execute(
        "UPDATE shares SET owner_ip='',owner_fp='' "
        "WHERE COALESCE(owner_ip,'')<>'' OR COALESCE(owner_fp,'')<>''")
    conn.execute(
        "UPDATE share_events SET ip='',fp='',referer='',ua='' "
        "WHERE COALESCE(ip,'')<>'' OR COALESCE(fp,'')<>'' "
        "OR COALESCE(referer,'')<>'' OR COALESCE(ua,'')<>''")
    conn.execute(
        "UPDATE page_views SET ua='',fp='' "
        "WHERE COALESCE(ua,'')<>'' OR COALESCE(fp,'')<>''")
    conn.execute("UPDATE request_logs SET ua='' WHERE COALESCE(ua,'')<>''")
    conn.execute("UPDATE reports SET ip='' WHERE COALESCE(ip,'')<>''")

    # 免费额度主体也不能含原始 IP/浏览器指纹；迁移时取 MAX 防止计数被意外叠加。
    rows = conn.execute(
        "SELECT day,subject,count FROM usage_daily "
        "WHERE (subject LIKE 'ip:%' AND subject NOT LIKE 'ip:h:%') "
        "OR (subject LIKE 'fp:%' AND subject NOT LIKE 'fp:h:%')"
    ).fetchall()
    for day, subject, count in rows:
        kind, raw = subject.split(":", 1)
        new_subject = f"{kind}:{_privacy_hash(f'quota-{kind}', raw, str(day))}"
        conn.execute(
            "INSERT INTO usage_daily(day,subject,count) VALUES(?,?,?) "
            "ON CONFLICT(day,subject) DO UPDATE SET count=MAX(count,excluded.count)",
            (day, new_subject, count))
        conn.execute("DELETE FROM usage_daily WHERE day=? AND subject=?",
                     (day, subject))

with _db_lock:
    _c = _db()
    _c.execute("PRAGMA journal_mode=WAL")      # 允许并发读，写不阻塞读
    _c.executescript(_SCHEMA)
    _c.execute("CREATE INDEX IF NOT EXISTS idx_reqlog_user ON request_logs(user_id)")
    # 就地补列（无迁移框架）：所有 ALTER 都按 PRAGMA 探测，老库可直接升级。
    def _ensure_columns(table, columns):
        have = {r[1] for r in _c.execute(f"PRAGMA table_info({table})")}
        for col, typ in columns:
            if col not in have:
                _c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {typ}")

    _ensure_columns("share_events", (
        ("source", "TEXT"), ("stage", "TEXT"), ("detail", "TEXT"), ("ms", "INTEGER"),
        ("next_src", "TEXT")))   # 该线路失败后，链上下一条将重试的线路（dy1/dy2/atc/proxy）
    _ensure_columns("api_keys", (
        ("reserved_cents", "INTEGER DEFAULT 0"), ("deleted_at", "INTEGER")))
    _ensure_columns("jobs", (
        ("price_cents", "INTEGER DEFAULT 0"), ("updated", "INTEGER"),
        ("request_id", "TEXT")))
    _ensure_columns("job_items", (
        ("price_cents", "INTEGER DEFAULT 0"), ("reserved", "INTEGER DEFAULT 0"),
        ("result", "TEXT"), ("error", "TEXT"),
        ("attempts", "INTEGER DEFAULT 0"), ("lease_owner", "TEXT"),
        ("lease_until", "INTEGER"), ("started", "INTEGER"), ("finished", "INTEGER")))
    _ensure_columns("api_logs", (("item_idx", "INTEGER"),))
    _c.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_jobs_idem ON jobs(key,request_id) "
        "WHERE request_id IS NOT NULL")
    _c.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_apilog_item ON api_logs(job_id,item_idx) "
        "WHERE item_idx IS NOT NULL")
    _c.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_ledger_item_reserve "
        "ON api_ledger(job_id,item_idx) WHERE event='reserve'")
    _c.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_ledger_item_settle "
        "ON api_ledger(job_id,item_idx) WHERE event IN ('charge','refund')")
    _c.execute(
        "CREATE INDEX IF NOT EXISTS idx_job_items_claim "
        "ON job_items(status,lease_until,job_id,idx)")
    _migrate_privacy_data(_c)
    _privacy_vacuum_needed = not _c.execute(
        "SELECT 1 FROM app_settings WHERE k='privacy_v2_vacuumed'"
    ).fetchone()
    _c.commit()
    if _privacy_vacuum_needed:
        try:
            _c.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            _c.execute("VACUUM")
            _c.execute(
                "INSERT OR REPLACE INTO app_settings(k,v) VALUES('privacy_v2_vacuumed','1')")
            _c.commit()
        except sqlite3.OperationalError:
            # 滚动部署期间老进程可能仍占用 WAL；下次启动继续尝试，绝不标记为完成。
            pass
    _c.close()


def _secure_data_permissions() -> None:
    """限制数据库、WAL、代理配置和应用密钥仅供服务账号读写。"""
    try:
        DATA_DIR.chmod(0o700)
    except OSError:
        pass
    for path in (DB_FILE, Path(str(DB_FILE) + "-wal"), Path(str(DB_FILE) + "-shm"),
                 STORE_FILE, DATA_DIR / ".app-secret"):
        try:
            if path.exists():
                path.chmod(0o600)
        except OSError:
            pass


_secure_data_permissions()


# ---------------------------------------------------------------- 防薅羊毛 / 限频

def _today() -> int:
    return int(time.time() // 86400)


# 只有来自可信反代时才采信 X-Forwarded-For，否则客户端可伪造头绕过所有基于 IP 的风控。
# 设 TRUST_PROXY=1 表示部署在反代后（Nginx/Cloudflare 等），此时才读 XFF。
TRUST_PROXY = os.environ.get("TRUST_PROXY", "").lower() in ("1", "true", "yes")
TRUST_PROXY_HOPS = max(1, int(os.environ.get("TRUST_PROXY_HOPS", "1")))
# 会话 cookie 是否加 Secure（仅走 HTTPS 发送）。生产（反代/HTTPS）应为真；本地 http 调试默认关。
COOKIE_SECURE = os.environ.get("COOKIE_SECURE", "").lower() in ("1", "true", "yes") or TRUST_PROXY


def _client_ip(request: Request) -> str:
    peer = request.client.host if request.client else "?"
    if TRUST_PROXY:
        xff = request.headers.get("x-forwarded-for")
        if xff:
            # proxy_add_x_forwarded_for 会保留客户端伪造的左侧值；从右侧按可信代理层数取值。
            parts = [p.strip() for p in xff.split(",") if p.strip()]
            if len(parts) >= TRUST_PROXY_HOPS:
                return parts[-TRUST_PROXY_HOPS][:64]
    return str(peer)[:64]


def _client_fp(request: Request) -> str:
    return (request.headers.get("x-fp") or "")[:64]


def _stored_ip(request: Request, purpose: str = "security",
               scope: str = "") -> str:
    return _privacy_hash(f"{purpose}-ip", _client_ip(request), scope)


def _stored_fp(request: Request, purpose: str = "security",
               scope: str = "") -> str:
    return _privacy_hash(f"{purpose}-visitor", _client_fp(request), scope)


def _coarse_ua(request: Request) -> str:
    """只保留兼容诊断需要的粗粒度环境，不落完整 UA、机型或 build 字符串。"""
    ua = request.headers.get("user-agent") or ""
    os_name = "ios" if re.search(r"iphone|ipad|ipod", ua, re.I) else (
        "android" if re.search(r"android", ua, re.I) else (
            "windows" if re.search(r"windows", ua, re.I) else (
                "macos" if re.search(r"mac os", ua, re.I) else "other")))
    browser = "wechat" if re.search(r"micromessenger", ua, re.I) else (
        "chrome" if re.search(r"(?:chrome|crios)/", ua, re.I) else (
            "safari" if re.search(r"safari/", ua, re.I) else "other"))
    m = re.search(r"(?:MicroMessenger|Chrome|CriOS|Version)/(\d+)", ua, re.I)
    return f"{browser}/{m.group(1) if m else 'x'} {os_name}"


def _log_link(value: str) -> str:
    """运营日志只存用途隔离的短摘要，不保存用户粘贴的完整链接/文案。"""
    if not value:
        return ""
    return _privacy_hash("submitted-link", value)[:26]


def _quota_subjects(request: Request):
    """返回 (计数主体列表, 限额)。登录用户按 user 计（10/天），匿名按指纹+IP（3/天）。"""
    u = current_user(request)
    if u:
        return [f"user:{u['id']}"], FREE_USER_DAILY
    scope = str(_today())
    subs = [f"ip:{_stored_ip(request, 'quota', scope)}"]
    fp = _stored_fp(request, "quota", scope)
    if fp:
        subs.append(f"fp:{fp}")
    return subs, FREE_ANON_DAILY


def quota_status(request: Request):
    """返回 (limit, used, remaining)。"""
    day = _today()
    subs, limit = _quota_subjects(request)
    marks = ",".join("?" for _ in subs)
    with _db_lock:
        conn = _db()
        try:
            rows = conn.execute(
                f"SELECT count FROM usage_daily WHERE day=? AND subject IN ({marks})",
                (day, *subs)).fetchall()
        finally:
            conn.close()
    used = max((int(r[0]) for r in rows), default=0)
    return limit, used, max(0, limit - used)


def reserve_quota(request: Request, n: int = 1, partial: bool = False,
                  endpoint: str = "web") -> dict:
    """在 BEGIN IMMEDIATE 事务中原子预占；失败调用可用 release_quota 精确退回。"""
    n = max(0, int(n))
    day = _today()
    subs, limit = _quota_subjects(request)
    marks = ",".join("?" for _ in subs)
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            rows = conn.execute(
                f"SELECT count FROM usage_daily WHERE day=? AND subject IN ({marks})",
                (day, *subs)).fetchall()
            used = max((int(r[0]) for r in rows), default=0)
            available = max(0, limit - used)
            take = min(n, available) if partial else (n if n <= available else 0)
            if take:
                for subject in subs:
                    conn.execute(
                        "INSERT INTO usage_daily(day,subject,count) VALUES(?,?,?) "
                        "ON CONFLICT(day,subject) DO UPDATE SET count=count+excluded.count",
                        (day, subject, take))
                reservation_id = "qr_" + secrets.token_urlsafe(12)
                now = int(time.time())
                conn.execute(
                    "INSERT INTO quota_reservations("
                    "id,day,subjects,units,committed_units,status,endpoint,created,lease_until"
                    ") VALUES(?,?,?,?,0,'pending',?,?,?)",
                    (reservation_id, day, json.dumps(subs, ensure_ascii=False),
                     take, endpoint[:40], now, now + QUOTA_RESERVATION_TTL))
            else:
                reservation_id = ""
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    return {"ok": take == n, "reserved": take, "limit": limit,
            "used_before": used, "used_after": used + take,
            "remaining": max(0, limit - used - take),
            "day": day, "subjects": subs, "id": reservation_id}


def settle_quota(reservation: Optional[dict], committed_units: int) -> None:
    """幂等结算持久化预占：成功次数保留，失败/未处理部分原子退款。"""
    if not reservation:
        return
    reservation_id = reservation.get("id", "")
    if not reservation_id:
        return
    committed_units = max(0, int(committed_units))
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            row = conn.execute(
                "SELECT * FROM quota_reservations WHERE id=? AND status='pending'",
                (reservation_id,)).fetchone()
            if not row:
                conn.rollback()
                return
            units = int(row["units"])
            committed = min(units, committed_units)
            refund = units - committed
            subjects = json.loads(row["subjects"] or "[]")
            for subject in subjects:
                conn.execute(
                    "UPDATE usage_daily SET count=MAX(0,count-?) WHERE day=? AND subject=?",
                    (refund, int(row["day"]), subject))
            status = "settled" if committed else "refunded"
            conn.execute(
                "UPDATE quota_reservations SET committed_units=?,status=?,settled=? "
                "WHERE id=? AND status='pending'",
                (committed, status, int(time.time()), reservation_id))
            conn.commit()
            reservation["reserved"] = committed
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()


def release_quota(reservation: Optional[dict]) -> None:
    settle_quota(reservation, 0)


def _refund_stale_quota_reservations() -> int:
    """进程崩溃后，租约到期的网页额度预占会在后台自动全额退回。"""
    now = int(time.time())
    refunded = 0
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            rows = conn.execute(
                "SELECT * FROM quota_reservations "
                "WHERE status='pending' AND lease_until<?", (now,)).fetchall()
            for row in rows:
                for subject in json.loads(row["subjects"] or "[]"):
                    conn.execute(
                        "UPDATE usage_daily SET count=MAX(0,count-?) "
                        "WHERE day=? AND subject=?",
                        (int(row["units"]), int(row["day"]), subject))
                conn.execute(
                    "UPDATE quota_reservations SET committed_units=0,status='refunded',settled=? "
                    "WHERE id=? AND status='pending'", (now, row["id"]))
                refunded += 1
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    return refunded


def log_request(request: Request, kind: str, link: str, ok: bool):
    try:
        u = current_user(request)
        db_exec("INSERT INTO request_logs(ts,kind,subject,ip,ua,link,ok,path,user_id) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                (int(time.time()), kind,
                 (f"user:{u['id']}" if u else (
                     _stored_fp(request, "request", str(_today()))
                     or _stored_ip(request, "request", str(_today())))),
                 _stored_ip(request, "request", str(_today())),
                 _coarse_ua(request),
                 _log_link(link), 1 if ok else 0, request.url.path,
                 u["id"] if u else None))
    except Exception:
        pass


def log_pageview(request: Request):
    try:
        day = str(_today())
        visitor = (_stored_fp(request, "analytics", day)
                   or _stored_ip(request, "analytics", day))
        db_exec("INSERT INTO page_views(ts,ip,ua,path,fp) VALUES(?,?,?,?,?)",
                (int(time.time()), visitor, "", request.url.path, ""))
    except Exception:
        pass


# ---------------------------------------------------------------- 用户鉴权 / 防机器人

def hash_pw(pw: str, salt: Optional[str] = None):
    salt = salt or secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 120_000).hex()
    return salt, h


def verify_pw(pw: str, salt: str, h: str) -> bool:
    return hmac.compare_digest(
        hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 120_000).hex(), h)


_user_sessions: dict = {}     # token -> (user_id, expiry)
USER_SESSION_TTL = 30 * 86400


def _new_user_session(uid: int) -> str:
    tok = secrets.token_urlsafe(24)
    _user_sessions[tok] = (uid, time.time() + USER_SESSION_TTL)
    return tok


def current_user(request: Request):
    """从 cookie 取当前登录用户（dict）或 None。"""
    tok = request.cookies.get("sess", "")
    ent = _user_sessions.get(tok)
    if not ent or ent[1] < time.time():
        _user_sessions.pop(tok, None)
        return None
    row = db_exec("SELECT * FROM users WHERE id=? AND disabled=0", (ent[0],), "one")
    return dict(row) if row else None


# ---- 滑块验证码（服务端 PNG 缺口 + 行为轨迹 + PoW + 蜜罐 + 一次性签名令牌）----
# 缺口坐标只存在服务端与像素里，绝不出现在返回的标记中——无法靠抓包/解析拿到答案。
_captchas: dict = {}          # cid -> (gap_x, gap_y, issued_at, ip)
CAPTCHA_W, CAPTCHA_H, PIECE = 300, 170, 50
POW_BITS = 14                 # 工作量证明，抬高批量自动化成本
CAPTCHA_SECRET = (os.environ.get("CAPTCHA_SECRET") or "").encode() or APP_SECRET
# 未配置时使用 DATA_DIR/.app-secret；多 worker 也能共享稳定密钥。
_passes: dict = {}            # pass_token -> expiry（一次性）


def _png(width: int, height: int, rows, alpha: bool = False) -> bytes:
    """极简 PNG 编码器（stdlib）。rows 为每行像素字节 bytearray。"""
    import struct
    import zlib
    ct = 6 if alpha else 2                              # RGBA / RGB
    raw = bytearray()
    for r in rows:
        raw.append(0)                                  # filter type 0
        raw += r

    def chunk(typ, data):
        return (struct.pack(">I", len(data)) + typ + data
                + struct.pack(">I", zlib.crc32(typ + data) & 0xFFFFFFFF))

    return (b"\x89PNG\r\n\x1a\n"
            + chunk(b"IHDR", struct.pack(">IIBBBBB", width, height, 8, ct, 0, 0, 0))
            + chunk(b"IDAT", zlib.compress(bytes(raw), 6))
            + chunk(b"IEND", b""))


def _draw_bg(W: int, H: int):
    import colorsys
    h0 = secrets.randbelow(360) / 360
    c1 = tuple(int(v * 255) for v in colorsys.hls_to_rgb(h0, 0.52, 0.55))
    c2 = tuple(int(v * 255) for v in colorsys.hls_to_rgb((h0 + 0.3) % 1, 0.42, 0.55))
    rows = []
    for y in range(H):
        ty = y / H
        row = bytearray(W * 3)
        for x in range(W):
            t = (x / W + ty) * 0.5
            o = x * 3
            row[o] = int(c1[0] * (1 - t) + c2[0] * t)
            row[o + 1] = int(c1[1] * (1 - t) + c2[1] * t)
            row[o + 2] = int(c1[2] * (1 - t) + c2[2] * t)
        rows.append(row)
    for _ in range(5):                                 # 干扰光斑（只遍历包围盒）
        cx, cy, cr = secrets.randbelow(W), secrets.randbelow(H), 12 + secrets.randbelow(22)
        dark = secrets.randbelow(2)
        for y in range(max(0, cy - cr), min(H, cy + cr)):
            base = rows[y]
            for x in range(max(0, cx - cr), min(W, cx + cr)):
                if (x - cx) ** 2 + (y - cy) ** 2 <= cr * cr:
                    o = x * 3
                    for k in range(3):
                        v = base[o + k]
                        base[o + k] = max(0, min(255, v * 7 // 10 if dark else v * 5 // 4))
    return rows


def _in_piece(dx: int, dy: int, PS: int, rad: int) -> bool:
    if dx < rad and dy < rad:
        return (dx - rad) ** 2 + (dy - rad) ** 2 <= rad * rad
    if dx >= PS - rad and dy < rad:
        return (dx - (PS - rad - 1)) ** 2 + (dy - rad) ** 2 <= rad * rad
    if dx < rad and dy >= PS - rad:
        return (dx - rad) ** 2 + (dy - (PS - rad - 1)) ** 2 <= rad * rad
    if dx >= PS - rad and dy >= PS - rad:
        return (dx - (PS - rad - 1)) ** 2 + (dy - (PS - rad - 1)) ** 2 <= rad * rad
    return True


def make_captcha(request: Request) -> dict:
    W, H, PS = CAPTCHA_W, CAPTCHA_H, PIECE
    cid = secrets.token_urlsafe(12)
    gap_x = 90 + secrets.randbelow(W - PS - 110)       # 答案：仅服务端 + 像素
    gap_y = 20 + secrets.randbelow(H - PS - 34)
    _captchas[cid] = (gap_x, gap_y, time.time(), _client_ip(request))
    now = time.time()
    if len(_captchas) > 3000:
        for k, v in list(_captchas.items()):
            if now - v[2] > 300:
                _captchas.pop(k, None)

    bg = _draw_bg(W, H)
    rad = 11
    piece_rows = []
    for dy in range(PS):
        prow = bytearray(PS * 4)
        for dx in range(PS):
            po = dx * 4
            if _in_piece(dx, dy, PS, rad):
                bo = (gap_x + dx) * 3
                srow = bg[gap_y + dy]
                r, g, b = srow[bo], srow[bo + 1], srow[bo + 2]
                edge = dx < 2 or dy < 2 or dx >= PS - 2 or dy >= PS - 2
                if edge:                               # 亮边，拼图更立体
                    prow[po], prow[po + 1], prow[po + 2] = min(255, r + 90), min(255, g + 90), min(255, b + 90)
                else:
                    prow[po], prow[po + 1], prow[po + 2] = r, g, b
                prow[po + 3] = 255
                srow[bo] = r * 4 // 10                 # 挖空处变暗成缺口
                srow[bo + 1] = g * 4 // 10
                srow[bo + 2] = b * 4 // 10
                if edge:
                    srow[bo] = min(255, srow[bo] + 40)
            else:
                prow[po + 3] = 0
        piece_rows.append(prow)

    import base64
    du = lambda p, m="png": f"data:image/{m};base64," + base64.b64encode(p).decode()
    return {"cid": cid, "bg": du(_png(W, H, bg)), "piece": du(_png(PS, PS, piece_rows, alpha=True)),
            "y": gap_y, "w": W, "h": H, "piece_size": PS, "pow_bits": POW_BITS}


def _pow_ok(cid: str, nonce: str) -> bool:
    if not isinstance(nonce, str) or len(nonce) > 40:
        return False
    digest = hashlib.sha256(f"{cid}:{nonce}".encode()).digest()
    return int.from_bytes(digest, "big").bit_length() <= 256 - POW_BITS   # 前 POW_BITS 位为 0


def verify_captcha(cid: str, x, trajectory, nonce: str, request: Request):
    c = _captchas.pop(cid, None)                       # cid 一次性
    if not c:
        return False, "验证已失效，请重新拖动滑块"
    gap_x, gap_y, t0, ip = c
    if _client_ip(request) != ip:
        return False, "环境变化，请重试"                # 绑定签发时的 IP
    if time.time() - t0 > 180:
        return False, "验证超时，请重试"
    if time.time() - t0 < 0.4:
        return False, "操作过快，请手动拖动"            # 秒过 = 脚本
    try:
        x = float(x)
    except Exception:
        return False, "参数错误"
    if abs(x - gap_x) > 6:
        return False, "拼图未对齐，请重试"
    tr = trajectory or []
    if not isinstance(tr, list) or len(tr) < 6:
        return False, "请手动拖动滑块完成验证"
    try:
        ts = [float(p["t"]) for p in tr]
        xs = [float(p["x"]) for p in tr]
    except Exception:
        return False, "轨迹异常"
    dur = ts[-1] - ts[0]
    if dur < 260 or dur > 30000:
        return False, "拖动速度异常，请重试"
    dxs = [xs[i + 1] - xs[i] for i in range(len(xs) - 1)]
    if max((abs(d) for d in dxs), default=0) > gap_x * 0.6:
        return False, "疑似脚本，请手动拖动"            # 一步跳到位
    if len(set(round(d, 1) for d in dxs)) < 4:
        return False, "疑似匀速脚本，请手动拖动"        # 速度无变化 = 线性脚本
    if not _pow_ok(cid, nonce):
        return False, "安全校验失败，请刷新重试"
    return True, None


def issue_pass(request: Request) -> str:
    """滑块通过后签发一次性、限时、绑定 IP 的 HMAC 通行令牌。"""
    exp = int(time.time()) + 120
    body = f"{_client_ip(request)}|{exp}|{secrets.token_urlsafe(9)}"
    sig = hmac.new(CAPTCHA_SECRET, body.encode(), hashlib.sha256).hexdigest()[:20]
    tok = f"{body}|{sig}"
    _passes[tok] = exp
    now = int(time.time())
    if len(_passes) > 5000:
        for k, v in list(_passes.items()):
            if v < now:
                _passes.pop(k, None)
    return tok


def consume_pass(tok: str, request: Request) -> bool:
    """注册/登录时校验并作废通行令牌——一次性、防重放、防伪造、绑定 IP。"""
    if not tok or not isinstance(tok, str):
        return False
    exp = _passes.pop(tok, None)                        # 一次性：用过即废
    if exp is None:
        return False
    try:
        t_ip, t_exp, rnd, sig = tok.split("|")
    except Exception:
        return False
    good = hmac.new(CAPTCHA_SECRET, f"{t_ip}|{t_exp}|{rnd}".encode(),
                    hashlib.sha256).hexdigest()[:20]
    return (hmac.compare_digest(sig, good)
            and t_ip == _client_ip(request)
            and int(t_exp) >= int(time.time()))


# ---- 注册/登录按 IP 限频（防爆破）----
_auth_hits: dict = {}          # ip -> [timestamps]
_captcha_hits: dict = {}       # ip -> [timestamps]（验证码签发限频，防 CPU-DoS）
AUTH_MAX_PER_HOUR = 20
CAPTCHA_MAX_PER_MIN = 40


def _rate_ok(store: dict, ip: str, window: float, cap: int) -> bool:
    now = time.time()
    hits = [t for t in store.get(ip, []) if now - t < window]
    hits.append(now)
    store[ip] = hits
    return len(hits) <= cap


def _auth_rate_ok(ip: str) -> bool:
    return _rate_ok(_auth_hits, ip, 3600, AUTH_MAX_PER_HOUR)


def _captcha_rate_ok(ip: str) -> bool:
    return _rate_ok(_captcha_hits, ip, 60, CAPTCHA_MAX_PER_MIN)


# 管理后台登录防爆破：单 IP 在窗口内失败超限即临时锁定（成功登录清零）。
# 注意与全站一致：只有 TRUST_PROXY=1 时 _client_ip 才采信 XFF，否则按直连 IP 计。
ADMIN_LOGIN_MAX_FAILS = 5
ADMIN_LOGIN_WINDOW = 900          # 15 分钟
_admin_fails: dict = {}           # ip -> [失败时间戳]


def _admin_fail_count(ip: str) -> int:
    now = time.time()
    fails = [t for t in _admin_fails.get(ip, []) if now - t < ADMIN_LOGIN_WINDOW]
    if fails:
        _admin_fails[ip] = fails
    else:
        _admin_fails.pop(ip, None)
    return len(fails)


def _admin_record_fail(ip: str):
    _admin_fails.setdefault(ip, []).append(time.time())


def _sweep_memory():
    """周期清理会话/令牌/限频等内存字典，防止无界增长。"""
    now = time.time()
    for tok, ent in list(_user_sessions.items()):
        if ent[1] < now:
            _user_sessions.pop(tok, None)
    for tok, exp in list(_passes.items()):
        if exp < now:
            _passes.pop(tok, None)
    for cid, v in list(_captchas.items()):
        if now - v[2] > 300:
            _captchas.pop(cid, None)
    for store, win in ((_auth_hits, 3600), (_captcha_hits, 60), (_share_hits, 3600),
                       (_admin_fails, ADMIN_LOGIN_WINDOW)):
        for ip, hits in list(store.items()):
            fresh = [t for t in hits if now - t < win]
            if fresh:
                store[ip] = fresh
            else:
                store.pop(ip, None)
    _sweep_media_limits()


_last_data_cleanup = 0.0


def _cleanup_retained_data(force: bool = False) -> None:
    """清理超过保留期的访问/播放/任务明细；汇总计数与账户余额不受影响。"""
    global _last_data_cleanup
    now = time.time()
    if not force and now - _last_data_cleanup < 300:
        return
    cutoff = int(now) - DATA_RETENTION_DAYS * 86400
    cutoff_day = _today() - DATA_RETENTION_DAYS
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            for table in ("request_logs", "page_views", "share_events", "api_logs"):
                conn.execute(f"DELETE FROM {table} WHERE ts<?", (cutoff,))
            # 投诉中的可选联系方式同样受保留期约束，不能因“未处理”而无限保存。
            conn.execute("DELETE FROM reports WHERE ts<?", (cutoff,))
            # 含今天在内只保留 30 个自然日桶，不能多留第 31 天。
            conn.execute("DELETE FROM usage_daily WHERE day<=?", (cutoff_day,))
            conn.execute(
                "DELETE FROM quota_reservations "
                "WHERE status<>'pending' AND COALESCE(settled,created)<?", (cutoff,))
            conn.execute(
                "DELETE FROM job_items WHERE job_id IN "
                "(SELECT id FROM jobs WHERE finished IS NOT NULL AND finished<?)",
                (cutoff,))
            conn.execute("DELETE FROM jobs WHERE finished IS NOT NULL AND finished<?",
                         (cutoff,))
            conn.execute("DELETE FROM api_ledger WHERE ts<?", (cutoff,))
            conn.execute(
                "DELETE FROM shares WHERE expires_at>0 AND expires_at<?",
                (int(now),))
            # 转发流量是无个人标识的按天聚合，保留 1 年供带宽/成本复盘
            conn.execute("DELETE FROM media_traffic WHERE day<?",
                         (_today() - 366,))
            conn.execute(
                "DELETE FROM api_keys WHERE enabled=0 AND deleted_at IS NOT NULL "
                "AND deleted_at<? AND COALESCE(reserved_cents,0)=0 "
                "AND NOT EXISTS(SELECT 1 FROM jobs WHERE jobs.key=api_keys.key)",
                (cutoff,))
            conn.commit()
            _last_data_cleanup = now
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()


def _sweeper():
    while True:
        time.sleep(300)
        try:
            _sweep_memory()
            _refund_stale_quota_reservations()
            _cleanup_retained_data()
            _flush_media_traffic()
            _atc_cleanup()
        except Exception:
            pass


threading.Thread(target=_sweeper, daemon=True).start()


# 移动端 UA 池（轮换降低指纹一致性）
UA_POOL = [
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_2 like Mac OS X) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/17.2 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 13; PixeI 7 Build/TQ3A.230805.001) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/119.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (Linux; Android 12; SM-G991B) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/118.0.0.0 Mobile Safari/537.36",
]

# 仅允许代理抖音系 CDN，防止服务被当作任意 URL 代理（SSRF）
ALLOWED_HOST_SUFFIXES = (
    "douyinpic.com", "douyinvod.com", "iesdouyin.com", "snssdk.com",
    "douyinstatic.com", "byteimg.com", "ibytedtos.com", "amemv.com",
    "zjcdn.com", "douyincdn.com", "bytecdn.cn", "douyin.com", "pstatp.com",
)

CACHE_TTL = 1800      # 解析结果缓存 30 分钟

# 代理测试目标
TEST_URL_IP = "https://api.ipify.org?format=json"     # 出口 IP
TEST_URL_DOUYIN = "https://www.iesdouyin.com/"        # 抖音可达性

SUPPORTED_SCHEMES = ("http", "https", "socks5", "socks5h", "socks4", "socks4a")

DEFAULT_SETTINGS = {
    "force_proxy": True,          # 无可用代理时拒绝直连（防真实 IP 暴露）
    "default_protocol": "socks5", # 无协议前缀的代理按此协议解析（代理多为 socks5）
    "rotation": "round_robin",    # round_robin | random | least_fail
    "retries": 3,                 # 单个请求最多尝试几个代理后放弃
    "auto_health": True,          # 后台定时健康检查
    "health_interval_min": 10,    # 健康检查间隔（分钟）
    "auto_disable_fail": 5,       # 连续失败达到此数自动禁用（0=不自动禁用）
    "test_reach_douyin": True,    # 测速时附带检测抖音可达
}

_ua_counter = 0


def pick_ua() -> str:
    global _ua_counter
    _ua_counter = (_ua_counter + 1) % len(UA_POOL)
    return UA_POOL[_ua_counter]


# ---------------------------------------------------------------- 代理池管理

class ProxyManager:
    """线程安全的代理池：持久化、轮换、失败计数、统计。"""

    def __init__(self):
        self._lock = threading.Lock()
        self._rr = 0
        self.proxies: list[dict] = []
        self.settings = dict(DEFAULT_SETTINGS)
        self.stats = {"total": 0, "via_proxy": 0, "direct": 0, "retries": 0, "banned": 0}
        self._load()

    # ---- 持久化 ----
    def _load(self):
        if STORE_FILE.exists():
            try:
                d = json.loads(STORE_FILE.read_text("utf-8"))
                self.proxies = d.get("proxies", [])
                self.settings.update(d.get("settings", {}))
            except Exception:
                pass

    def _save(self):
        tmp = STORE_FILE.with_name(STORE_FILE.name + ".tmp")
        tmp.write_text(json.dumps(
            {"proxies": self.proxies, "settings": self.settings},
            ensure_ascii=False, indent=2), "utf-8")
        try:
            tmp.chmod(0o600)
        except OSError:
            pass
        os.replace(tmp, STORE_FILE)

    # ---- 解析：兼容多种代理书写格式 ----
    @staticmethod
    def parse_proxy(raw: str, default_scheme: str = "socks5") -> Optional[str]:
        """把各种常见格式统一成 `scheme://[user:pass@]host:port`。

        支持：
          scheme://user:pass@host:port      scheme://host:port
          user:pass@host:port               host:port
          host:port:user:pass               ip:port（4 段 / 2 段冒号分隔）
          带协议前缀：http/https/socks5/socks5h/socks4/socks4a
        无协议前缀时按 default_scheme（默认 socks5，代理多为 socks5）。
        """
        raw = raw.strip().strip('"\'')
        if not raw:
            return None
        scheme = default_scheme
        m = re.match(r"^(https?|socks5h|socks5|socks4a|socks4)://(.*)$", raw, re.I)
        if m:
            scheme, rest = m.group(1).lower(), m.group(2)
        else:
            rest = raw
        if scheme not in SUPPORTED_SCHEMES:
            return None

        user = pw = None
        if "@" in rest:                                   # user:pass@host:port
            cred, _, hostport = rest.rpartition("@")
            if ":" in cred:
                user, _, pw = cred.partition(":")
            else:
                user = cred
            hp = hostport
        else:
            parts = rest.split(":")
            if len(parts) == 4:                           # host:port:user:pass
                host, port, user, pw = parts
                hp = f"{host}:{port}"
            elif len(parts) == 3:                         # host:port:user
                host, port, user = parts
                hp = f"{host}:{port}"
            else:                                         # host:port
                hp = rest

        hm = re.match(r"^([^:/\s@]+):(\d{1,5})$", hp)
        if not hm or not (0 < int(hm.group(2)) < 65536):
            return None
        host, port = hm.group(1), hm.group(2)
        auth = ""
        if user:
            auth = urlparse.quote(user, safe="")
            if pw:
                auth += ":" + urlparse.quote(pw, safe="")
            auth += "@"
        return f"{scheme}://{auth}{host}:{port}"

    # ---- 增删改 ----
    def add_many(self, raw: str, note: str = "") -> dict:
        default_scheme = self.settings.get("default_protocol", "socks5")
        added, skipped = [], []
        with self._lock:
            existing = {p["url"] for p in self.proxies}
            for line in re.split(r"[\r\n,;]+|\s{2,}", raw.strip()):
                line = line.strip()
                if not line:
                    continue
                url = self.parse_proxy(line, default_scheme)
                if not url:
                    skipped.append(line)
                    continue
                if url in existing:
                    skipped.append(url)
                    continue
                existing.add(url)
                self.proxies.append({
                    "id": secrets.token_hex(4), "url": url, "enabled": True,
                    "auto_off": False, "note": note, "added_at": int(time.time()),
                    "ok": 0, "fail": 0, "last_used": None, "last_ok": None,
                    "latency_ms": None, "exit_ip": None, "douyin_ok": None,
                    "banned": False, "banned_at": None, "banned_reason": None,
                })
                added.append(url)
            self._save()
        return {"added": len(added), "skipped": skipped}

    def remove(self, pid: str) -> bool:
        with self._lock:
            n = len(self.proxies)
            self.proxies = [p for p in self.proxies if p["id"] != pid]
            self._save()
            return len(self.proxies) < n

    def toggle(self, pid: str) -> Optional[bool]:
        with self._lock:
            for p in self.proxies:
                if p["id"] == pid:
                    p["enabled"] = not p["enabled"]
                    p["auto_off"] = False        # 手动操作，取消自动禁用标记
                    if p["enabled"]:
                        p["fail"] = 0
                        p["banned"] = False      # 手动启用即解除封禁标记
                        p["banned_reason"] = None
                    self._save()
                    return p["enabled"]
        return None

    def remove_many(self, ids: set) -> int:
        """批量删除（跳过 managed 托管条目，其生命周期归 mihomo 面板管理）。"""
        with self._lock:
            n = len(self.proxies)
            self.proxies = [p for p in self.proxies
                            if p["id"] not in ids or p.get("managed")]
            self._save()
            return n - len(self.proxies)

    def set_enabled_many(self, ids: set, enabled: bool) -> int:
        """批量启停（跳过 managed 条目），语义与 toggle() 一致。"""
        changed = 0
        with self._lock:
            for p in self.proxies:
                if p["id"] not in ids or p.get("managed"):
                    continue
                p["enabled"] = enabled
                p["auto_off"] = False
                if enabled:
                    p["fail"] = 0
                    p["banned"] = False          # 手动启用即解除封禁标记
                    p["banned_reason"] = None
                changed += 1
            if changed:
                self._save()
        return changed

    def mark_banned(self, p: dict, reason: str):
        """代理 IP 被抖音封禁：落库标记、自动禁用、计数。"""
        with self._lock:
            p["banned"] = True
            p["banned_at"] = int(time.time())
            p["banned_reason"] = reason
            p["enabled"] = False
            p["auto_off"] = True
            p["fail"] = p.get("fail", 0) + 1
            self.stats["banned"] = self.stats.get("banned", 0) + 1
            self._save()

    def get(self, pid: str) -> Optional[dict]:
        return next((p for p in self.proxies if p["id"] == pid), None)

    def set_setting(self, key: str, val):
        with self._lock:
            self.settings[key] = val
            self._save()

    def sync_managed(self, url: Optional[str], enabled: bool,
                     note: str = "内置机场加速（mihomo）"):
        """维护唯一一条「托管」代理条目（内置 mihomo 落地的本地端口）。
        url=None 时移除该条目；否则 upsert 并按 enabled 启停。与用户手动加的代理隔离。"""
        with self._lock:
            m = next((p for p in self.proxies if p.get("managed")), None)
            if url is None:
                if m:
                    self.proxies = [p for p in self.proxies if not p.get("managed")]
                    self._save()
                return
            if m:
                changed = (m["url"] != url) or (m["enabled"] != enabled)
                m["url"], m["note"] = url, note
                m["enabled"] = enabled
                if changed:
                    m["banned"] = False
                    m["banned_reason"] = None
            else:
                self.proxies.append({
                    "id": "mihomo", "url": url, "enabled": enabled, "managed": True,
                    "auto_off": False, "note": note, "added_at": int(time.time()),
                    "ok": 0, "fail": 0, "last_used": None, "last_ok": None,
                    "latency_ms": None, "exit_ip": None, "douyin_ok": None,
                    "banned": False, "banned_at": None, "banned_reason": None,
                })
            self._save()

    # ---- 选择与打点 ----
    @property
    def force_proxy(self) -> bool:
        return bool(self.settings.get("force_proxy", True))

    @property
    def retries(self) -> int:
        return max(1, int(self.settings.get("retries", 3)))

    def candidates(self) -> list[dict]:
        """按轮换策略返回本次请求的代理尝试顺序（仅启用中的）。"""
        with self._lock:
            active = [p for p in self.proxies if p["enabled"]]
            if not active:
                return []
            strategy = self.settings.get("rotation", "round_robin")
            if strategy == "random":
                ordered = active[:]
                random.shuffle(ordered)
            elif strategy == "least_fail":
                ordered = sorted(active, key=lambda p: (p["fail"], -p["ok"]))
            else:                                    # round_robin：轮换起点 + 健康优先
                self._rr = (self._rr + 1) % len(active)
                ordered = active[self._rr:] + active[:self._rr]
                ordered.sort(key=lambda p: p["fail"])
            return ordered

    def _auto_disable_if_needed(self, p: dict):
        thr = int(self.settings.get("auto_disable_fail", 5))
        if thr > 0 and p["enabled"] and p["fail"] >= thr:
            p["enabled"] = False
            p["auto_off"] = True

    def mark_ok(self, p: Optional[dict], latency_ms: Optional[int] = None):
        with self._lock:
            self.stats["total"] += 1
            if p is None:
                self.stats["direct"] += 1
                return
            self.stats["via_proxy"] += 1
            p["ok"] += 1
            p["fail"] = 0
            p["last_used"] = p["last_ok"] = int(time.time())
            if latency_ms is not None:
                p["latency_ms"] = latency_ms
            self._save()

    def mark_fail(self, p: dict):
        with self._lock:
            p["fail"] += 1
            p["last_used"] = int(time.time())
            self._auto_disable_if_needed(p)
            self._save()

    def note_retry(self):
        with self._lock:
            self.stats["retries"] += 1

    def record_probe(self, p: dict, ok: bool, latency_ms=None,
                     exit_ip=None, douyin_ok=None):
        """健康检查/手动测试后回写状态，并处理自动禁用 / 自愈。"""
        with self._lock:
            if ok:
                p["fail"] = 0
                p["last_ok"] = int(time.time())
                if latency_ms is not None:
                    p["latency_ms"] = latency_ms
                if exit_ip is not None:
                    p["exit_ip"] = exit_ip
                if douyin_ok is not None:
                    p["douyin_ok"] = douyin_ok
                if p.get("banned"):              # 封禁的代理测通了 → 解封自愈
                    p["banned"] = False
                    p["banned_reason"] = None
                if p.get("auto_off"):            # 自动禁用过的，恢复可用 → 自愈
                    p["enabled"] = True
                    p["auto_off"] = False
            else:
                p["fail"] += 1
                self._auto_disable_if_needed(p)
            self._save()


proxy_mgr = ProxyManager()


# ---------------------------------------------------------------- 内置 mihomo 内核（机场订阅）
#
# 机场订阅里是 vmess/vless/trojan 等加密协议，本项目出站层（urllib+PySocks）不会解，
# 无法直接进代理池。这里内置一个 mihomo（Clash.Meta 内核）子进程：吃订阅 → 在本地
# 落地成一个 socks5 端口 → 作为一条「托管」代理喂给代理池。多节点测速/切换由 mihomo 负责。
#
# 隔离底线（保证只有本项目能用、绝不影响同服务器其他项目）：
#   · 不设任何系统代理环境变量、不开 TUN/透明代理 → 别的项目联网完全无感
#   · 只绑 127.0.0.1 + allow-lan:false      → 外部机器连不到
#   · 随机高位端口 + 账号密码鉴权 + skip-auth-prefixes 置空（本机也必须带凭证）
#                                           → 同机别的进程即使连到端口也被拒
#   · 子进程只写 data/mihomo/，以本服务同一用户运行，不碰别的项目文件
#
# 默认关闭：只有在后台配置了订阅 URL 后才下载内核并启动。单 worker 运行前提下才安全
# （多 worker 会重复拉起子进程），本项目本就要求单 worker。

MIHOMO_DIR = DATA_DIR / "mihomo"
MIHOMO_BIN = MIHOMO_DIR / "mihomo"
MIHOMO_CFG = MIHOMO_DIR / "config.yaml"
MIHOMO_PID = MIHOMO_DIR / "mihomo.pid"
MIHOMO_LOG = MIHOMO_DIR / "run.log"
MIHOMO_VERSION = os.environ.get("MIHOMO_VERSION", "v1.18.10")
# 国内服务器连不上 github 时，用 MIHOMO_DL_BASE 换镜像（形如 .../releases/download）
MIHOMO_DL_BASE = os.environ.get(
    "MIHOMO_DL_BASE", "https://github.com/MetaCubeX/mihomo/releases/download").rstrip("/")
MIHOMO_OFF = os.environ.get("MIHOMO_OFF", "").lower() in ("1", "true", "yes")


def _mihomo_asset() -> str:
    """按当前平台拼 mihomo release 资源名。"""
    system = platform.system().lower()          # linux / darwin
    machine = platform.machine().lower()
    if machine in ("x86_64", "amd64"):
        arch = "amd64"
    elif machine in ("aarch64", "arm64"):
        arch = "arm64"
    else:
        raise RuntimeError(f"不支持的 CPU 架构：{machine}")
    if system not in ("linux", "darwin"):
        raise RuntimeError(f"不支持的系统：{system}")
    # linux-amd64 用 compatible 变体，兼容不支持 x86-64-v3 指令集的老 CPU
    if system == "linux" and arch == "amd64":
        return f"mihomo-linux-amd64-compatible-{MIHOMO_VERSION}.gz"
    return f"mihomo-{system}-{arch}-{MIHOMO_VERSION}.gz"


_MIHOMO_CFG_TMPL = """\
# 由 server.py 自动生成，请勿手改（改后会被覆盖）。含订阅 token，权限 600。
mixed-port: {port}
bind-address: 127.0.0.1
allow-lan: false
authentication:
  - "{user}:{password}"
skip-auth-prefixes: []
tun:
  enable: false
mode: rule
log-level: warning
# 一个永远连不通的占位节点：保证 auto 组永不为空，从而阻止 mihomo 注入
# COMPATIBLE(=DIRECT) 兜底节点。机场无可用节点时流量落到它 → 直接失败（fail-closed），
# 绝不退回服务器真实 IP 直连——这是本项目「绝不暴露服务器 IP」底线在 mihomo 层的落实。
proxies:
  - name: blackhole
    type: socks5
    server: 127.0.0.1
    port: 1
proxy-providers:
  jichang:
    type: http
    url: "{sub_url}"
    path: ./providers/jichang.yaml
    interval: 3600
    health-check:
      enable: true
      url: https://www.gstatic.com/generate_204
      interval: 300
proxy-groups:
  # fallback：按顺序选第一个「健康」的节点，机场节点全挂时只剩 blackhole → 失败。
  # 不用 url-test：空 provider 的 url-test 会被注入 COMPATIBLE 直连节点而漏 IP。
  - name: auto
    type: fallback
    use: [jichang]
    proxies: [blackhole]
    url: https://www.gstatic.com/generate_204
    interval: 300
rules:
  - MATCH,auto
"""


class MihomoManager:
    """内置 mihomo 子进程的下载、配置、启停与守护。线程安全。"""

    def __init__(self):
        self._lock = threading.Lock()
        self._proc: Optional[subprocess.Popen] = None
        self._state = "stopped"        # stopped/downloading/starting/running/error
        self._last_error = ""
        self._next_try = 0.0           # 失败退避：下次允许启动的时间戳

    # ---- 凭证与本地代理地址（首次生成后持久化）----
    def _creds(self) -> tuple[str, str, str]:
        port = app_setting("mihomo_port")
        user = app_setting("mihomo_user")
        pw = app_setting("mihomo_pass")
        if not (port and user and pw):
            port = str(random.randint(20000, 60000))
            user = "dy" + secrets.token_hex(3)
            pw = secrets.token_hex(8)
            set_app_setting("mihomo_port", port)
            set_app_setting("mihomo_user", user)
            set_app_setting("mihomo_pass", pw)
        return port, user, pw

    def proxy_url(self) -> str:
        port, user, pw = self._creds()
        return f"socks5://{user}:{pw}@127.0.0.1:{port}"

    def sub_url(self) -> str:
        return (app_setting("mihomo_sub_url") or "").strip()

    # ---- 二进制 ----
    def ensure_binary(self):
        if MIHOMO_BIN.exists() and os.access(MIHOMO_BIN, os.X_OK):
            return
        MIHOMO_DIR.mkdir(parents=True, exist_ok=True)
        asset = _mihomo_asset()
        url = f"{MIHOMO_DL_BASE}/{MIHOMO_VERSION}/{asset}"
        gz = MIHOMO_DIR / asset
        self._state = "downloading"
        req = urlreq.Request(url, headers={"User-Agent": "douyin-dl"})
        with urlreq.urlopen(req, timeout=180) as r, open(gz, "wb") as f:
            shutil.copyfileobj(r, f)
        with gzip.open(gz, "rb") as fi, open(MIHOMO_BIN, "wb") as fo:
            shutil.copyfileobj(fi, fo)
        os.chmod(MIHOMO_BIN, 0o755)
        try:
            gz.unlink()
        except OSError:
            pass

    # ---- 配置 ----
    def write_config(self):
        MIHOMO_DIR.mkdir(parents=True, exist_ok=True)
        port, user, pw = self._creds()
        cfg = _MIHOMO_CFG_TMPL.format(port=port, user=user, password=pw,
                                      sub_url=self.sub_url())
        MIHOMO_CFG.write_text(cfg, "utf-8")
        try:
            os.chmod(MIHOMO_CFG, 0o600)          # 含订阅 token
        except OSError:
            pass

    # ---- 进程 ----
    @staticmethod
    def _alive(pid: int) -> bool:
        try:
            os.kill(pid, 0)
            return True
        except OSError:
            return False

    def _kill_stale(self):
        """清理上一次残留的 mihomo（防孤儿/端口占用）。"""
        if not MIHOMO_PID.exists():
            return
        try:
            pid = int(MIHOMO_PID.read_text().strip())
        except (ValueError, OSError):
            MIHOMO_PID.unlink(missing_ok=True)
            return
        if self._proc and self._proc.pid == pid:
            return
        if self._alive(pid):
            try:
                os.kill(pid, 15)
                for _ in range(20):
                    if not self._alive(pid):
                        break
                    time.sleep(0.1)
                if self._alive(pid):
                    os.kill(pid, 9)
            except OSError:
                pass
        MIHOMO_PID.unlink(missing_ok=True)

    def running(self) -> bool:
        return self._proc is not None and self._proc.poll() is None

    def _start_locked(self):
        if not self.sub_url() or self.running():
            return
        self.ensure_binary()
        self.write_config()
        self._kill_stale()
        self._state = "starting"
        logf = open(MIHOMO_LOG, "ab", buffering=0)
        # 用绝对路径：DATA_DIR 可能是相对路径，exec 不受进程 cwd 影响
        self._proc = subprocess.Popen(
            [str(MIHOMO_BIN.resolve()), "-d", str(MIHOMO_DIR.resolve())],
            stdout=logf, stderr=logf,
        )
        MIHOMO_PID.write_text(str(self._proc.pid))
        time.sleep(1.5)
        if self.running():
            self._state = "running"
            self._last_error = ""
        else:
            self._state = "error"
            self._last_error = self._log_tail() or f"mihomo 启动即退出（码 {self._proc.returncode}）"
            self._next_try = time.time() + 30

    def _log_tail(self, n: int = 500) -> str:
        try:
            data = MIHOMO_LOG.read_bytes()[-n:]
            return data.decode("utf-8", "replace").strip()
        except OSError:
            return ""

    def stop(self):
        with self._lock:
            if self._proc and self._proc.poll() is None:
                try:
                    self._proc.terminate()
                    self._proc.wait(timeout=5)
                except (subprocess.TimeoutExpired, OSError):
                    try:
                        self._proc.kill()
                    except OSError:
                        pass
            self._proc = None
            self._kill_stale()
            self._state = "stopped"

    def reload(self):
        """订阅或凭证变更后：有订阅则重建配置并重启，无订阅则停掉。"""
        with self._lock:
            if not self.sub_url():
                self._next_try = 0
        if not self.sub_url():
            self.stop()
            proxy_mgr.sync_managed(None, False)
            return
        self.stop()
        with self._lock:
            self._next_try = 0
            try:
                self._start_locked()
            except Exception as e:            # noqa: BLE001 下载/启动失败不能崩主服务
                self._state = "error"
                self._last_error = str(e)
                self._next_try = time.time() + 30
        proxy_mgr.sync_managed(self.proxy_url(), self.running())

    def _tick(self):
        sub = self.sub_url()
        if not sub:
            if self.running():
                self.stop()
            proxy_mgr.sync_managed(None, False)
            return
        with self._lock:
            if not self.running() and time.time() >= self._next_try:
                try:
                    self._start_locked()
                except Exception as e:        # noqa: BLE001
                    self._state = "error"
                    self._last_error = str(e)
                    self._next_try = time.time() + 30
        proxy_mgr.sync_managed(self.proxy_url(), self.running())

    def supervise(self):
        while True:
            time.sleep(5)
            if MIHOMO_OFF:
                continue
            try:
                self._tick()
            except Exception as e:            # noqa: BLE001 守护线程绝不能挂
                self._last_error = str(e)

    def status(self) -> dict:
        sub = self.sub_url()
        managed = next((p for p in proxy_mgr.proxies if p.get("managed")), None)
        return {
            "enabled": bool(sub),
            "sub_url_masked": _mask_secret(sub) if sub else "",
            "state": self._state,
            "running": self.running(),
            "binary_ready": MIHOMO_BIN.exists() and os.access(MIHOMO_BIN, os.X_OK),
            "version": MIHOMO_VERSION,
            "last_error": self._last_error[-300:],
            "exit_ip": managed.get("exit_ip") if managed else None,
            "douyin_ok": managed.get("douyin_ok") if managed else None,
            "latency_ms": managed.get("latency_ms") if managed else None,
        }


def _mask_secret(s: str) -> str:
    """打码：保留头尾，中间星号（用于订阅 URL/token 回显）。"""
    if len(s) <= 12:
        return s[:2] + "***" + s[-2:] if len(s) > 4 else "***"
    return s[:18] + "***" + s[-6:]


def _proxy_public_label(proxy: Optional[dict]) -> str:
    """公开错误中只显示协议和节点，不回显 userinfo/密码。"""
    if not proxy:
        return "direct"
    try:
        p = urlparse.urlsplit(proxy.get("url", ""))
        host = p.hostname or "unknown"
        if ":" in host and not host.startswith("["):
            host = f"[{host}]"
        return f"{p.scheme or 'proxy'}://{host}{':' + str(p.port) if p.port else ''}"
    except Exception:
        return "proxy"


def _redact_proxy_error(value) -> str:
    text = str(value or "")
    # 异常库有时会把完整代理 URL 带回来，先替换已知配置，再兜底清理任意 URL userinfo。
    for proxy in getattr(proxy_mgr, "proxies", []):
        raw = proxy.get("url", "")
        if raw:
            text = text.replace(raw, _proxy_public_label(proxy))
    return re.sub(r"([a-zA-Z][\w+.-]*://)[^/@\s]+@", r"\1***@", text)[:180]


mihomo_mgr = MihomoManager()


# ---------------------------------------------------------------- 应用设置 + 开放 API 计费

NEW_KEY_BALANCE = int(os.environ.get("NEW_KEY_BALANCE", "100"))   # 新 Key 试用余额（分）


def app_setting(key: str, default: str = "") -> str:
    row = db_exec("SELECT v FROM app_settings WHERE k=?", (key,), "one")
    return row["v"] if row else default


def set_app_setting(key: str, val) -> None:
    db_exec("INSERT INTO app_settings(k,v) VALUES(?,?) "
            "ON CONFLICT(k) DO UPDATE SET v=?", (key, str(val), str(val)))


def api_price_cents() -> int:
    try:
        return max(0, int(app_setting("api_price_cents", "1")))    # 默认 1 分/次
    except Exception:
        return 1


def create_api_key(user_id: Optional[int], name: str) -> dict:
    key = "dy_" + secrets.token_urlsafe(24)
    now = int(time.time())
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            conn.execute(
                "INSERT INTO api_keys("
                "key,user_id,name,created,enabled,balance_cents,spent_cents,calls,reserved_cents"
                ") VALUES(?,?,?,?,1,?,0,0,0)",
                (key, user_id, (name or "未命名")[:60], now, NEW_KEY_BALANCE))
            conn.execute(
                "INSERT INTO api_ledger(ts,key,event,balance_delta,reason) "
                "VALUES(?,?,?,?,?)",
                (now, key, "opening", NEW_KEY_BALANCE, "new_key_balance"))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    return get_api_key(key)


def get_api_key(key: str) -> Optional[dict]:
    row = db_exec("SELECT * FROM api_keys WHERE key=?", (key,), "one")
    return dict(row) if row else None


def list_api_keys(user_id: Optional[int] = None) -> list:
    if user_id is None:
        rows = db_exec("SELECT * FROM api_keys ORDER BY created DESC", (), "all")
    else:
        rows = db_exec(
            "SELECT * FROM api_keys WHERE user_id=? AND enabled=1 "
            "AND deleted_at IS NULL ORDER BY created DESC",
            (user_id,), "all")
    return [dict(r) for r in rows]


def revoke_api_key(key: str, user_id: Optional[int] = None) -> bool:
    k = get_api_key(key)
    if not k or (user_id is not None and k["user_id"] != user_id):
        return False
    db_exec("UPDATE api_keys SET enabled=0,deleted_at=? WHERE key=?",
            (int(time.time()), key))
    return True


def recharge_key(key: str, cents: int) -> bool:
    if not get_api_key(key):
        return False
    cents = int(cents)
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            n = conn.execute(
                "UPDATE api_keys SET balance_cents=balance_cents+? WHERE key=?",
                (cents, key)).rowcount
            if not n:
                conn.rollback()
                return False
            conn.execute(
                "INSERT INTO api_ledger(ts,key,event,balance_delta,reason) "
                "VALUES(?,?,?,?,?)",
                (int(time.time()), key, "recharge", cents, "admin_recharge"))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    return True


def api_key_check(key: str):
    """校验 key（不扣费）。返回 (rec, error)。"""
    if not key:
        return None, "缺少 API Key（请通过 X-API-Key 请求头传入）"
    rec = get_api_key(key)
    if not rec or not rec["enabled"] or rec.get("deleted_at"):
        return None, "无效或已禁用的 API Key"
    return rec, None


# ---------------------------------------------------------------- HTTP 出站层

class NoRedirect(urlreq.HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):
        return None


def _proxy_handler(proxy: dict):
    """根据代理 URL 构造 urllib handler。socks5 走 PySocks。"""
    url = proxy["url"]
    scheme = url.split("://", 1)[0].lower()
    if scheme in ("http", "https"):
        return urlreq.ProxyHandler({"http": url, "https": url})
    if scheme.startswith("socks"):
        import socks
        from sockshandler import SocksiPyHandler
        parts = urlparse.urlsplit(url)
        stype = socks.SOCKS4 if scheme.startswith("socks4") else socks.SOCKS5
        rdns = scheme in ("socks5h", "socks4a")     # 远端解析 DNS，避免 DNS 泄露
        user = urlparse.unquote(parts.username) if parts.username else None
        pw = urlparse.unquote(parts.password) if parts.password else None
        return SocksiPyHandler(stype, parts.hostname, parts.port or 1080,
                               rdns=rdns, username=user, password=pw)
    raise ValueError(f"不支持的代理协议: {scheme}")


def _raw_open(url: str, follow: bool, headers: dict, timeout: int, proxy: Optional[dict]):
    handlers = []
    if proxy:
        handlers.append(_proxy_handler(proxy))
    if not follow:
        handlers.append(NoRedirect())
    opener = urlreq.build_opener(*handlers)
    req = urlreq.Request(url, headers=headers)
    try:
        return opener.open(req, timeout=timeout)
    except urlerr.HTTPError as e:
        if not follow and e.code in (301, 302, 303, 307, 308):
            return e          # 重定向对短链解析而言是"成功"
        raise


def open_url(url: str, follow: bool = True, headers: Optional[dict] = None,
             timeout: int = 30, retry_http_statuses: tuple = (),
             ban_on_auth_error: bool = True):
    """出站请求核心：一律经代理，失败自动转移。

    所有到抖音的服务器请求都走这里 —— **绝不服务器直连**，避免暴露服务器 IP。
    仅当管理后台关闭「禁止直连」(force_proxy=False) 且无可用代理时，才退回直连。
    返回 (response, proxy_used_or_None)。
    """
    hdrs = {"User-Agent": pick_ua()}
    if headers:
        hdrs.update(headers)

    cands = proxy_mgr.candidates()
    if not cands:                                   # 无可用代理
        if proxy_mgr.force_proxy:
            raise ApiError(503, "没有可用代理，且已开启「禁止服务器直连」——为避免暴露服务器 IP，"
                                "不会直连抖音。请在管理后台添加并启用代理。")
        r = _raw_open(url, follow, hdrs, timeout, None)   # 仅在管理员显式允许时直连
        proxy_mgr.mark_ok(None)
        return r, None

    cands = cands[:proxy_mgr.retries]               # 每请求最多尝试 N 个代理
    errors = []
    for i, p in enumerate(cands):
        if i > 0:
            proxy_mgr.note_retry()                  # 记录一次自动重试（换代理）
        t0 = time.time()
        try:
            r = _raw_open(url, follow, hdrs, timeout, p)
            proxy_mgr.mark_ok(p, int((time.time() - t0) * 1000))
            return r, p
        except urlerr.HTTPError as e:
            if e.code in (403, 401):                # 抖音封禁该代理 IP → 落库+禁用+换代理
                if ban_on_auth_error:
                    proxy_mgr.mark_banned(p, f"抖音返回 {e.code}，IP 被封禁")
                    errors.append(f"{_proxy_public_label(p)} → 被封禁(HTTP {e.code})")
                else:
                    # 媒体 CDN 的 401/403 也可能只针对当前资源/域名，不能据此永久封禁出口。
                    errors.append(
                        f"{_proxy_public_label(p)} → 媒体请求 HTTP {e.code}")
                try:
                    e.close()
                except Exception:
                    pass
                continue
            if e.code in retry_http_statuses:
                # 代理网关也会生成 5xx；媒体请求应换出口验证，不能把故障代理记为健康。
                # 同时不能累计连接失败：源站 429/5xx 可能只针对当前资源，避免毒死代理池。
                errors.append(f"{_proxy_public_label(p)} → HTTP {e.code}")
                try:
                    e.close()
                except Exception:
                    pass
                continue
            # 其他 4xx/5xx 是源站问题，不怪代理，直接上抛
            proxy_mgr.mark_ok(p, int((time.time() - t0) * 1000))
            raise
        except Exception as e:                      # 连接/超时 → 代理故障，自动转移
            msg = str(e).lower()
            if "403" in msg or "forbidden" in msg or "tunnel connection failed" in msg:
                if ban_on_auth_error:
                    proxy_mgr.mark_banned(p, "代理无法连接抖音（403/被封禁）")
                    errors.append(f"{_proxy_public_label(p)} → 被封禁(403)")
                else:
                    # HTTP CONNECT/tunnel 失败未必代表出口 IP 被抖音永久封禁；
                    # 媒体 CDN 线路只轮换本次请求，不能污染代理池持久健康状态。
                    errors.append(
                        f"{_proxy_public_label(p)} → 媒体连接失败(403)")
            else:
                proxy_mgr.mark_fail(p)
                errors.append(
                    f"{_proxy_public_label(p)} → {type(e).__name__}: "
                    f"{_redact_proxy_error(e)}")

    # 所有代理都连不通
    if proxy_mgr.force_proxy:
        raise ApiError(502, "全部代理均不可用，且已禁止服务器直连抖音。"
                            "请在管理后台检查代理状态。")
    r = _raw_open(url, follow, hdrs, timeout, None)   # 仅在管理员显式允许时直连
    proxy_mgr.mark_ok(None)
    return r, None


# ---------------------------------------------------------------- 工具函数

app = FastAPI(title="抖音无水印下载器", version=APP_VERSION)


@app.middleware("http")
async def _private_api_responses(request: Request, call_next):
    """API 响应可能含签名地址、账号或密钥，禁止浏览器与共享代理持久化。"""
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        response.headers["Cache-Control"] = "private, no-store"
        response.headers["Pragma"] = "no-cache"
    return response


class ApiError(Exception):
    def __init__(self, status: int, message: str, headers: Optional[dict] = None):
        self.status, self.message = status, message
        self.headers = headers or {}


@app.exception_handler(ApiError)
async def _api_error(_: Request, exc: ApiError):
    return JSONResponse(status_code=exc.status, content={"error": exc.message},
                        headers=exc.headers)


def _host_allowed(url: str) -> bool:
    if not isinstance(url, str) or len(url) > 4096:
        return False
    try:
        p = urlparse.urlsplit(url)
        if p.scheme not in ("http", "https") or p.username or p.password:
            return False
        if p.port not in (None, 80, 443):
            return False
        host = (p.hostname or "").lower().rstrip(".")
    except (ValueError, TypeError):
        return False
    return any(host == s or host.endswith("." + s) for s in ALLOWED_HOST_SUFFIXES)


def _find_key(obj, key):
    if isinstance(obj, dict):
        if key in obj:
            yield obj[key]
        for v in obj.values():
            yield from _find_key(v, key)
    elif isinstance(obj, list):
        for v in obj:
            yield from _find_key(v, key)


def _safe_name(desc: str, fallback: str) -> str:
    name = re.sub(r"#\S+", "", desc).strip()
    name = re.sub(r'[\\/:*?"<>|\s]+', "_", name).strip("_")
    return (name or fallback)[:60]


def _media_signature(kind: str, resource: str, exp: int) -> str:
    payload = f"media:v1\n{kind}\n{resource}\n{int(exp)}".encode()
    return hmac.new(APP_SECRET, payload, hashlib.sha256).hexdigest()


def _media_token(kind: str, resource: str, ttl: int = MEDIA_TOKEN_TTL) -> tuple:
    exp = int(time.time()) + max(60, int(ttl))
    return exp, _media_signature(kind, resource, exp)


def _require_media_token(kind: str, resource: str, exp: int, sig: str) -> None:
    now = int(time.time())
    if (not re.fullmatch(r"[0-9a-f]{64}", sig or "")
            or exp < now or exp > now + MEDIA_TOKEN_TTL + 300):
        raise ApiError(403, "媒体链接已过期，请重新解析或刷新页面")
    expected = _media_signature(kind, resource, exp)
    if not hmac.compare_digest(sig, expected):
        raise ApiError(403, "媒体链接签名无效")


def _video_proxy_url(vid: str) -> str:
    exp, sig = _media_token("video", vid)
    return f"/api/video/{vid}?" + urlparse.urlencode({"exp": exp, "sig": sig})


def _stream(resp, chunk=256 * 1024, on_close=None):
    try:
        while True:
            block = resp.read(chunk)
            if not block:
                break
            yield block
    finally:
        if on_close:
            on_close()
        else:
            resp.close()


def _parse_content_range(value: str):
    """解析单段 Content-Range，返回 (start, end, total|None)。"""
    match = re.fullmatch(
        r"bytes\s+(\d+)-(\d+)/(\d+|\*)", (value or "").strip(),
        flags=re.IGNORECASE)
    if not match:
        return None
    start, end = int(match.group(1)), int(match.group(2))
    total = None if match.group(3) == "*" else int(match.group(3))
    if end < start or (total is not None and (total <= 0 or end >= total)):
        return None
    return start, end, total


def _video_response_shape(resp, requested_range: str):
    """校验上游 Range 语义，返回流的绝对边界与期望字节数。

    返回 (status, start, end|None, total|None, expected|None)。缺少长度时
    expected 为 None，此时仍可流式传输，但无法判定提前 EOF。
    """
    status = resp.status if hasattr(resp, "status") else resp.getcode()
    content_range = _parse_content_range(
        resp.headers.get("Content-Range") or "")
    raw_length = (resp.headers.get("Content-Length") or "").strip()
    if raw_length and not re.fullmatch(r"\d+", raw_length):
        raise ValueError("invalid video Content-Length")
    content_length = int(raw_length) if raw_length else None

    if requested_range:
        if status != 206 or content_range is None:
            raise ValueError("video upstream did not honor Range")
        start, end, total = content_range
        expected = end - start + 1
        if content_length is not None and content_length != expected:
            raise ValueError("video range length mismatch")

        match = re.fullmatch(r"bytes=(\d*)-(\d*)", requested_range.strip())
        if not match:
            raise ValueError("invalid requested video Range")
        req_start, req_end = match.groups()
        if req_start:
            if start != int(req_start):
                raise ValueError("video range start mismatch")
            if req_end:
                wanted_end = int(req_end)
                if total is not None:
                    wanted_end = min(wanted_end, total - 1)
                if end != wanted_end:
                    raise ValueError("video range end mismatch")
            elif total is not None and end != total - 1:
                raise ValueError("open video range ended early")
        else:
            # 后缀 Range 必须知道资源总长，才能证明返回的是最后 N 字节。
            suffix = int(req_end)
            if total is None or end != total - 1:
                raise ValueError("invalid suffix video range")
            if start != max(0, total - suffix):
                raise ValueError("video suffix range mismatch")
        return status, start, end, total, expected

    if status != 200:
        raise ValueError("unexpected partial response for full video")
    if content_range is not None:
        raise ValueError("unexpected Content-Range for full video")
    if content_length is None:
        return status, 0, None, None, None
    return status, 0, content_length - 1, content_length, content_length


class _ResumeBudgetExceeded(Exception):
    pass


class _ResumableVideoStream:
    """在上游长连接提前 EOF/读取异常后，从精确字节偏移续传。"""
    _MAX_CONSECUTIVE_RESUME_FAILURES = MEDIA_RESUME_MAX_FAILURES
    _MAX_TOTAL_RESUME_ATTEMPTS = MEDIA_RESUME_MAX_ATTEMPTS
    _MAX_RESUME_SECONDS = MEDIA_RESUME_MAX_SECONDS

    def __init__(self, vid: str, initial, request_headers: dict,
                 start: int, end: Optional[int], total: Optional[int],
                 expected: Optional[int], chunk: int = 256 * 1024):
        self.vid = vid
        self.request_headers = dict(request_headers)
        self.start = start
        self.end = end
        self.total = total
        self.expected = expected
        self.chunk = chunk
        self.sent = 0
        self.current = initial
        self.closed = False
        self._lock = threading.Lock()
        self._on_close = None
        self._resume_attempts = 0
        self._resume_started = None

    def set_on_close(self, callback) -> None:
        self._on_close = callback

    def _take_current(self):
        with self._lock:
            return None if self.closed else self.current

    def _replace_current(self, replacement) -> bool:
        with self._lock:
            if self.closed:
                accepted = False
            else:
                self.current = replacement
                accepted = True
        if not accepted:
            _close_upstream(replacement)
        return accepted

    def _resume(self):
        with self._lock:
            if self.closed:
                return None
            now = time.monotonic()
            if self._resume_started is None:
                self._resume_started = now
            if (self._resume_attempts >= self._MAX_TOTAL_RESUME_ATTEMPTS
                    or now - self._resume_started
                    >= self._MAX_RESUME_SECONDS):
                raise _ResumeBudgetExceeded()
            self._resume_attempts += 1
        offset = self.start + self.sent
        if self.end is not None and offset > self.end:
            return None
        headers = dict(self.request_headers)
        headers["Range"] = (
            f"bytes={offset}-{self.end}"
            if self.end is not None else f"bytes={offset}-")

        def validate(candidate):
            status, start, end, total, expected = _video_response_shape(
                candidate, headers["Range"])
            if status != 206 or start != offset:
                raise ValueError("resumed video range start mismatch")
            if self.end is not None and end is not None and end > self.end:
                raise ValueError("resumed video range exceeded response")
            if self.total is not None and total != self.total:
                raise ValueError("video size changed while resuming")
            if expected is not None and self.expected is not None:
                remaining = self.expected - self.sent
                if expected > remaining:
                    raise ValueError("resumed video range is too long")

        replacement = _open_video_upstream(
            self.vid, headers, validator=validate)
        if (time.monotonic() - self._resume_started
                >= self._MAX_RESUME_SECONDS):
            _close_upstream(replacement)
            raise _ResumeBudgetExceeded()
        if not self._replace_current(replacement):
            return None
        return replacement

    def close(self) -> None:
        with self._lock:
            if self.closed:
                return
            self.closed = True
            current, self.current = self.current, None
        if current is not None:
            _close_upstream(current)

    def __iter__(self):
        consecutive_failures = 0
        try:
            while self.expected is None or self.sent < self.expected:
                current = self._take_current()
                if current is None:
                    break
                remaining = (
                    self.chunk if self.expected is None
                    else min(self.chunk, self.expected - self.sent))
                try:
                    block = current.read(remaining)
                except Exception as exc:
                    # http.client.IncompleteRead 等异常会携带已收到的 partial；
                    # 必须先转发并推进 offset，否则重开 Range 会重复这些字节。
                    partial = getattr(exc, "partial", b"")
                    block = bytes(partial) if isinstance(
                        partial, (bytes, bytearray, memoryview)) else b""
                    if self.expected is not None:
                        block = block[:self.expected - self.sent]
                    if block:
                        self.sent += len(block)
                        consecutive_failures = 0
                        yield block
                    if self.expected is None or self.sent >= self.expected:
                        break
                    _close_upstream(current)
                    consecutive_failures += 1
                else:
                    if block:
                        if self.expected is not None:
                            block = block[:self.expected - self.sent]
                        self.sent += len(block)
                        consecutive_failures = 0
                        yield block
                        continue
                    if self.expected is None or self.sent >= self.expected:
                        break
                    # 已声明长度却提前 EOF：关闭断流连接，从 sent 对应的
                    # 绝对偏移重开单段 Range，避免重复或缺失字节。
                    _close_upstream(current)
                    consecutive_failures += 1

                if (consecutive_failures
                        > self._MAX_CONSECUTIVE_RESUME_FAILURES):
                    raise OSError(
                        "video upstream repeatedly ended early") from None
                while True:
                    try:
                        resumed = self._resume()
                        break
                    except _ResumeBudgetExceeded:
                        raise OSError(
                            "video upstream resume budget exhausted") from None
                    except Exception:
                        consecutive_failures += 1
                        if (consecutive_failures
                                > self._MAX_CONSECUTIVE_RESUME_FAILURES):
                            raise OSError(
                                "video upstream resume failed") from None
                if resumed is None:
                    break
        finally:
            callback = self._on_close
            if callback:
                callback()
            else:
                self.close()


_media_limit_lock = threading.Lock()
_media_hits: dict = {}
_media_active: dict = {}


class _MediaLease:
    def __init__(self, key: str):
        self.key = key
        self.released = False


def _media_lease(request: Request) -> _MediaLease:
    """为一次媒体流申请 IP 级请求/并发租约；仅在流关闭时释放并发计数。"""
    key = _client_ip(request)
    now = time.time()
    with _media_limit_lock:
        hits = [t for t in _media_hits.get(key, []) if now - t < 60]
        if len(hits) >= MEDIA_REQUESTS_PER_MIN:
            raise ApiError(429, "媒体请求过于频繁，请稍后再试",
                           {"Retry-After": "60"})
        if _media_active.get(key, 0) >= MEDIA_MAX_CONCURRENT:
            raise ApiError(429, "同时播放或下载的媒体过多，请稍后再试",
                           {"Retry-After": "2"})
        hits.append(now)
        _media_hits[key] = hits
        _media_active[key] = _media_active.get(key, 0) + 1
    return _MediaLease(key)


def _media_release(lease: _MediaLease) -> None:
    with _media_limit_lock:
        if lease.released:
            return
        lease.released = True
        active = _media_active.get(lease.key, 0) - 1
        if active > 0:
            _media_active[lease.key] = active
        else:
            _media_active.pop(lease.key, None)


class _MediaStreamingResponse(StreamingResponse):
    """无论 ASGI 在响应头、首块或流中何处中断，都释放媒体并发租约。"""
    def __init__(self, *args, finalize, **kwargs):
        self._media_finalize = finalize
        super().__init__(*args, **kwargs)

    async def __call__(self, scope, receive, send):
        try:
            await super().__call__(scope, receive, send)
        finally:
            self._media_finalize()


def _media_finalizer(upstream, lease: _MediaLease, on_close=None):
    """返回线程安全、幂等的上游关闭 + 并发租约释放函数。

    on_close 在首次 finalize 时执行一次（用于流量计量等收尾统计），
    必须是廉价的内存操作——finalize 可能跑在事件循环线程，阻塞不得。"""
    lock = threading.Lock()
    closed = False

    def finalize():
        nonlocal closed
        with lock:
            if closed:
                return
            closed = True
        try:
            upstream.close()
        except Exception:
            pass
        finally:
            _media_release(lease)
            if on_close:
                try:
                    on_close()
                except Exception:
                    pass

    return finalize


def _sweep_media_limits() -> None:
    now = time.time()
    with _media_limit_lock:
        for key, hits in list(_media_hits.items()):
            fresh = [t for t in hits if now - t < 60]
            if fresh:
                _media_hits[key] = fresh
            else:
                _media_hits.pop(key, None)


# 媒体转发流量统计（后台「转发流量统计」）：只统计经 /api/video 同源转发的字节，
# 浏览器直连抖音 CDN 的流量不经过本服务器、无法也不需要统计。
# 内存累加 + 定期落库（media_traffic 按天/用途聚合，无任何个人标识）——
# 不能在流结束回调里直接写 SQLite：finalize 可能跑在事件循环线程。
_traffic_lock = threading.Lock()
_traffic_pending: dict = {}          # (day, scope) -> [requests, bytes]


def _traffic_add(scope: str, nbytes: int) -> None:
    key = (_today(), scope)
    with _traffic_lock:
        cur = _traffic_pending.setdefault(key, [0, 0])
        cur[0] += 1
        cur[1] += max(0, int(nbytes or 0))


def _flush_media_traffic() -> None:
    global _traffic_pending
    with _traffic_lock:
        if not _traffic_pending:
            return
        pending, _traffic_pending = _traffic_pending, {}
    try:
        with _db_lock:
            conn = _db()
            try:
                conn.execute("BEGIN IMMEDIATE")
                for (day, scope), (n, b) in pending.items():
                    conn.execute(
                        "INSERT INTO media_traffic(day,scope,requests,bytes) "
                        "VALUES(?,?,?,?) ON CONFLICT(day,scope) DO UPDATE SET "
                        "requests=requests+excluded.requests, "
                        "bytes=bytes+excluded.bytes",
                        (day, scope, n, b))
                conn.commit()
            except Exception:
                conn.rollback()
                raise
            finally:
                conn.close()
    except Exception:
        # 落库失败把计数放回内存，等下一轮 sweeper 重试，不丢数据
        with _traffic_lock:
            for key, (n, b) in pending.items():
                cur = _traffic_pending.setdefault(key, [0, 0])
                cur[0] += n
                cur[1] += b


def _valid_single_range(value: str) -> bool:
    """仅接受单段 bytes Range，拒绝多段请求放大与异常长请求头。"""
    if not value:
        return True
    value = value.strip()
    if len(value) > 100:
        return False
    match = re.fullmatch(r"bytes=(\d*)-(\d*)", value)
    if not match or not any(match.groups()):
        return False
    start, end = match.groups()
    return not ((start and end and int(start) > int(end))
                or (not start and int(end) <= 0))


def _content_disposition(name: str) -> str:
    safe = re.sub(r'[\\/:*?"<>|]+', "_", name)[:80]
    return f"attachment; filename*=UTF-8''{urlparse.quote(safe)}"


# ---------------------------------------------------------------- 核心解析

_cache: dict = {}
_author_cache: dict = {}          # item_id -> (ts, 作者结构化详情)
CDN_HEADERS = {
    "Referer": "https://www.douyin.com/",
    "Accept-Encoding": "identity",
}


def _play_api(vid: str) -> str:
    """无水印播放接口地址。交给用户浏览器直接请求：

    浏览器 GET 该地址 → 302 → 跟随到 CDN 直链（按浏览器自身 IP/地区解析）→ 播放。
    普通浏览器优先用它，让视频字节不经过本服务器（省带宽、不暴露服务器 IP），
    且 CDN 直链与浏览器同 IP，避免"服务器/代理 IP 解析的直链换个 IP 打不开"。
    实测该接口对桌面 UA / 无 UA 均返回 200，浏览器可直连。
    """
    return f"https://aweme.snssdk.com/aweme/v1/play/?video_id={vid}&ratio=1080p&line=0"


def _play_api_alt(vid: str) -> str:
    """备用播放域名。与 aweme.snssdk.com 互为备份（见 docs/产品文档.md §风险表）。

    微信内不同机型/内核对这两个域名的可达性不一致（部分环境 snssdk 被拦、
    iesdouyin 可播，反之亦然）。所有环境（含微信）都在服务器代理前尝试
    此线路以节省带宽，同源代理只做兜底。
    """
    return f"https://www.iesdouyin.com/aweme/v1/play/?video_id={vid}&ratio=1080p&line=0"


def _video_download_url(vid: str, filename: str = "video.mp4") -> str:
    """同源下载地址：经 Range 预检后由浏览器原生流式保存。"""
    exp, sig = _media_token("video", vid)
    return f"/api/video/{vid}?" + urlparse.urlencode({
        "exp": exp,
        "sig": sig,
        "dl": "1",
        "name": filename or "video.mp4",
    })


def _card_cover(cover: str) -> str:
    """把抖音封面直链转成"适合当社交卡片图"的形式：**去签名 + 转 JPEG**。

    抖音给的封面是 `https://p26-sign.douyinpic.com/...webp?x-expires=...&x-signature=...`，
    当 og:image 有两个硬伤：① `.webp` 微信卡片缩略图支持不稳定；② 签名 ~14 天过期，
    过期后存量分享页全变无图卡片。

    实测（见 README 更新日志 v1.7.0）：把主机的 `-sign` 去掉、扩展名换成 `.jpeg`，
    抖音会返回 **无签名、不过期的 JPEG**（同一张图，体积略大）。签名覆盖了路径，
    所以只换扩展名不去 -sign 主机会 403，两步必须一起做。

    只认白名单内的抖音图床，转换失败就原样返回（宁可用 webp，也不要吐出个坏链接）。
    """
    if not cover or not cover.startswith("https://"):
        return cover
    try:
        if not _host_allowed(cover):
            return cover
        p = urlparse.urlsplit(cover)
        host, path = p.netloc, p.path
        if "-sign." not in host or not path.lower().endswith((".webp", ".jpeg", ".jpg")):
            return cover
        host = host.replace("-sign.", ".", 1)
        path = re.sub(r"\.webp$", ".jpeg", path, flags=re.I)
        return f"https://{host}{path}"        # 丢掉 query（签名参数），无签名主机不需要
    except Exception:
        return cover


def _parse_share(text: str) -> dict:
    m = re.search(r"https://v\.douyin\.com/[\w-]+/?", text)
    if not m:
        raise ApiError(400, "未找到抖音分享链接，请确认文案里包含 v.douyin.com 短链")
    short = m.group(0)

    try:
        resp, _ = open_url(short, follow=False)
    except ApiError:
        raise
    except Exception:
        raise ApiError(502, "短链请求失败，请检查网络/代理后重试")
    try:
        location = resp.headers.get("Location", "") if hasattr(resp, "headers") else ""
    finally:
        try:
            resp.close()
        except Exception:
            pass
    km = re.search(r"/share/(video|note|slides)/(\d+)", location)
    if not km:
        if "/share/live" in location:
            raise ApiError(400, "这是直播分享链接，暂不支持下载直播内容")
        raise ApiError(404, "链接已失效或指向不支持的内容类型")
    kind, item_id = km.group(1), km.group(2)
    if kind == "slides":
        kind = "note"
    return _parse_item(kind, item_id)


def _parse_item(kind: str, item_id: str) -> dict:
    """抓分享页并提取元数据。与短链解析分开，便于分享页按 item_id 刷新过期直链。"""
    try:
        page, used_proxy = open_url(f"https://www.iesdouyin.com/share/{kind}/{item_id}/")
        try:
            html = page.read().decode("utf-8", "ignore")
        finally:
            page.close()
    except ApiError:
        raise
    except Exception:
        raise ApiError(502, "分享页请求失败，请稍后重试")

    dm = re.search(r"window\._ROUTER_DATA\s*=\s*(\{.*?\})\s*</script>", html, re.S)
    if not dm:
        # 分享页正常必有 _ROUTER_DATA；没有 = 被返回验证/风控页。
        # 若是经代理请求，判定该代理 IP 被封禁：落库、禁用、上抛。
        if used_proxy:
            proxy_mgr.mark_banned(used_proxy, "分享页返回验证/无数据，IP 被风控封禁")
            raise ApiError(502, "代理 IP 被抖音风控（返回验证页），已自动封禁并禁用该代理，请重试")
        raise ApiError(502, "分享页无数据，可能被风控或页面结构变更，请稍后重试")
    data = json.loads(dm.group(1))

    items = next((i for i in _find_key(data, "item_list") if i), None)
    if not items:
        raise ApiError(404, "视频不存在、已被删除，或作者设为私密/仅粉丝可见")
    item = items[0]

    desc = item.get("desc", "") or ""
    au = item.get("author") or {}
    author = au.get("nickname") or next(_find_key(au, "nickname"), "") or ""
    avatar_list = next(_find_key(au, "url_list"), None) or []
    avatar = avatar_list[0] if avatar_list else ""
    base = _safe_name(desc, item_id)

    # 作者结构化详情：缓存到服务端，供前端悬停 2s 拉取做浮层
    sec_uid = au.get("sec_uid") or ""
    homepage = f"https://www.douyin.com/user/{sec_uid}" if sec_uid else ""
    author_detail = {
        "nickname": author,
        "avatar": avatar,
        "sec_uid": sec_uid,
        "douyin_id": au.get("unique_id") or au.get("short_id") or "",
        "signature": (au.get("signature") or "").strip(),
        "aweme_count": au.get("aweme_count"),
        "following_count": au.get("following_count"),
        "follower_count": au.get("mplatform_followers_count") or au.get("follower_count"),
        "total_favorited": au.get("total_favorited"),      # 获赞总数（分享页多为空，浮层时富化）
        "homepage": homepage,
        "enriched": False,
    }
    _author_cache[item_id] = (time.time(), author_detail)

    # 作品互动数据（点赞/评论/收藏/分享）—— 分享页直接给，无需额外请求
    st = item.get("statistics") or {}
    stats = {
        "digg": st.get("digg_count"),        # 点赞
        "comment": st.get("comment_count"),  # 评论
        "collect": st.get("collect_count"),  # 收藏
        "share": st.get("share_count"),      # 分享
    }

    # 更多可直接读取的元数据
    tags = [t.get("hashtag_name") for t in (item.get("text_extra") or []) if t.get("hashtag_name")]
    mu = item.get("music") or {}
    music = {"title": mu.get("title"), "author": mu.get("author")} if mu.get("title") else None
    poi = item.get("aweme_poi_info") or {}
    location = poi.get("poi_name") or (item.get("anchor_info") or {}).get("name") or None

    # 缩略图（封面/头像）直接给 CDN 直链，由浏览器直连加载
    result = {
        "kind": kind, "item_id": item_id, "title": desc or "（无标题）",
        "author": author,
        "avatar": avatar,
        "author_url": homepage,
        "create_time": item.get("create_time"),
        "stats": stats,
        "tags": tags,
        "music": music,
        "location": location,
        "base": base,
    }

    if kind == "note":
        images = item.get("images") or []
        if not images:
            raise ApiError(404, "图集中未找到图片")
        urls = [img["url_list"][0] for img in images if img.get("url_list")]
        # 直链交给浏览器直接查看 / 下载
        result["images"] = [{"url": u, "filename": f"{base}_{i:02d}.jpeg"}
                            for i, u in enumerate(urls, 1)]
        return result

    video = item.get("video") or {}
    play = next(_find_key(video.get("play_addr") or {}, "url_list"), None) or []
    if not play:
        raise ApiError(404, "未找到播放地址")
    vm = re.search(r"video_id=([\w-]+)", play[0])
    if not vm:
        raise ApiError(502, "播放地址格式已变更，无法提取 video_id")
    vid = vm.group(1)
    cover_list = next(_find_key(video.get("cover") or {}, "url_list"), None) or []

    result.update({
        "duration_ms": video.get("duration") or 0,
        "cover": cover_list[0] if cover_list else "",
        "video": {
            "url": _play_api(vid),                    # 浏览器直连播放（自行跟随 302）
            "alt_url": _play_api_alt(vid),            # 备用抖音域名（线路顺序由前端按运行环境决定）
            "proxy_url": _video_proxy_url(vid),       # 同源签名流：微信优先，普通浏览器兜底
            "filename": f"{base}.mp4",
            "download_url": _video_download_url(vid, f"{base}.mp4"),
            "width": video.get("width"),
            "height": video.get("height"),
        },
    })
    return result


def _parse_cached(text: str) -> dict:
    key = text.strip()
    now = time.time()
    hit = _cache.get(key)
    if hit and now - hit[0] < CACHE_TTL:
        return hit[1]
    data = _parse_share(text)
    _cache[key] = (now, data)
    _cache[data["item_id"]] = (now, data)
    if len(_cache) > 500:
        for k, (ts, _) in list(_cache.items()):
            if now - ts > CACHE_TTL:
                _cache.pop(k, None)
    return data


# ---------------------------------------------------------------- 分享页
#
# 目标：抖音链接发到微信打不开 —— 生成一个「微信里点开就能看」的作品页。
# 原则（与全站一致）：只存元数据快照 + video_id，**不落地任何媒体字节**。
#   · 视频：每次渲染用 _play_api(vid) 重拼播放地址（该地址无签名/无时效，天然长期有效）
#   · 图集：CDN 直链会过期 → 超过 SHARE_REFRESH_TTL 或前端上报失败时按 item_id 惰性重解析
#   · 源作品被删 → 刷新失败 → status='dead'，页面展示"已被原作者删除"并保留署名

SHARE_TTL_ANON = int(os.environ.get("SHARE_TTL_ANON_DAYS", "7")) * 86400
SHARE_TTL_USER = int(os.environ.get("SHARE_TTL_USER_DAYS", "30")) * 86400
SHARE_REFRESH_TTL = 12 * 3600          # 图集直链超过这个时长就在下次访问时刷新
SHARE_MAX_PER_HOUR = 30                # 匿名创建限频（每 IP）
_share_hits: dict = {}

# ---- 分享域名池 ----
# 微信封"下载/侵权类"域名是常态而非意外，因此分享链接与主站域名物理隔离，并可轮换。
# 短码与域名解耦：同一个 sid 在任意域名下都能打开，某域名被封时切换即可救活存量分享。
# 配置：SHARE_DOMAINS="https://s1.example.com,https://s2.example.com"
SHARE_DOMAINS = [d.strip().rstrip("/") for d in
                 os.environ.get("SHARE_DOMAINS", "").split(",") if d.strip()]
_share_dom_rr = 0


def _domains_off() -> set:
    """被管理员标记为"已被封"的域名，暂时不再分配给新链接。"""
    try:
        return set(json.loads(app_setting("share_domains_off", "[]")))
    except Exception:
        return set()


def _share_origin(request: Request) -> str:
    """给**新生成的分享链接**分配域名。
    优先级：后台配置的主分享域名（app_settings.share_primary_domain）→ SHARE_DOMAINS
    域名池（轮换）→ 当前请求来源。主域名让所有新链接固定落在同一个"微信可打开"的域名上，
    无需改环境变量、后台即时生效；被标记封禁后自动退回域名池/请求来源。"""
    global _share_dom_rr
    primary = app_setting("share_primary_domain", "").strip().rstrip("/")
    if primary and primary not in _domains_off():
        return primary
    pool = [d for d in SHARE_DOMAINS if d not in _domains_off()]
    if not pool:
        return _origin(request)
    _share_dom_rr = (_share_dom_rr + 1) % len(pool)
    return pool[_share_dom_rr]

# 短码字母表：去掉 0/O/1/l/I 等易混字符
_SID_ALPHABET = "23456789abcdefghijkmnpqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ"


def _new_sid(n: int = 7) -> str:
    for _ in range(6):
        sid = "".join(secrets.choice(_SID_ALPHABET) for _ in range(n))
        if not db_exec("SELECT id FROM shares WHERE id=?", (sid,), "one"):
            return sid
    return "".join(secrets.choice(_SID_ALPHABET) for _ in range(n + 3))


def _is_wechat(request: Request) -> bool:
    return "micromessenger" in (request.headers.get("user-agent") or "").lower()


def _share_ttl(request: Request) -> int:
    return SHARE_TTL_USER if current_user(request) else SHARE_TTL_ANON


def _share_state(row: dict) -> str:
    """ok | expired | dead | takedown —— 决定页面展示哪种状态。"""
    if row["status"] in ("dead", "takedown"):
        return row["status"]
    if row["expires_at"] and row["expires_at"] < time.time():
        return "expired"
    return "ok"


def _refresh_share(row: dict) -> dict:
    """图集直链过期时按 item_id 重新解析。失败则标记 dead（源多半已被删）。"""
    try:
        data = _parse_item(row["kind"], row["item_id"])
    except Exception:
        db_exec("UPDATE shares SET status='dead', refreshed_at=? WHERE id=?",
                (int(time.time()), row["id"]))
        row["status"] = "dead"
        return row
    vid = row.get("vid") or ""
    if row["kind"] != "note":
        play_url = ((data.get("video") or {}).get("url") or "")
        match = re.search(r"[?&]video_id=([\w-]+)", play_url)
        if match:
            vid = match.group(1)
    db_exec("UPDATE shares SET payload=?, cover=?, vid=?, refreshed_at=?, status='ok' "
            "WHERE id=?",
            (json.dumps(data, ensure_ascii=False), data.get("cover", ""), vid,
             int(time.time()), row["id"]))
    row["payload"] = json.dumps(data, ensure_ascii=False)
    row["cover"] = data.get("cover", "")
    row["vid"] = vid
    row["status"] = "ok"
    return row


def _share_view(row: dict, origin: str = "") -> dict:
    """把 shares 行转成分享页要用的数据结构（含重拼后的播放地址）。"""
    data = json.loads(row["payload"] or "{}")
    cfg = _atc_cfg()                                   # 一次读取，避免重复查库
    if row["kind"] != "note" and row["vid"]:
        data.setdefault("video", {})
        data["video"]["url"] = _play_api(row["vid"])          # 每次重拼，保持新鲜
        data["video"]["alt_url"] = _play_api_alt(row["vid"])   # 备用抖音域名
        data["video"]["proxy_url"] = _video_proxy_url(row["vid"])
        filename = data["video"].get("filename") or (
            _safe_name(row["title"] or "", row["item_id"]) + ".mp4")
        data["video"]["filename"] = filename
        data["video"]["download_url"] = _video_download_url(row["vid"], filename)
        # ATC 增强线路：地址新鲜才注入；过期则后台惰性重新入队，本次走其他线路
        if cfg["enabled"] and cfg["play_enhance"]:
            cached = _atc_cache_get(row["item_id"])
            if _atc_url_fresh(cached, cfg["url_ttl"]):
                data["video"]["atc_url"] = cached["video_url"]
            elif cached:
                _atc_enqueue(row["item_id"], purpose="play")
    # 播放线路优先级（后台可拖拽排序；atc 无地址时前端自动跳过）
    data["play_priority"] = cfg["play_priority"]
    return {
        "sid": row["id"],
        "kind": row["kind"],
        "item_id": row["item_id"],
        "title": row["custom_title"] or row["title"] or "（无标题）",
        "author": row["author"] or "",
        "avatar": row["avatar"] or "",
        "cover": row["cover"] or "",
        # 社交分享用的封面（无签名 JPEG、不过期）——JS-SDK 卡片图与海报都用它
        "card_cover": _card_cover(row["cover"] or ""),
        "created": row["created"],
        "expires_at": row["expires_at"],
        "state": _share_state(row),
        "url": f"{origin}/s/{row['id']}" if origin else f"/s/{row['id']}",
        "views": row["views"], "plays": row["plays"], "downloads": row["downloads"],
        "data": data,
    }


def _share_create(request: Request, data: dict, custom_title: str = "") -> dict:
    u = current_user(request)
    sid = _new_sid()
    now = int(time.time())
    vid = ""
    if data.get("video", {}).get("url"):
        vm = re.search(r"video_id=([\w-]+)", data["video"]["url"])
        vid = vm.group(1) if vm else ""
    db_exec(
        "INSERT INTO shares(id,item_id,kind,vid,owner_user_id,owner_fp,owner_ip,"
        "title,author,avatar,cover,payload,custom_title,visibility,expires_at,"
        "refreshed_at,status,created) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (sid, data.get("item_id", ""), data.get("kind", "video"), vid,
         u["id"] if u else None, "", "",
         (data.get("title") or "")[:300], (data.get("author") or "")[:100],
         data.get("avatar", ""), data.get("cover", ""),
         json.dumps(data, ensure_ascii=False), (custom_title or "")[:300],
         "link", now + _share_ttl(request), now, "ok", now))
    row = dict(db_exec("SELECT * FROM shares WHERE id=?", (sid,), "one"))
    # ATC 播放地址增强：后台异步入队（分钟级），不阻塞分享页返回
    if data.get("kind") != "note" and data.get("item_id"):
        try:
            _atc_enqueue(data["item_id"], purpose="play")
        except Exception:
            pass
    return _share_view(row, _share_origin(request))      # 新链接按域名池分配


def _share_event(request: Request, sid: str, kind: str, source: str = "",
                 stage: str = "", detail: str = "", ms: int = 0,
                 next_src: str = ""):
    """记录分享页埋点。播放类事件额外带 source/stage/detail/ms/next_src，用于诊断
    「微信里哪些视频能播、走的哪条线路、失败在哪一步、失败后接着重试哪条」。
    注意：只记线路名，不记带签名的完整媒体地址（隐私红线）。"""
    col = {"view": "views", "play": "plays",
           "download": "downloads", "cta": "cta_clicks"}.get(kind)
    try:
        if col:
            db_exec(f"UPDATE shares SET {col}={col}+1 WHERE id=?", (sid,))
        db_exec("INSERT INTO share_events(ts,sid,kind,ip,ua,referer,wechat,fp,"
                "source,stage,detail,ms,next_src) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (int(time.time()), sid, kind, "", _coarse_ua(request), "",
                 1 if _is_wechat(request) else 0, "",
                 source[:24], stage[:24], detail[:120], int(ms or 0),
                 next_src[:24]))
    except Exception:
        pass


class ShareBody(BaseModel):
    text: str = ""
    item_id: str = ""
    title: str = ""


@app.post("/api/share")
def api_share_create(body: ShareBody, request: Request):
    """生成分享页。已解析过的作品命中缓存 → 不重复解析、不再扣配额。"""
    if not _rate_ok(_share_hits, _client_ip(request), 3600, SHARE_MAX_PER_HOUR):
        raise ApiError(429, "创建分享页过于频繁，请稍后再试")

    data = None
    if body.item_id:
        hit = _cache.get(body.item_id.strip())
        if hit and time.time() - hit[0] < CACHE_TTL:
            data = hit[1]
        else:
            # 缓存已过期：本站已有该作品的分享页时可直接复用快照，仍然零解析成本
            row = db_exec("SELECT * FROM shares WHERE item_id=? AND status='ok' "
                          "ORDER BY created DESC LIMIT 1", (body.item_id.strip(),), "one")
            if row:
                data = json.loads(dict(row)["payload"] or "{}")
    if data is None:
        if not body.text.strip():
            raise ApiError(400, "解析结果已过期，请重新粘贴链接后再生成分享页")
        reservation = reserve_quota(request, 1, endpoint="share_parse")
        if not reservation["ok"]:
            raise _quota_error(reservation["limit"])
        try:
            data = _parse_cached(body.text)
        except Exception:
            release_quota(reservation)
            raise
        settle_quota(reservation, 1)
    if not data.get("item_id"):
        raise ApiError(400, "解析数据不完整，无法生成分享页")
    return _share_create(request, data, body.title)


@app.get("/api/share/{sid}")
def api_share_get(sid: str, request: Request):
    row = db_exec("SELECT * FROM shares WHERE id=?", (sid,), "one")
    if not row:
        raise ApiError(404, "分享页不存在或已被删除")
    return _share_view(dict(row), _origin(request))


class ShareEventBody(BaseModel):
    kind: str
    source: str = ""            # 播放线路：dy1(aweme.snssdk) / dy2(iesdouyin) / atc(增强) / proxy(服务器兜底)
    stage: str = ""             # 该线路的结果：start / ok / error / timeout / giveup
    detail: str = ""            # 失败细节（media error code、readyState 等）
    ms: int = 0                 # 从该线路开始到出结果的耗时
    next: str = ""              # 失败后链上下一条将重试的线路名（无则空 = 已是最后一条）


# 播放诊断事件：play_try/play_ok/play_fail 只写 share_events，不累加 shares 计数，
# 避免把「尝试次数」混进 plays（plays 仍只由 play 事件累加，代表一次成功起播）。
SHARE_EVENT_KINDS = ("view", "play", "download", "cta", "fallback",
                     "play_try", "play_ok", "play_fail")


@app.post("/api/share/{sid}/event")
def api_share_event(sid: str, body: ShareEventBody, request: Request):
    if body.kind in SHARE_EVENT_KINDS:
        _share_event(request, sid, body.kind, body.source, body.stage,
                     body.detail, body.ms, body.next)
    return {"ok": True}


def _qr_bytes(sid: str, request: Request, kind: str, scale: int):
    try:
        import segno
    except ImportError:
        raise ApiError(501, "服务器未安装二维码依赖 segno")
    buf = io.BytesIO()
    segno.make(f"{_share_origin(request)}/s/{sid}", error="m").save(
        buf, kind=kind, scale=scale, border=2, dark="#111418", light="#ffffff")
    return buf.getvalue()


@app.get("/s/{sid}/qr.svg")
def share_qr(sid: str, request: Request):
    return Response(_qr_bytes(sid, request, "svg", 6), media_type="image/svg+xml",
                    headers={"Cache-Control": "public, max-age=86400"})


@app.get("/s/{sid}/qr.png")
def share_qr_png(sid: str, request: Request):
    """海报合成用：同源 PNG，画进 canvas 不会污染画布（SVG 在部分浏览器会）。"""
    return Response(_qr_bytes(sid, request, "png", 8), media_type="image/png",
                    headers={"Cache-Control": "public, max-age=86400"})


@app.get("/api/shares")
def api_my_shares(request: Request, limit: int = 50):
    u = current_user(request)
    if not u:
        raise ApiError(401, "请先登录后查看我的分享")
    rows = db_exec("SELECT * FROM shares WHERE owner_user_id=? ORDER BY created DESC LIMIT ?",
                   (u["id"], max(1, min(200, limit))), "all")
    origin = _share_origin(request)     # 复制出去的链接始终用当前可用域名
    return {"shares": [_share_view(dict(r), origin) for r in rows]}


@app.delete("/api/shares/{sid}")
def api_share_delete(sid: str, request: Request):
    u = current_user(request)
    if not u:
        raise ApiError(401, "请先登录")
    n = db_exec("DELETE FROM shares WHERE id=? AND owner_user_id=?", (sid, u["id"]), "rowcount")
    if not n:
        raise ApiError(404, "分享页不存在或不属于你")
    return {"ok": True}


class ReportBody(BaseModel):
    sid: str
    reason: str
    contact: str = ""


@app.post("/api/report")
def api_report(body: ReportBody, request: Request):
    """侵权/违规投诉入口（无需登录）。管理员在后台处理后可下架。"""
    if not body.reason.strip():
        raise ApiError(400, "请填写投诉理由")
    db_exec("INSERT INTO reports(ts,sid,reason,contact,ip) VALUES(?,?,?,?,?)",
            (int(time.time()), body.sid[:32], body.reason[:1000],
             body.contact[:200], ""))
    return {"ok": True, "message": "已收到，我们会尽快处理"}


# ---- 微信 JS-SDK 分享卡片签名 ----
#
# 裸页面在微信里的分享卡片由微信自行抓取，样式朴素；接入 JS-SDK 才能精确控制
# 标题/描述/缩略图。需要「已认证服务号 + 已备案域名」，在后台填 AppID/AppSecret 启用。
# jsapi_ticket 全局唯一、7200s 有效且有调用频次上限 → **必须存 app_settings 表**，
# 存内存会导致多 worker 各自刷新互相顶掉。

def _wx_api(url: str) -> dict:
    """请求微信开放接口。**刻意不走代理池**——公众号要求服务器出口 IP 在白名单内，
    走代理会因 IP 不匹配而失败；且这里请求的是微信而非抖音，不涉及被抖音封的问题。"""
    req = urlreq.Request(url, headers={"User-Agent": "douyin-dl/1.0"})
    with urlreq.urlopen(req, timeout=10) as r:
        return json.loads(r.read().decode("utf-8", "ignore"))


def _wx_ticket():
    """返回 (appid, jsapi_ticket)；未配置返回 (None, None)。带 app_settings 级缓存。"""
    appid = app_setting("wx_appid").strip()
    secret = app_setting("wx_secret").strip()
    if not (appid and secret):
        return None, None
    now = time.time()
    cached = app_setting("wx_ticket")
    try:
        exp = float(app_setting("wx_ticket_exp") or 0)
    except ValueError:
        exp = 0
    if cached and exp > now + 60:
        return appid, cached
    tok = _wx_api("https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential"
                  f"&appid={urlparse.quote(appid)}&secret={urlparse.quote(secret)}")
    if not tok.get("access_token"):
        raise ApiError(502, f"获取微信 access_token 失败：{tok.get('errmsg') or tok}")
    tk = _wx_api("https://api.weixin.qq.com/cgi-bin/ticket/getticket?type=jsapi"
                 f"&access_token={urlparse.quote(tok['access_token'])}")
    if not tk.get("ticket"):
        raise ApiError(502, f"获取 jsapi_ticket 失败：{tk.get('errmsg') or tk}")
    set_app_setting("wx_ticket", tk["ticket"])
    set_app_setting("wx_ticket_exp", now + int(tk.get("expires_in", 7200)) - 300)
    return appid, tk["ticket"]


@app.get("/api/wx/jssdk")
def wx_jssdk(request: Request, url: str = ""):
    """给分享页签名。未配置公众号时返回 enabled=false，前端静默降级。"""
    # 先校验 URL 归属，再取 ticket：避免外站请求也能触发微信接口调用（有频次上限）
    # 含后台主分享域名：反代头缺失（如 Cloudflare flexible SSL）导致 _origin 取错时仍能签名
    primary = app_setting("share_primary_domain", "").strip().rstrip("/")
    allowed = list(SHARE_DOMAINS) + [_origin(request)] + ([primary] if primary else [])
    page = (url or "").split("#")[0]
    if not any(page.startswith(a) for a in allowed):
        raise ApiError(403, "该 URL 不属于本站，拒绝签名")
    try:
        appid, ticket = _wx_ticket()
    except ApiError as e:
        return {"enabled": False, "error": e.message}
    if not appid:
        return {"enabled": False}
    nonce = secrets.token_hex(8)
    ts = int(time.time())
    raw = (f"jsapi_ticket={ticket}&noncestr={nonce}&timestamp={ts}&url={page}")
    return {"enabled": True, "appId": appid, "timestamp": ts, "nonceStr": nonce,
            "signature": hashlib.sha1(raw.encode()).hexdigest()}


# ---------------------------------------------------------------- AnyToCopy 增强线路（ATC）
#
# 第三方增值服务（可选，默认关闭）：语音转文字文案提取 + 分享页增强播放地址。
# 硬约束：
#   · 永不进入同步主解析（/api/parse 不感知本分区）；ATC 是异步任务（提交→轮询，分钟级）
#   · 总开关 atc_enabled=0 时整条线路完全静默，任何主流程不受影响
#   · ATC 并发上限 5 → 全部请求经 atc_jobs 表串行排队，用户请求绝不直连 ATC
#   · 只存 URL 与文案元数据，不落地媒体字节；缓存随 DATA_RETENTION_DAYS 清理
#   · 出站刻意不走代理池（同微信 JS-SDK 先例：第三方 API 要求出口稳定，且非抖音无封 IP 风险）

ATC_DEFAULT_BASE = "https://api.anytocopy.com/vip/open-api/v1"
ATC_POLL_INTERVAL = 4          # 官方建议 3-5 秒
ATC_JOB_TIMEOUT = 300          # 单任务最长 5 分钟
ATC_INFLIGHT_MAX = 2           # 同时在轮询的任务数（对方并发上限 5，留余量给其网页端）
ATC_WORK_URL = "https://www.douyin.com/video/{item_id}"   # 由 item_id 还原作品链接
SHARE_PLAY_SOURCES = ("dy1", "dy2", "atc", "proxy")


def _atc_cfg() -> dict:
    """读取运行时配置（app_settings，后台改即时生效）。未启用/未配密钥 → enabled=False。"""
    key = app_setting("atc_api_key").strip()
    secret = app_setting("atc_api_secret").strip()
    try:
        daily = max(0, min(100, int(app_setting("atc_transcript_daily", "5"))))
    except ValueError:
        daily = 5
    try:
        # 实测 ATC 地址约 1 小时后 403（2026-08-07 实测），默认 1 小时并提前换线
        ttl = max(600, min(86400, int(app_setting("atc_url_ttl", "3600"))))
    except ValueError:
        ttl = 3600
    try:
        priority = json.loads(app_setting("share_play_priority", ""))
        if not (isinstance(priority, list)
                and sorted(priority) == sorted(SHARE_PLAY_SOURCES)):
            raise ValueError
    except (ValueError, TypeError):
        priority = list(SHARE_PLAY_SOURCES)
    return {
        "enabled": app_setting("atc_enabled") == "1" and bool(key and secret),
        "key": key, "secret": secret,
        "base": (app_setting("atc_base_url") or ATC_DEFAULT_BASE).rstrip("/"),
        "play_enhance": app_setting("atc_play_enhance", "1") == "1",
        "transcript_enabled": app_setting("atc_transcript_enabled", "1") == "1",
        "transcript_daily": daily,
        "url_ttl": ttl,
        "play_priority": priority,
    }


def _atc_request(method: str, path: str, params: dict, cfg: dict) -> dict:
    """调 ATC 开放 API。返回解析后的 JSON；网络/协议错误抛异常。"""
    url = cfg["base"] + path + "?" + urlparse.urlencode(params)
    req = urlreq.Request(url, method=method)
    req.add_header("X-API-Key", cfg["key"])
    req.add_header("X-API-Secret", cfg["secret"])
    req.add_header("User-Agent", pick_ua())
    with urlreq.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8", "replace"))


def _atc_cache_get(item_id: str) -> Optional[dict]:
    row = db_exec("SELECT * FROM atc_cache WHERE item_id=?", (item_id,), "one")
    return dict(row) if row else None


def _atc_url_fresh(row: Optional[dict], ttl: int) -> bool:
    """缓存里的 API 播放地址是否仍在有效期内（签名链接会过期）。"""
    return bool(row and row.get("video_url") and row.get("url_fetched_at")
                and time.time() - row["url_fetched_at"] < ttl)


def _atc_enqueue(item_id: str, work_url: str = "", purpose: str = "play") -> bool:
    """入队一个 ATC 任务。幂等：同 item_id 有在途任务或缓存仍新鲜 → 不再入队。"""
    cfg = _atc_cfg()
    if not cfg["enabled"] or not item_id:
        return False
    if purpose == "play" and not cfg["play_enhance"]:
        return False
    if purpose == "transcript" and not cfg["transcript_enabled"]:
        return False
    cached = _atc_cache_get(item_id)
    if purpose == "play" and _atc_url_fresh(cached, cfg["url_ttl"]):
        return False
    if purpose == "transcript" and cached and cached.get("text_content"):
        return False
    if db_exec("SELECT id FROM atc_jobs WHERE item_id=? AND status IN ('pending','submitted')",
               (item_id,), "one"):
        return False
    # 防任务空转：近期已跑完一轮但仍没有新鲜地址（对方也取不到）→ 冷却期内不再入队
    cooldown = db_exec(
        "SELECT updated FROM atc_jobs WHERE item_id=? AND status IN ('done','failed') "
        "ORDER BY updated DESC LIMIT 1", (item_id,), "one")
    if cooldown and time.time() - cooldown[0] < cfg["url_ttl"]:
        return False
    now = int(time.time())
    db_exec("INSERT INTO atc_jobs(item_id,work_url,purpose,status,created,updated) "
            "VALUES(?,?,?,'pending',?,?)",
            (item_id, (work_url or ATC_WORK_URL.format(item_id=item_id))[:300],
             purpose, now, now))
    return True


def _atc_save_result(item_id: str, data: dict) -> None:
    """任务成功：upsert 缓存。只覆盖返回里实际带值的字段（转录任务不该清掉旧播放地址）。"""
    now = int(time.time())
    old = _atc_cache_get(item_id) or {}
    video_url = data.get("videoUrl") or old.get("video_url") or ""
    fetched = now if data.get("videoUrl") else (old.get("url_fetched_at") or 0)
    db_exec(
        "INSERT INTO atc_cache(item_id,video_url,url_fetched_at,content,text_content,"
        "audio_url,duration,created,updated) VALUES(?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT(item_id) DO UPDATE SET video_url=?,url_fetched_at=?,content=?,"
        "text_content=?,audio_url=?,duration=?,updated=?",
        (item_id, video_url, fetched,
         (data.get("content") or old.get("content") or "")[:2000],
         (data.get("textContent") or old.get("text_content") or ""),
         (data.get("audioUrl") or old.get("audio_url") or ""),
         data.get("duration") or old.get("duration"),
         old.get("created") or now, now,
         video_url, fetched,
         (data.get("content") or old.get("content") or "")[:2000],
         (data.get("textContent") or old.get("text_content") or ""),
         (data.get("audioUrl") or old.get("audio_url") or ""),
         data.get("duration") or old.get("duration"), now))


def _atc_worker():
    """守护线程（5s 一轮）：提交 pending 任务、轮询 submitted 任务、写缓存。
    同时在途任务不超过 ATC_INFLIGHT_MAX；重启后 submitted 任务凭 task_id 直接续查。"""
    while True:
        time.sleep(5)
        try:
            cfg = _atc_cfg()
            if not cfg["enabled"]:
                continue
            now = int(time.time())
            inflight = db_exec(
                "SELECT COUNT(*) FROM atc_jobs WHERE status='submitted'", (), "one")[0]
            if inflight < ATC_INFLIGHT_MAX:
                job = db_exec(
                    "SELECT * FROM atc_jobs WHERE status='pending' ORDER BY id LIMIT 1",
                    (), "one")
                if job:
                    job = dict(job)
                    try:
                        resp = _atc_request("POST", "/video/extract",
                                            {"workUrl": job["work_url"],
                                             "taskType": "TEXT"}, cfg)
                        if resp.get("code") == 200 and resp.get("data"):
                            db_exec("UPDATE atc_jobs SET status='submitted',task_id=?,"
                                    "updated=? WHERE id=?", (str(resp["data"]), now, job["id"]))
                        elif "并发" in str(resp.get("msg") or ""):
                            # 对方并发已满：保持 pending 等下一轮，不判死
                            db_exec("UPDATE atc_jobs SET updated=?,error=? WHERE id=?",
                                    (now, "对方并发已满，排队重试中", job["id"]))
                        else:
                            raise RuntimeError(str(resp.get("msg") or resp)[:200])
                    except Exception as e:
                        db_exec("UPDATE atc_jobs SET status='failed',error=?,updated=? "
                                "WHERE id=?",
                                (f"提交失败: {type(e).__name__}: {e}"[:300], now, job["id"]))
                    continue
            job = db_exec(
                "SELECT * FROM atc_jobs WHERE status='submitted' ORDER BY id LIMIT 1",
                (), "one")
            if not job:
                continue
            job = dict(job)
            if now - (job["updated"] or now) < ATC_POLL_INTERVAL:
                continue                       # 距上次轮询不足 4 秒
            if now - job["created"] > ATC_JOB_TIMEOUT:
                db_exec("UPDATE atc_jobs SET status='failed',error='轮询超时',updated=? "
                        "WHERE id=?", (now, job["id"]))
                continue
            try:
                resp = _atc_request("GET", "/video/query", {"taskId": job["task_id"]}, cfg)
                data = resp.get("data") or {}
                status = data.get("status", "")
                if status == "SUCCESS":
                    _atc_save_result(job["item_id"], data)
                    db_exec("UPDATE atc_jobs SET status='done',updated=? WHERE id=?",
                            (now, job["id"]))
                elif status in ("FAILED", "FAILURE"):
                    db_exec("UPDATE atc_jobs SET status='failed',error=?,updated=? WHERE id=?",
                            ((data.get("errorMessage") or "任务失败")[:300], now, job["id"]))
                else:
                    db_exec("UPDATE atc_jobs SET updated=? WHERE id=?", (now, job["id"]))
            except Exception as e:
                # 单次轮询网络错误不判死，只刷新时间戳；超时由上面的 created 判定兜底
                db_exec("UPDATE atc_jobs SET updated=?,error=? WHERE id=?",
                        (now, f"轮询异常: {type(e).__name__}"[:200], job["id"]))
        except Exception:
            pass


def _atc_cleanup() -> int:
    """缓存按保留期清理；终态任务记录保留 7 天。由 _sweeper 调用。"""
    now = int(time.time())
    n = db_exec("DELETE FROM atc_cache WHERE updated<?",
                (now - DATA_RETENTION_DAYS * 86400,), "rowcount") or 0
    n += db_exec("DELETE FROM atc_jobs WHERE status IN ('done','failed') AND updated<?",
                 (now - 7 * 86400,), "rowcount") or 0
    return n


def _atc_status() -> dict:
    """后台状态面板数据（不含密钥本体）。"""
    cfg = _atc_cfg()
    today0 = _today() * 86400
    row = db_exec(
        "SELECT COUNT(*) AS total, "
        "SUM(CASE WHEN status='done' THEN 1 ELSE 0 END) AS done, "
        "SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) AS failed "
        "FROM atc_jobs WHERE created>=?", (today0,), "one")
    pending = db_exec(
        "SELECT COUNT(*) FROM atc_jobs WHERE status IN ('pending','submitted')", (), "one")[0]
    last_err = db_exec(
        "SELECT error FROM atc_jobs WHERE status='failed' AND error IS NOT NULL "
        "ORDER BY updated DESC LIMIT 1", (), "one")
    secret = cfg["secret"]
    return {
        "enabled": cfg["enabled"],
        "configured": bool(cfg["key"] and secret),
        "master_on": app_setting("atc_enabled") == "1",
        "api_key": cfg["key"],
        "api_secret_masked": (secret[:3] + "****" + secret[-2:]) if len(secret) > 5 else "",
        "base_url": cfg["base"],
        "play_enhance": cfg["play_enhance"],
        "transcript_enabled": cfg["transcript_enabled"],
        "transcript_daily": cfg["transcript_daily"],
        "url_ttl": cfg["url_ttl"],
        "play_priority": cfg["play_priority"],
        "queue_pending": pending,
        "today_total": int(row[0] or 0), "today_done": int(row[1] or 0),
        "today_failed": int(row[2] or 0),
        "last_error": (last_err[0] if last_err else "") or "",
        "test": app_setting("atc_test_state", ""),
    }


# ---- 文案提取（注册用户专属，每日限额；缓存命中不扣次）----

def _atc_transcript_status(user_id: int) -> tuple[int, int, int]:
    """返回 (limit, used, remaining)。计数主体与网页解析配额隔离（atc: 前缀）。"""
    cfg = _atc_cfg()
    limit = cfg["transcript_daily"]
    row = db_exec("SELECT count FROM usage_daily WHERE day=? AND subject=?",
                  (_today(), f"atc:user:{user_id}"), "one")
    used = int(row[0]) if row else 0
    return limit, used, max(0, limit - used)


def _atc_transcript_reserve(user_id: int) -> dict:
    """原子预占一次文案提取额度（与 reserve_quota 同款 BEGIN IMMEDIATE 模式）。"""
    cfg = _atc_cfg()
    day = _today()
    subject = f"atc:user:{user_id}"
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            row = conn.execute(
                "SELECT count FROM usage_daily WHERE day=? AND subject=?",
                (day, subject)).fetchone()
            used = int(row[0]) if row else 0
            take = 1 if used < cfg["transcript_daily"] else 0
            reservation_id = ""
            if take:
                conn.execute(
                    "INSERT INTO usage_daily(day,subject,count) VALUES(?,?,1) "
                    "ON CONFLICT(day,subject) DO UPDATE SET count=count+1",
                    (day, subject))
                reservation_id = "qr_" + secrets.token_urlsafe(12)
                now = int(time.time())
                conn.execute(
                    "INSERT INTO quota_reservations("
                    "id,day,subjects,units,committed_units,status,endpoint,created,lease_until"
                    ") VALUES(?,?,?,1,0,'pending','atc_transcript',?,?)",
                    (reservation_id, day, json.dumps([subject]), now,
                     now + QUOTA_RESERVATION_TTL))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    return {"ok": bool(take), "id": reservation_id,
            "limit": cfg["transcript_daily"], "used_after": used + take,
            "remaining": max(0, cfg["transcript_daily"] - used - take)}


class AtcTranscriptBody(BaseModel):
    item_id: str = ""


@app.post("/api/atc/transcript")
def api_atc_transcript(body: AtcTranscriptBody, request: Request):
    """提交文案提取。缓存命中秒回不扣次；否则扣一次额度并异步入队。"""
    cfg = _atc_cfg()
    if not (cfg["enabled"] and cfg["transcript_enabled"]):
        raise ApiError(404, "文案提取功能未开启")
    u = current_user(request)
    if not u:
        raise ApiError(401, "获取文案需要登录，注册后每天免费提取 "
                         f"{cfg['transcript_daily']} 次")
    item_id = (body.item_id or "").strip()[:40]
    if not item_id:
        raise ApiError(400, "缺少 item_id，请先解析作品")
    cached = _atc_cache_get(item_id)
    if cached and cached.get("text_content"):
        limit, used, remaining = _atc_transcript_status(u["id"])
        return {"state": "ready", "text": cached["text_content"],
                "audio_url": cached.get("audio_url") or "",
                "duration": cached.get("duration"), "cached": True,
                "remaining": remaining, "daily": limit}
    reservation = _atc_transcript_reserve(u["id"])
    if not reservation["ok"]:
        raise ApiError(429, f"今日文案提取次数已用完（每天 {reservation['limit']} 次）")
    try:
        _atc_enqueue(item_id, purpose="transcript")
    except Exception:
        release_quota(reservation)
        raise
    settle_quota(reservation, 1)
    return {"state": "processing", "cached": False,
            "remaining": reservation["remaining"], "daily": reservation["limit"]}


@app.get("/api/atc/transcript")
def api_atc_transcript_get(item_id: str, request: Request):
    """轮询提取状态：ready / processing / none。"""
    cfg = _atc_cfg()
    if not (cfg["enabled"] and cfg["transcript_enabled"]):
        raise ApiError(404, "文案提取功能未开启")
    u = current_user(request)
    if not u:
        raise ApiError(401, "请先登录")
    limit, used, remaining = _atc_transcript_status(u["id"])
    item_id = (item_id or "").strip()[:40]
    cached = _atc_cache_get(item_id) if item_id else None
    if cached and cached.get("text_content"):
        return {"state": "ready", "text": cached["text_content"],
                "audio_url": cached.get("audio_url") or "",
                "duration": cached.get("duration"),
                "remaining": remaining, "daily": limit}
    if item_id and db_exec(
            "SELECT id FROM atc_jobs WHERE item_id=? AND status IN ('pending','submitted')",
            (item_id,), "one"):
        return {"state": "processing", "remaining": remaining, "daily": limit}
    failed = db_exec(
        "SELECT error FROM atc_jobs WHERE item_id=? AND status='failed' "
        "ORDER BY updated DESC LIMIT 1", (item_id,), "one") if item_id else None
    if failed:
        return {"state": "failed", "error": "提取失败，请稍后重试",
                "remaining": remaining, "daily": limit}
    return {"state": "none", "remaining": remaining, "daily": limit}


# ---------------------------------------------------------------- 公共 API

class ParseBody(BaseModel):
    text: str


def _quota_error(limit: int):
    return ApiError(429, f"今日免费次数已用完（每天 {limit} 次）。注册登录后每天可用 "
                         f"{FREE_USER_DAILY} 次，或使用开放 API 按量调用。")


@app.post("/api/parse")
def api_parse(body: ParseBody, request: Request):
    reservation = reserve_quota(request, 1, endpoint="parse")
    if not reservation["ok"]:
        raise _quota_error(reservation["limit"])
    try:
        data = _parse_cached(body.text)
    except Exception:
        release_quota(reservation)
        log_request(request, "web", body.text[:100], False)
        raise
    settle_quota(reservation, 1)
    log_request(request, "web", body.text[:100], True)
    return data


# ---------------------------------------------------------------- 开放 API v1（异步任务 + 计费）

def _extract_links(text: str) -> list:
    links = re.findall(r"https://v\.douyin\.com/[\w-]+/?", text or "")
    seen, uniq = set(), []
    for l in links:
        if l not in seen:
            seen.add(l)
            uniq.append(l)
    return uniq


API_JOB_LEASE_SECONDS = max(120, int(os.environ.get("API_JOB_LEASE_SECONDS", "600")))
API_JOB_HEARTBEAT_SECONDS = max(
    10, min(API_JOB_LEASE_SECONDS // 3,
            int(os.environ.get("API_JOB_HEARTBEAT_SECONDS", "30"))))
_JOB_TERMINAL = ("succeeded", "failed", "cancelled")
_job_stop = threading.Event()
_job_wakeup = queue.Queue(maxsize=1)
_job_threads: list[threading.Thread] = []
_job_workers_guard = threading.Lock()
_job_instance = "jw_" + secrets.token_urlsafe(9)


def _api_price_from_conn(conn) -> int:
    row = conn.execute(
        "SELECT v FROM app_settings WHERE k='api_price_cents'").fetchone()
    try:
        return max(0, int(row["v"] if row else "1"))
    except Exception:
        return 1


def _json_list(value) -> list:
    try:
        out = json.loads(value or "[]")
        return out if isinstance(out, list) else []
    except Exception:
        return []


def _job_item_result(row) -> Optional[dict]:
    item = dict(row)
    status = item.get("status")
    if status == "succeeded":
        try:
            data = json.loads(item.get("result") or "{}")
        except Exception:
            data = {}
        out = {"link": item.get("link") or "", "ok": True, "data": data}
        if item.get("error"):
            out["warning"] = item["error"]
        return out
    if status in ("failed", "cancelled"):
        out = {"link": item.get("link") or "", "ok": False,
               "error": item.get("error") or "解析失败"}
        if status == "cancelled":
            out["code"] = "cancelled"
        return out
    return None


def _refresh_job_aggregate(conn, job_id: str, now: Optional[int] = None) -> None:
    """在调用者事务内，从 item 事实表重建 job 聚合；只在完成时写一次 results 快照。"""
    now = int(now or time.time())
    job = conn.execute("SELECT total FROM jobs WHERE id=?", (job_id,)).fetchone()
    if not job:
        return
    agg = conn.execute(
        "SELECT "
        "SUM(CASE WHEN status IN ('succeeded','failed','cancelled') THEN 1 ELSE 0 END) done,"
        "SUM(CASE WHEN status='succeeded' THEN 1 ELSE 0 END) ok,"
        "COALESCE(SUM(CASE WHEN status='succeeded' THEN price_cents ELSE 0 END),0) cost,"
        "SUM(CASE WHEN status='running' THEN 1 ELSE 0 END) running "
        "FROM job_items WHERE job_id=?", (job_id,)).fetchone()
    done = int(agg["done"] or 0)
    ok_n = int(agg["ok"] or 0)
    cost = int(agg["cost"] or 0)
    total = int(job["total"] or 0)
    if done >= total:
        status, finished = "done", now
        rows = conn.execute(
            "SELECT * FROM job_items WHERE job_id=? ORDER BY idx", (job_id,)).fetchall()
        results = [r for r in (_job_item_result(x) for x in rows) if r is not None]
        conn.execute(
            "UPDATE jobs SET status=?,done=?,ok=?,cost_cents=?,results=?,"
            "updated=?,finished=COALESCE(finished,?) WHERE id=?",
            (status, done, ok_n, cost, json.dumps(results, ensure_ascii=False),
             now, finished, job_id))
    else:
        status = "running" if int(agg["running"] or 0) or done else "pending"
        conn.execute(
            "UPDATE jobs SET status=?,done=?,ok=?,cost_cents=?,updated=?,finished=NULL "
            "WHERE id=?", (status, done, ok_n, cost, now, job_id))


def _claim_job_item(owner: str) -> Optional[dict]:
    """用数据库租约/CAS 领取一项；过期 running 可恢复，但绝不再次预扣。"""
    now = int(time.time())
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            row = conn.execute(
                "SELECT ji.*,j.key,j.user_id FROM job_items ji "
                "JOIN jobs j ON j.id=ji.job_id "
                "WHERE ji.reserved=1 AND j.status IN ('pending','running') AND "
                "(ji.status IN ('pending','reserved') OR "
                "(ji.status='running' AND COALESCE(ji.lease_until,0)<?)) "
                # 持久性异常导致某项重复出租约时，先让未重试项继续前进，避免队首饥饿。
                "ORDER BY COALESCE(ji.attempts,0),j.created,ji.idx LIMIT 1",
                (now,)).fetchone()
            if not row:
                conn.rollback()
                return None
            n = conn.execute(
                "UPDATE job_items SET status='running',lease_owner=?,lease_until=?,"
                "attempts=COALESCE(attempts,0)+1,started=COALESCE(started,?) "
                "WHERE job_id=? AND idx=? AND reserved=1 AND "
                "(status IN ('pending','reserved') OR "
                "(status='running' AND COALESCE(lease_until,0)<?))",
                (owner, now + API_JOB_LEASE_SECONDS, now,
                 row["job_id"], row["idx"], now)).rowcount
            if n != 1:
                conn.rollback()
                return None
            conn.execute(
                "UPDATE jobs SET status='running',updated=? "
                "WHERE id=? AND status<>'done'", (now, row["job_id"]))
            conn.commit()
            item = dict(row)
            item["lease_owner"] = owner
            item["attempts"] = int(row["attempts"] or 0) + 1
            return item
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()


def _release_job_lease(item: dict) -> None:
    """本进程遇到临时内部错误时立即让出租约，避免 heartbeat 把孤儿项永久续租。"""
    db_exec(
        "UPDATE job_items SET lease_until=0 WHERE job_id=? AND idx=? "
        "AND status='running' AND lease_owner=?",
        (item["job_id"], item["idx"], item["lease_owner"]))
    _wake_job_workers()


def _finish_job_item(item: dict, ok: bool, data: Optional[dict] = None,
                     error_message: str = "") -> bool:
    """CAS 完成 item，并在同一事务内扣 reserved/入 spent 或精确退款、写账本和日志。"""
    now = int(time.time())
    terminal = "succeeded" if ok else "failed"
    result_json = json.dumps(data or {}, ensure_ascii=False) if ok else None
    error_message = (error_message or "")[:300]
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            row = conn.execute(
                "SELECT ji.*,j.key,j.user_id FROM job_items ji "
                "JOIN jobs j ON j.id=ji.job_id "
                "WHERE ji.job_id=? AND ji.idx=? AND ji.status='running' "
                "AND ji.reserved=1 AND ji.lease_owner=?",
                (item["job_id"], item["idx"], item["lease_owner"])).fetchone()
            if not row:
                conn.rollback()       # 租约已被别的 worker 接管或已经结算
                return False
            changed = conn.execute(
                "UPDATE job_items SET status=?,reserved=0,result=?,error=?,finished=?,"
                "lease_owner=NULL,lease_until=NULL WHERE job_id=? AND idx=? "
                "AND status='running' AND reserved=1 AND lease_owner=?",
                (terminal, result_json, error_message or None, now,
                 row["job_id"], row["idx"], item["lease_owner"])).rowcount
            if changed != 1:
                conn.rollback()
                return False

            price = max(0, int(row["price_cents"] or 0))
            if ok:
                account_changed = conn.execute(
                    "UPDATE api_keys SET reserved_cents=reserved_cents-?,"
                    "spent_cents=spent_cents+?,calls=calls+1,last_used=? "
                    "WHERE key=? AND COALESCE(reserved_cents,0)>=?",
                    (price, price, now, row["key"], price)).rowcount
                event, balance_delta, reserved_delta = "charge", 0, -price
                spent_delta, calls_delta, reason = price, 1, "parse_succeeded"
            else:
                account_changed = conn.execute(
                    "UPDATE api_keys SET reserved_cents=reserved_cents-?,"
                    "balance_cents=balance_cents+? "
                    "WHERE key=? AND COALESCE(reserved_cents,0)>=?",
                    (price, price, row["key"], price)).rowcount
                event, balance_delta, reserved_delta = "refund", price, -price
                spent_delta, calls_delta, reason = 0, 0, "parse_failed"
            if account_changed != 1:
                raise RuntimeError("API 预授权账户不存在或 reserved_cents 对账失败")

            conn.execute(
                "INSERT INTO api_ledger("
                "ts,key,job_id,item_idx,event,balance_delta,reserved_delta,"
                "spent_delta,calls_delta,reason) VALUES(?,?,?,?,?,?,?,?,?,?)",
                (now, row["key"], row["job_id"], row["idx"], event,
                 balance_delta, reserved_delta, spent_delta, calls_delta, reason))
            conn.execute(
                "INSERT INTO api_logs("
                "ts,key,user_id,link,ok,cost_cents,job_id,item_idx"
                ") VALUES(?,?,?,?,?,?,?,?)",
                (now, row["key"], row["user_id"], row["link"], 1 if ok else 0,
                 price if ok else 0, row["job_id"], row["idx"]))
            _refresh_job_aggregate(conn, row["job_id"], now)
            conn.commit()
            return True
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()


def _run_claimed_job_item(item: dict) -> None:
    try:
        data = _parse_cached(item["link"])
    except ApiError as e:
        _finish_job_item(item, False, error_message=e.message)
    except Exception as e:
        _finish_job_item(
            item, False,
            error_message="内部解析错误：" + _redact_proxy_error(e))
    else:
        _finish_job_item(item, True, data=data)


def _wake_job_workers() -> None:
    try:
        _job_wakeup.put_nowait(True)
    except queue.Full:
        pass


def _job_worker_loop(worker_no: int) -> None:
    owner = f"{_job_instance}:{worker_no}"
    while not _job_stop.is_set():
        try:
            item = _claim_job_item(owner)
        except Exception:
            _job_stop.wait(1)
            continue
        if item is None:
            try:
                _job_wakeup.get(timeout=1)
            except queue.Empty:
                pass
            continue
        try:
            _run_claimed_job_item(item)
        except Exception:
            # 数据库瞬时故障时不改变预授权；释放租约后由本/下一进程重试。
            try:
                _release_job_lease(item)
            except Exception:
                pass
            _job_stop.wait(0.5)


def _job_heartbeat_loop() -> None:
    owner_prefix = _job_instance + ":"
    while not _job_stop.wait(API_JOB_HEARTBEAT_SECONDS):
        try:
            db_exec(
                "UPDATE job_items SET lease_until=? WHERE status='running' "
                "AND substr(lease_owner,1,?)=?",
                (int(time.time()) + API_JOB_LEASE_SECONDS,
                 len(owner_prefix), owner_prefix))
        except Exception:
            pass


def _recover_legacy_api_jobs() -> dict:
    """一次性终结旧 daemon 遗留任务；无法证明的至多一笔预扣按用户有利原则退款。"""
    now = int(time.time())
    recovered = refunded = 0
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            legacy = conn.execute(
                "SELECT j.* FROM jobs j WHERE j.status<>'done' AND NOT EXISTS "
                "(SELECT 1 FROM job_items ji WHERE ji.job_id=j.id) "
                "ORDER BY j.created").fetchall()
            current_price = _api_price_from_conn(conn)
            for job in legacy:
                links = [str(x) for x in _json_list(job["links"])]
                old_results = _json_list(job["results"])
                logs = conn.execute(
                    "SELECT * FROM api_logs WHERE job_id=? ORDER BY id",
                    (job["id"],)).fetchall()
                total = max(int(job["total"] or 0), len(links),
                            len(old_results), len(logs))
                positive_prices = [int(x["cost_cents"] or 0) for x in logs
                                   if int(x["cost_cents"] or 0) > 0]
                fallback_price = max(
                    [int(job["price_cents"] or 0), current_price, *positive_prices])
                built = []
                for idx in range(total):
                    prior = old_results[idx] if idx < len(old_results) else None
                    log = logs[idx] if idx < len(logs) else None
                    link = (links[idx] if idx < len(links) else
                            (prior.get("link", "") if isinstance(prior, dict) else
                             (log["link"] if log else "")))
                    if isinstance(prior, dict):
                        succeeded = bool(prior.get("ok"))
                        result = prior.get("data") if succeeded else None
                        err = "" if succeeded else str(prior.get("error") or "解析失败")
                    elif log is not None:
                        succeeded = bool(log["ok"])
                        result = {} if succeeded else None
                        err = ("旧任务已计费，但结果在重启前未完整落库"
                               if succeeded else "旧任务解析失败")
                    else:
                        succeeded, result = False, None
                        err = "服务升级时任务尚未完成，已取消并执行保守退款"
                    status = "succeeded" if succeeded else (
                        "failed" if log is not None or prior is not None else "cancelled")
                    item_price = (int(log["cost_cents"] or 0)
                                  if succeeded and log is not None else fallback_price)
                    conn.execute(
                        "INSERT OR IGNORE INTO job_items("
                        "job_id,idx,link,status,price_cents,reserved,result,error,"
                        "attempts,started,finished) VALUES(?,?,?,?,?,0,?,?,0,?,?)",
                        (job["id"], idx, link, status, item_price,
                         json.dumps(result or {}, ensure_ascii=False) if succeeded else None,
                         err or None, job["created"], now))
                    if log is not None and log["item_idx"] is None:
                        conn.execute(
                            "UPDATE api_logs SET item_idx=? WHERE id=? AND item_idx IS NULL",
                            (idx, log["id"]))
                    built.append(_job_item_result({
                        "link": link, "status": status,
                        "result": (json.dumps(result or {}, ensure_ascii=False)
                                   if succeeded else None),
                        "error": err or None,
                    }))

                # 旧执行器逐项串行，同一 job 在崩溃点至多有一笔“已扣余额但未结算”。
                # 旧 schema 没有证据能区分它是否发生，故只做一次、偏向用户的安全退款。
                key_row = conn.execute(
                    "SELECT 1 FROM api_keys WHERE key=?", (job["key"],)).fetchone()
                if key_row and fallback_price > 0:
                    conn.execute(
                        "UPDATE api_keys SET balance_cents=balance_cents+? WHERE key=?",
                        (fallback_price, job["key"]))
                    conn.execute(
                        "INSERT INTO api_ledger("
                        "ts,key,job_id,event,balance_delta,reason"
                        ") VALUES(?,?,?,?,?,?)",
                        (now, job["key"], job["id"], "legacy_safety_refund",
                         fallback_price, "legacy_daemon_state_ambiguous"))
                    refunded += fallback_price
                conn.execute(
                    "UPDATE jobs SET total=?,status='done',updated=?,finished=?,"
                    "results=? WHERE id=?",
                    (total, now, now,
                     json.dumps([x for x in built if x], ensure_ascii=False), job["id"]))
                _refresh_job_aggregate(conn, job["id"], now)
                recovered += 1
            conn.execute(
                "INSERT OR REPLACE INTO app_settings(k,v) "
                "VALUES('api_jobs_v2_legacy_recovered',?)", (str(now),))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    return {"jobs": recovered, "refunded_cents": refunded}


def _reconcile_api_job_accounts() -> dict:
    """启动对账：item 是预授权事实源；差异只按“不让用户少余额”的方向修复并记账。"""
    now = int(time.time())
    repaired = 0
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            keys = conn.execute(
                "SELECT key,COALESCE(balance_cents,0) balance_cents,"
                "COALESCE(reserved_cents,0) reserved_cents FROM api_keys").fetchall()
            for key_row in keys:
                expected_row = conn.execute(
                    "SELECT COALESCE(SUM(ji.price_cents),0) n FROM job_items ji "
                    "JOIN jobs j ON j.id=ji.job_id WHERE j.key=? AND ji.reserved=1 "
                    "AND ji.status IN ('pending','reserved','running')",
                    (key_row["key"],)).fetchone()
                expected = int(expected_row["n"] or 0)
                actual = int(key_row["reserved_cents"] or 0)
                if actual > expected:
                    delta = actual - expected
                    conn.execute(
                        "UPDATE api_keys SET reserved_cents=?,balance_cents=balance_cents+? "
                        "WHERE key=?", (expected, delta, key_row["key"]))
                    conn.execute(
                        "INSERT INTO api_ledger("
                        "ts,key,event,balance_delta,reserved_delta,reason"
                        ") VALUES(?,?,?,?,?,?)",
                        (now, key_row["key"], "reconcile_refund",
                         delta, -delta, "orphan_reserved_surplus"))
                    repaired += 1
                elif actual < expected:
                    # 正常事务不可能走到这里；若磁盘/人工改库造成差额，补足 reserved
                    # 而不再扣 available，避免恢复过程让用户二次付费。
                    delta = expected - actual
                    conn.execute(
                        "UPDATE api_keys SET reserved_cents=? WHERE key=?",
                        (expected, key_row["key"]))
                    conn.execute(
                        "INSERT INTO api_ledger("
                        "ts,key,event,reserved_delta,reason) VALUES(?,?,?,?,?)",
                        (now, key_row["key"], "reconcile_reserve",
                         delta, "missing_reserve_repaired_user_favor"))
                    repaired += 1
            job_ids = conn.execute(
                "SELECT DISTINCT job_id FROM job_items").fetchall()
            for row in job_ids:
                _refresh_job_aggregate(conn, row["job_id"], now)
            conn.execute(
                "INSERT OR REPLACE INTO app_settings(k,v) "
                "VALUES('api_jobs_last_reconciled',?)", (str(now),))
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    return {"accounts_repaired": repaired}


def _prepare_api_jobs() -> None:
    """启动线程前完成会依赖旧 api_logs 的恢复与账户对账。"""
    _recover_legacy_api_jobs()
    _reconcile_api_job_accounts()


def _start_api_job_workers(prepared: bool = False) -> None:
    with _job_workers_guard:
        _job_threads[:] = [t for t in _job_threads if t.is_alive()]
        if any(t.is_alive() for t in _job_threads):
            return
        if not prepared:
            _prepare_api_jobs()
        _job_stop.clear()
        while True:
            try:
                _job_wakeup.get_nowait()
            except queue.Empty:
                break
        heartbeat = threading.Thread(
            target=_job_heartbeat_loop, name="api-job-heartbeat", daemon=False)
        _job_threads.append(heartbeat)
        for i in range(API_JOB_WORKERS):
            _job_threads.append(threading.Thread(
                target=_job_worker_loop, args=(i,),
                name=f"api-job-worker-{i}", daemon=False))
        for thread in _job_threads:
            thread.start()
        _wake_job_workers()


def _stop_api_job_workers() -> None:
    with _job_workers_guard:
        _job_stop.set()
        _wake_job_workers()
        for thread in list(_job_threads):
            thread.join(timeout=API_JOB_LEASE_SECONDS + 5)
        _job_threads[:] = [t for t in _job_threads if t.is_alive()]


class JobBody(BaseModel):
    links: list = []
    text: str = ""


def _api_key_from(request: Request) -> str:
    # API Key 是长期计费凭据，只允许请求头；查询参数会泄露到 URL 历史和代理访问日志。
    return request.headers.get("X-API-Key") or ""


@app.post("/api/v1/jobs")
def api_v1_create_job(body: JobBody, request: Request):
    """原子预授权整批费用并持久化 item；Idempotency-Key 重放返回同一任务。"""
    key = _api_key_from(request)
    rec, err = api_key_check(key)
    if err:
        raise ApiError(401, err)
    links = list(body.links or []) or _extract_links(body.text)
    links = [str(l) for l in links
             if re.match(r"https://v\.douyin\.com/[\w-]+", str(l))][:100]
    if not links:
        raise ApiError(400, "links 为空或没有合法的 v.douyin.com 链接")
    request_id = (request.headers.get("Idempotency-Key") or "").strip() or None
    if request_id and len(request_id) > 128:
        raise ApiError(400, "Idempotency-Key 最长 128 个字符")
    links_json = json.dumps(links, ensure_ascii=False)
    now = int(time.time())
    replay = None
    with _db_lock:
        conn = _db()
        try:
            conn.execute("BEGIN IMMEDIATE")
            if request_id:
                replay = conn.execute(
                    "SELECT * FROM jobs WHERE key=? AND request_id=?",
                    (key, request_id)).fetchone()
                if replay:
                    if _json_list(replay["links"]) != links:
                        raise ApiError(409, "同一 Idempotency-Key 不能提交不同链接")
                    conn.commit()
                else:
                    replay = None
            if not replay:
                key_row = conn.execute(
                    "SELECT * FROM api_keys WHERE key=?", (key,)).fetchone()
                if (not key_row or not key_row["enabled"]
                        or key_row["deleted_at"] is not None):
                    raise ApiError(401, "无效或已禁用的 API Key")
                price = _api_price_from_conn(conn)
                total_cost = price * len(links)
                if int(key_row["balance_cents"] or 0) < total_cost:
                    raise ApiError(
                        402, f"余额不足（当前 {key_row['balance_cents']} 分，"
                        f"本任务需预授权 {total_cost} 分），请充值")
                job_id = "job_" + secrets.token_urlsafe(12)
                conn.execute(
                    "UPDATE api_keys SET balance_cents=balance_cents-?,"
                    "reserved_cents=COALESCE(reserved_cents,0)+? WHERE key=?",
                    (total_cost, total_cost, key))
                conn.execute(
                    "INSERT INTO jobs("
                    "id,key,user_id,status,total,done,ok,cost_cents,links,results,"
                    "created,price_cents,updated,request_id"
                    ") VALUES(?,?,?,?,?,0,0,0,?,'[]',?,?,?,?)",
                    (job_id, key, key_row["user_id"], "pending", len(links),
                     links_json, now, price, now, request_id))
                for idx, link in enumerate(links):
                    conn.execute(
                        "INSERT INTO job_items("
                        "job_id,idx,link,status,price_cents,reserved,attempts"
                        ") VALUES(?,?,?,'reserved',?,1,0)",
                        (job_id, idx, link, price))
                    conn.execute(
                        "INSERT INTO api_ledger("
                        "ts,key,job_id,item_idx,event,balance_delta,reserved_delta,reason"
                        ") VALUES(?,?,?,?,?,?,?,?)",
                        (now, key, job_id, idx, "reserve",
                         -price, price, "job_pre_authorized"))
                conn.commit()
                replay = conn.execute(
                    "SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    _wake_job_workers()
    j = dict(replay)
    return {"code": 0, "message": "accepted", "data": {
        "job_id": j["id"], "total": j["total"], "status": j["status"],
        "price_cents": j["price_cents"],
        "estimated_cost_cents": int(j["price_cents"] or 0) * int(j["total"] or 0),
        "query_url": f"/api/v1/jobs/{j['id']}"}}


@app.get("/api/v1/jobs/{job_id}")
def api_v1_get_job(job_id: str, request: Request):
    """查询任务结果。需带同一 API Key。"""
    rec, err = api_key_check(_api_key_from(request))
    if err:
        raise ApiError(401, err)
    with _db_lock:
        conn = _db()
        try:
            row = conn.execute(
                "SELECT * FROM jobs WHERE id=? AND key=?",
                (job_id, rec["key"])).fetchone()
            if not row:
                raise ApiError(404, "任务不存在或无权访问")
            items = conn.execute(
                "SELECT * FROM job_items WHERE job_id=? ORDER BY idx",
                (job_id,)).fetchall()
        finally:
            conn.close()
    j = dict(row)
    if items:
        terminal = [x for x in items if x["status"] in _JOB_TERMINAL]
        done = len(terminal)
        ok_n = sum(1 for x in terminal if x["status"] == "succeeded")
        cost = sum(int(x["price_cents"] or 0) for x in terminal
                   if x["status"] == "succeeded")
        status = ("done" if done >= int(j["total"] or 0) else
                  ("running" if done or any(x["status"] == "running" for x in items)
                   else "pending"))
        results = [r for r in (_job_item_result(x) for x in items) if r is not None]
    else:                       # v1 已完成任务仍可按旧 results 快照查询
        done, ok_n, cost, status = j["done"], j["ok"], j["cost_cents"], j["status"]
        results = _json_list(j["results"])
    return {"code": 0, "message": "ok", "data": {
        "job_id": j["id"], "status": status, "total": j["total"],
        "done": done, "ok": ok_n, "cost_cents": cost,
        "created": j["created"], "finished": j["finished"],
        "results": results}}


@app.get("/api/v1/balance")
def api_v1_balance(request: Request):
    """查询当前 Key 的余额与用量。"""
    rec, err = api_key_check(_api_key_from(request))
    if err:
        raise ApiError(401, err)
    return {"code": 0, "data": {
        "balance_cents": rec["balance_cents"], "spent_cents": rec["spent_cents"],
        "reserved_cents": int(rec.get("reserved_cents") or 0),
        "calls": rec["calls"], "price_cents": api_price_cents()}}


def _fetch_user_info(sec_uid: str) -> dict:
    """经代理拉取作者主页统计（免签名 reflow 接口）：粉丝数、获赞数、作品数等。"""
    url = f"https://www.iesdouyin.com/web/api/v2/user/info/?sec_uid={sec_uid}"
    # 浏览器无法跨域取（抖音接口无 CORS），只能服务器代拉 —— 走代理，不暴露服务器 IP
    resp, _ = open_url(url, headers={"Referer": "https://www.iesdouyin.com/"})
    try:
        ui = (json.loads(resp.read().decode("utf-8", "ignore")) or {}).get("user_info") or {}
    finally:
        try:
            resp.close()
        except Exception:
            pass
    return {
        "follower_count": ui.get("mplatform_followers_count"),
        "total_favorited": ui.get("total_favorited"),
        "following_count": ui.get("following_count"),
        "aweme_count": ui.get("aweme_count"),
        "douyin_id": ui.get("unique_id") or "",
        "signature": (ui.get("signature") or "").strip(),
    }


@app.get("/api/author")
def api_author(item_id: str):
    """作者结构化详情（供前端悬停浮层）。

    基础字段来自解析时缓存的分享页 author 对象；首次请求时再直连（不走代理）拉一次
    user/info 富化粉丝数/获赞数（分享页不给这两项），结果服务端缓存 10 分钟。
    注：抖音该接口无 CORS/JSONP，浏览器无法跨域直取，故由服务器直连（非代理）代拉。
    """
    hit = _author_cache.get(item_id)
    if not hit:
        raise ApiError(404, "作者信息不存在或已过期，请重新解析该视频")
    detail = hit[1]
    if detail.get("enriched") and time.time() - hit[0] < 600:
        return detail
    sec = detail.get("sec_uid")
    if sec:
        try:
            merged = {**detail, **{k: v for k, v in _fetch_user_info(sec).items() if v is not None},
                      "enriched": True}
            _author_cache[item_id] = (time.time(), merged)
            return merged
        except Exception:
            pass                         # 富化失败就返回基础字段，不影响头像浮层
    return detail


class BatchBody(BaseModel):
    text: str


@app.post("/api/parse/batch")
def api_parse_batch(body: BatchBody, request: Request):
    """批量解析：每条链接算一次配额，超出今日免费额度的部分不解析。"""
    links = re.findall(r"https://v\.douyin\.com/[\w-]+/?", body.text)
    seen, uniq = set(), []
    for l in links:
        if l not in seen:
            seen.add(l)
            uniq.append(l)
    if not uniq:
        raise ApiError(400, "未找到任何 v.douyin.com 分享链接")

    uniq = uniq[:50]
    reservation = reserve_quota(
        request, len(uniq), partial=True, endpoint="parse_batch")
    if reservation["reserved"] <= 0:
        raise _quota_error(reservation["limit"])
    limit = reservation["limit"]
    process = uniq[:reservation["reserved"]]
    over = uniq[reservation["reserved"]:]   # 超额部分不解析

    out, spent = [], 0
    for l in process:
        try:
            out.append({"ok": True, "link": l, "data": _parse_cached(l)})
            spent += 1
            log_request(request, "web", l, True)
        except ApiError as e:
            out.append({"ok": False, "link": l, "error": e.message})
            log_request(request, "web", l, False)
        except Exception:
            out.append({"ok": False, "link": l, "error": "内部解析错误，请稍后重试"})
            log_request(request, "web", l, False)
    for l in over:
        out.append({"ok": False, "link": l,
                    "error": f"今日免费次数不足未解析（每天 {limit} 次，登录后 {FREE_USER_DAILY} 次）"})
    settle_quota(reservation, spent)
    remaining = quota_status(request)[2]
    return {"count": len(out), "results": out,
            "quota": {"limit": limit, "remaining": remaining}}


class ExportBody(BaseModel):
    items: list          # 前端解析好的结果数组（仅元数据/文案，不含媒体字节）


@app.post("/api/export/xlsx")
def export_xlsx(body: ExportBody):
    """把批量解析结果导出为真正的 Excel(.xlsx)。仅整理元数据，不下载任何视频。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "抖音批量解析"
    headers = ["序号", "类型", "标题/文案", "作者", "作品ID", "时长(秒)", "分辨率",
               "点赞", "评论", "收藏", "分享", "发布时间", "话题标签", "背景音乐",
               "拍摄位置", "视频/图片地址", "作者主页", "原分享链接"]
    ws.append(headers)
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="E0234E")
    for c in ws[1]:
        c.font, c.fill = hf, hfill
        c.alignment = Alignment(vertical="center")

    for i, d in enumerate(body.items or [], 1):
        d = d or {}
        is_note = d.get("kind") == "note"
        st = d.get("stats") or {}
        v = d.get("video") or {}
        ct = d.get("create_time")
        cts = time.strftime("%Y-%m-%d %H:%M", time.localtime(ct)) if ct else ""
        media = (" | ".join(im.get("url", "") for im in (d.get("images") or []))
                 if is_note else v.get("url", ""))
        res = f"{v.get('width')}×{v.get('height')}" if v.get("width") else ""
        dur = round((d.get("duration_ms") or 0) / 1000, 1) if not is_note else ""
        ws.append([
            i, "图集" if is_note else "视频", d.get("title", ""), d.get("author", ""),
            d.get("item_id", ""), dur, res,
            st.get("digg"), st.get("comment"), st.get("collect"), st.get("share"),
            cts, " ".join("#" + t for t in (d.get("tags") or [])),
            (d.get("music") or {}).get("title") or "", d.get("location") or "",
            media, d.get("author_url", ""), d.get("_link", ""),
        ])

    widths = [5, 6, 40, 14, 20, 8, 11, 8, 8, 8, 8, 17, 24, 24, 20, 46, 40, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(38 + col)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = time.strftime("douyin_batch_%Y%m%d_%H%M.xlsx")
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": _content_disposition(fname)})


def _close_upstream(resp) -> None:
    try:
        resp.close()
    except Exception:
        pass


def _open_video_upstream(vid: str, headers: dict, validator=None):
    """打开一个可信的视频响应；主线路异常时切到备用播放域名。"""
    no_proxy_error = None
    for upstream in (_play_api(vid), _play_api_alt(vid)):
        resp = None
        accepted = False
        try:
            resp, _ = open_url(
                upstream,
                headers=headers,
                retry_http_statuses=(408, 425, 429, 500, 502, 503, 504),
                ban_on_auth_error=False,
            )
            status = resp.status if hasattr(resp, "status") else resp.getcode()
            if status not in (200, 206):
                raise ValueError(f"unexpected video status {status}")

            content_type = (resp.headers.get("Content-Type") or "").split(";", 1)[0]
            content_type = content_type.strip().lower()
            if not content_type or not (
                content_type.startswith("video/")
                or content_type in {
                    "application/mp4",
                    "application/octet-stream",
                    "binary/octet-stream",
                }
            ):
                raise ValueError(f"unexpected video content type {content_type}")

            geturl = getattr(resp, "geturl", None)
            final_url = geturl() if callable(geturl) else ""
            if final_url and not _host_allowed(final_url):
                raise ValueError("video redirect left the Douyin media allowlist")
            if validator:
                validator(resp)

            accepted = True
            return resp
        except ApiError as exc:
            if exc.status == 503:
                no_proxy_error = exc
                break
        except urlerr.HTTPError as exc:
            _close_upstream(exc)
        except Exception:
            pass
        finally:
            if resp is not None and not accepted:
                _close_upstream(resp)

    if no_proxy_error:
        raise no_proxy_error
    raise ApiError(502, "视频下载线路暂时不可用，请稍后重试")


@app.get("/api/video/{vid}")
def api_video(vid: str, request: Request, exp: int = 0, sig: str = "",
              dl: str = "", name: str = "video.mp4"):
    if not re.fullmatch(r"[\w-]{8,120}", vid):
        raise ApiError(400, "非法的视频 ID")
    _require_media_token("video", vid, exp, sig)
    range_header = request.headers.get("range", "")
    if not _valid_single_range(range_header):
        raise ApiError(416, "仅支持单段 bytes Range 请求")
    lease = _media_lease(request)
    extra = dict(CDN_HEADERS)
    if range_header:
        extra["Range"] = range_header
    try:
        # 同源播放/下载线路：经代理，绝不直连暴露服务器 IP。
        # 主播放域名被风控、返回网关页或临时 5xx 时，自动切换备用域名。
        resp = _open_video_upstream(
            vid, extra,
            validator=lambda candidate: _video_response_shape(
                candidate, range_header))
        status, start, end, total, expected = _video_response_shape(
            resp, range_header)
    except ApiError:
        _media_release(lease)
        raise
    except Exception:
        _media_release(lease)
        raise ApiError(502, "视频下载线路暂时不可用，请稍后重试")

    headers = {
        "Accept-Ranges": "bytes",
        "Cache-Control": "private, no-store",
        "X-Content-Type-Options": "nosniff",
    }
    if expected is not None:
        headers["Content-Length"] = str(expected)
    if status == 206:
        headers["Content-Range"] = (
            f"bytes {start}-{end}/"
            f"{total if total is not None else '*'}")
    if dl:
        headers["Content-Disposition"] = _content_disposition(name or "video.mp4")
    stream = _ResumableVideoStream(
        vid, resp, extra, start, end, total, expected)
    # 流关闭时按实际转发字节计量（stream.sent 即已发给客户端的字节数）
    scope = "download" if dl else "play"
    finalize = _media_finalizer(
        stream, lease, on_close=lambda: _traffic_add(scope, stream.sent))
    stream.set_on_close(finalize)
    return _MediaStreamingResponse(
        stream,
        finalize=finalize, status_code=status,
        media_type="video/mp4", headers=headers)


# 注：图集打包 ZIP 需服务器逐张下载再压缩，会走服务器 IP/带宽，
# 图集打包与"图片下载走浏览器直连"的设计冲突，已改为前端逐张下载（downloadAll）。


# ---------------------------------------------------------------- 管理后台

_sessions: dict[str, float] = {}     # token -> 过期时间
SESSION_TTL = 12 * 3600


def _new_session() -> str:
    tok = secrets.token_urlsafe(24)
    _sessions[tok] = time.time() + SESSION_TTL
    return tok


def _require_admin(request: Request):
    tok = request.cookies.get("admin_session", "")
    exp = _sessions.get(tok)
    if not exp or exp < time.time():
        _sessions.pop(tok, None)
        raise ApiError(401, "未登录或会话已过期，请重新登录管理后台")


class LoginBody(BaseModel):
    password: str


@app.post("/api/admin/login")
def admin_login(body: LoginBody, request: Request):
    ip = _client_ip(request)
    if _admin_fail_count(ip) >= ADMIN_LOGIN_MAX_FAILS:
        raise ApiError(429, "登录失败次数过多，账号已临时锁定，请 15 分钟后再试")
    if not secrets.compare_digest(body.password, ADMIN_PASSWORD):
        _admin_record_fail(ip)
        left = ADMIN_LOGIN_MAX_FAILS - _admin_fail_count(ip)
        raise ApiError(403, f"密码错误，还可尝试 {left} 次" if left > 0
                       else "密码错误，账号已临时锁定，请 15 分钟后再试")
    _admin_fails.pop(ip, None)          # 成功登录 → 清零失败计数
    tok = _new_session()
    resp = JSONResponse({"ok": True})
    resp.set_cookie("admin_session", tok, httponly=True, samesite="lax",
                    secure=COOKIE_SECURE, max_age=SESSION_TTL)
    return resp


@app.post("/api/admin/logout")
def admin_logout(request: Request):
    tok = request.cookies.get("admin_session", "")
    _sessions.pop(tok, None)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("admin_session")
    return resp


@app.get("/api/admin/state")
def admin_state(request: Request):
    _require_admin(request)
    return {
        "proxies": proxy_mgr.proxies,
        "settings": proxy_mgr.settings,
        "stats": proxy_mgr.stats,
        "ua_pool_size": len(UA_POOL),
        "captcha": dict(_captcha_stats),     # 滑块漏斗（内存计数，重启清零）
    }


class AddProxyBody(BaseModel):
    urls: str
    note: str = ""


@app.post("/api/admin/proxies")
def admin_add_proxy(body: AddProxyBody, request: Request):
    _require_admin(request)
    return proxy_mgr.add_many(body.urls, body.note)


@app.delete("/api/admin/proxies/{pid}")
def admin_del_proxy(pid: str, request: Request):
    _require_admin(request)
    if not proxy_mgr.remove(pid):
        raise ApiError(404, "代理不存在")
    return {"ok": True}


@app.post("/api/admin/proxies/{pid}/toggle")
def admin_toggle_proxy(pid: str, request: Request):
    _require_admin(request)
    state = proxy_mgr.toggle(pid)
    if state is None:
        raise ApiError(404, "代理不存在")
    return {"ok": True, "enabled": state}


PROBE_TIMEOUT = 25    # 住宅代理较慢，给足超时


def _probe_proxy(p: dict, reach_douyin: bool = True) -> dict:
    """测试代理：出口 IP + 延迟，可选附带抖音可达性检测。回写状态并处理自愈/禁用。

    住宅代理每请求轮换 IP 且延迟高，抖音可达性重试 2 次以降低误报。
    """
    t0 = time.time()
    try:
        r = _raw_open(TEST_URL_IP, True, {"User-Agent": pick_ua()}, PROBE_TIMEOUT, p)
        body = r.read().decode("utf-8", "ignore")
        r.close()
        ip = json.loads(body).get("ip", "?")
        latency = int((time.time() - t0) * 1000)
    except Exception as e:
        proxy_mgr.record_probe(p, ok=False)
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}

    douyin_ok = None
    if reach_douyin:
        douyin_ok = False
        for _ in range(2):                       # 轮换住宅代理：重试降低误报
            try:
                r = _raw_open(TEST_URL_DOUYIN, True, {"User-Agent": pick_ua()},
                              PROBE_TIMEOUT, p)
                r.read(128)
                r.close()
                douyin_ok = True
                break
            except Exception:
                continue

    proxy_mgr.record_probe(p, ok=True, latency_ms=latency, exit_ip=ip, douyin_ok=douyin_ok)
    return {"ok": True, "exit_ip": ip, "latency_ms": latency, "douyin_ok": douyin_ok}


def _probe_all(proxies: list[dict], reach_douyin: bool) -> list[dict]:
    """并发测试一批代理。"""
    import concurrent.futures as cf
    out = [None] * len(proxies)
    with cf.ThreadPoolExecutor(max_workers=min(8, max(1, len(proxies)))) as ex:
        futs = {ex.submit(_probe_proxy, p, reach_douyin): i for i, p in enumerate(proxies)}
        for f in cf.as_completed(futs):
            i = futs[f]
            out[i] = {"id": proxies[i]["id"], **f.result()}
    return out


class ValidateBody(BaseModel):
    urls: str


@app.post("/api/admin/proxies/validate")
def admin_validate(body: ValidateBody, request: Request):
    """预览解析结果：把用户粘贴的内容规范化成标准格式，不落库。"""
    _require_admin(request)
    scheme = proxy_mgr.settings.get("default_protocol", "socks5")
    out = []
    for line in re.split(r"[\r\n,;]+|\s{2,}", body.urls.strip()):
        line = line.strip()
        if not line:
            continue
        parsed = ProxyManager.parse_proxy(line, scheme)
        out.append({"raw": line, "parsed": parsed, "ok": bool(parsed)})
    return {"results": out}


@app.post("/api/admin/proxies/{pid}/test")
def admin_test_proxy(pid: str, request: Request):
    _require_admin(request)
    p = proxy_mgr.get(pid)
    if not p:
        raise ApiError(404, "代理不存在")
    return _probe_proxy(p, proxy_mgr.settings.get("test_reach_douyin", True))


@app.post("/api/admin/proxies/test-all")
def admin_test_all(request: Request):
    _require_admin(request)
    reach = proxy_mgr.settings.get("test_reach_douyin", True)
    results = _probe_all(list(proxy_mgr.proxies), reach)
    ok = sum(1 for r in results if r and r.get("ok"))
    return {"results": results, "ok": ok, "total": len(results)}


class ProxyBatchBody(BaseModel):
    action: str                            # delete / enable / disable / test
    ids: list[str]


@app.post("/api/admin/proxies/batch")
def admin_batch_proxy(body: ProxyBatchBody, request: Request):
    """批量操作选中的代理；managed 托管条目跳过增删启停（归 mihomo 面板管）。"""
    _require_admin(request)
    ids = set(body.ids)
    if not ids:
        raise ApiError(400, "未选择任何代理")
    if len(ids) > 500:
        raise ApiError(400, "单次批量操作最多 500 个")
    if body.action == "delete":
        return {"ok": True, "affected": proxy_mgr.remove_many(ids)}
    if body.action in ("enable", "disable"):
        return {"ok": True, "affected": proxy_mgr.set_enabled_many(ids, body.action == "enable")}
    if body.action == "test":
        selected = [p for p in proxy_mgr.proxies if p["id"] in ids]
        reach = proxy_mgr.settings.get("test_reach_douyin", True)
        results = _probe_all(selected, reach)
        ok = sum(1 for r in results if r and r.get("ok"))
        return {"results": results, "ok": ok, "total": len(results)}
    raise ApiError(400, "不支持的批量操作")


class SettingBody(BaseModel):
    force_proxy: Optional[bool] = None
    default_protocol: Optional[str] = None
    rotation: Optional[str] = None
    retries: Optional[int] = None
    auto_health: Optional[bool] = None
    health_interval_min: Optional[int] = None
    auto_disable_fail: Optional[int] = None
    test_reach_douyin: Optional[bool] = None


@app.post("/api/admin/settings")
def admin_settings(body: SettingBody, request: Request):
    _require_admin(request)
    vals = body.dict(exclude_none=True)
    if "default_protocol" in vals and vals["default_protocol"] not in SUPPORTED_SCHEMES:
        raise ApiError(400, "不支持的默认协议")
    if "rotation" in vals and vals["rotation"] not in ("round_robin", "random", "least_fail"):
        raise ApiError(400, "不支持的轮换策略")
    if "retries" in vals:
        vals["retries"] = max(1, min(10, int(vals["retries"])))
    if "health_interval_min" in vals:
        vals["health_interval_min"] = max(1, min(1440, int(vals["health_interval_min"])))
    if "auto_disable_fail" in vals:
        vals["auto_disable_fail"] = max(0, min(100, int(vals["auto_disable_fail"])))
    for k, v in vals.items():
        proxy_mgr.set_setting(k, v)
    return {"ok": True, "settings": proxy_mgr.settings}


# ---- 开放 API 密钥管理（管理员）----

@app.get("/api/admin/apikeys")
def admin_list_keys(request: Request):
    _require_admin(request)
    return {"keys": list_api_keys(), "free_ip_daily": FREE_ANON_DAILY,
            "price_cents": api_price_cents()}


class NewKeyBody(BaseModel):
    name: str = ""


@app.post("/api/admin/apikeys")
def admin_create_key(body: NewKeyBody, request: Request):
    _require_admin(request)
    return create_api_key(None, body.name)


class KeyBody(BaseModel):
    key: str


@app.post("/api/admin/apikeys/revoke")
def admin_revoke_key(body: KeyBody, request: Request):
    _require_admin(request)
    if not revoke_api_key(body.key):
        raise ApiError(404, "API Key 不存在")
    return {"ok": True}


class RechargeBody(KeyBody):
    cents: int


@app.post("/api/admin/apikeys/recharge")
def admin_recharge_key(body: RechargeBody, request: Request):
    _require_admin(request)
    if not recharge_key(body.key, body.cents):
        raise ApiError(404, "API Key 不存在")
    return {"ok": True, "key": get_api_key(body.key)}


class PriceBody(BaseModel):
    price_cents: int


@app.post("/api/admin/api-price")
def admin_set_price(body: PriceBody, request: Request):
    _require_admin(request)
    set_app_setting("api_price_cents", max(0, int(body.price_cents)))
    return {"ok": True, "price_cents": api_price_cents()}


# ---- 数据分析 ----

def _series(sql: str, days: int = 14):
    """返回最近 days 天的 {day: value} 序列（day 为 epoch 天）。sql 需 SELECT day, val。"""
    since = (_today() - days + 1) * 86400
    rows = db_exec(sql, (since,), "all") or []
    m = {r[0]: r[1] for r in rows}
    return [{"day": _today() - i, "v": m.get(_today() - i, 0)} for i in range(days - 1, -1, -1)]


@app.get("/api/admin/analytics")
def admin_analytics(request: Request):
    _require_admin(request)
    today = _today()
    day0 = today * 86400

    def one(sql, params=()):
        r = db_exec(sql, params, "one")
        return (r[0] or 0) if r else 0

    total_users = one("SELECT COUNT(*) FROM users")
    new_users_today = one("SELECT COUNT(*) FROM users WHERE created_at>=?", (day0,))
    pv_today = one("SELECT COUNT(*) FROM page_views WHERE ts>=?", (day0,))
    uv_today = one("SELECT COUNT(DISTINCT ip) FROM page_views WHERE ts>=?", (day0,))
    web_today = one("SELECT COUNT(*) FROM request_logs WHERE ok=1 AND ts>=?", (day0,))
    api_today = one("SELECT COUNT(*) FROM api_logs WHERE ok=1 AND ts>=?", (day0,))
    rev_today = one("SELECT COALESCE(SUM(cost_cents),0) FROM api_logs WHERE ts>=?", (day0,))
    rev_total = one("SELECT COALESCE(SUM(cost_cents),0) FROM api_logs")
    # 回访率：注册后又回来过（last_login 比注册晚 1 天以上）
    returned = one("SELECT COUNT(*) FROM users WHERE last_login-created_at>=86400")
    retention = round(returned / total_users * 100, 1) if total_users else 0.0

    return {
        "cards": {
            "total_users": total_users, "new_users_today": new_users_today,
            "pv_today": pv_today, "uv_today": uv_today,
            "usage_today": web_today + api_today, "api_today": api_today,
            "revenue_today_cents": rev_today, "revenue_total_cents": rev_total,
            "retention_pct": retention,
        },
        "series": {
            "new_users": _series("SELECT created_at/86400, COUNT(*) FROM users WHERE created_at>=? GROUP BY 1"),
            "pv": _series("SELECT ts/86400, COUNT(*) FROM page_views WHERE ts>=? GROUP BY 1"),
            "uv": _series("SELECT ts/86400, COUNT(DISTINCT ip) FROM page_views WHERE ts>=? GROUP BY 1"),
            "parses": _series("SELECT ts/86400, COUNT(*) FROM request_logs WHERE ok=1 AND ts>=? GROUP BY 1"),
            "api_calls": _series("SELECT ts/86400, COUNT(*) FROM api_logs WHERE ok=1 AND ts>=? GROUP BY 1"),
            "revenue": _series("SELECT ts/86400, COALESCE(SUM(cost_cents),0) FROM api_logs WHERE ts>=? GROUP BY 1"),
        },
    }


@app.get("/api/admin/users")
def admin_users(request: Request, limit: int = 100):
    _require_admin(request)
    rows = db_exec(
        "SELECT u.id,u.email,u.created_at,u.last_login,u.disabled,u.reg_ip,"
        "(SELECT COUNT(*) FROM request_logs r WHERE r.user_id=u.id AND r.ok=1) AS parses,"
        "(SELECT COALESCE(SUM(spent_cents),0) FROM api_keys k WHERE k.user_id=u.id) AS spent,"
        "(SELECT COUNT(*) FROM api_keys k WHERE k.user_id=u.id) AS keys "
        "FROM users u ORDER BY u.created_at DESC LIMIT ?", (min(limit, 500),), "all")
    return {"users": [dict(r) for r in rows]}


class UserToggleBody(BaseModel):
    disabled: bool


@app.post("/api/admin/users/{uid}/toggle")
def admin_toggle_user(uid: int, body: UserToggleBody, request: Request):
    _require_admin(request)
    db_exec("UPDATE users SET disabled=? WHERE id=?", (1 if body.disabled else 0, uid))
    return {"ok": True}


@app.get("/api/admin/logs/web")
def admin_web_logs(request: Request, limit: int = 100):
    _require_admin(request)
    rows = db_exec("SELECT ts,ip,ua,link,ok,user_id FROM request_logs ORDER BY id DESC LIMIT ?",
                   (min(limit, 500),), "all")
    return {"logs": [dict(r) for r in rows]}


@app.get("/api/admin/logs/api")
def admin_api_logs(request: Request, limit: int = 100):
    _require_admin(request)
    rows = db_exec("SELECT a.ts,a.key,a.user_id,a.link,a.ok,a.cost_cents,a.job_id,u.email "
                   "FROM api_logs a LEFT JOIN users u ON u.id=a.user_id ORDER BY a.id DESC LIMIT ?",
                   (min(limit, 500),), "all")
    return {"logs": [dict(r) for r in rows]}


# ---- AnyToCopy 增强线路（配置 / 测试 / 播放优先级）----

@app.get("/api/admin/atc")
def admin_atc_get(request: Request):
    _require_admin(request)
    status = _atc_status()
    # 有在途的测试任务时顺带推进一次（管理操作，频率极低）
    try:
        test = json.loads(app_setting("atc_test_state", "") or "{}")
    except ValueError:
        test = {}
    if test.get("state") == "submitted" and test.get("task_id"):
        cfg = _atc_cfg()
        try:
            resp = _atc_request("GET", "/video/query",
                                {"taskId": test["task_id"]}, cfg)
            data = resp.get("data") or {}
            st = data.get("status", "")
            if st == "SUCCESS":
                test = {"state": "success", "ms": test.get("ms"),
                        "duration": data.get("duration"),
                        "has_text": bool(data.get("textContent"))}
            elif st in ("FAILED", "FAILURE"):
                test = {"state": "failed",
                        "error": (data.get("errorMessage") or "任务失败")[:200]}
            set_app_setting("atc_test_state", json.dumps(test, ensure_ascii=False))
            status["test"] = json.dumps(test, ensure_ascii=False)
        except Exception:
            pass
    return status


class AtcSettingsBody(BaseModel):
    api_key: Optional[str] = None
    api_secret: Optional[str] = None
    base_url: Optional[str] = None
    enabled: Optional[bool] = None
    play_enhance: Optional[bool] = None
    transcript_enabled: Optional[bool] = None
    transcript_daily: Optional[int] = None
    url_ttl: Optional[int] = None
    play_priority: Optional[list] = None


@app.post("/api/admin/atc")
def admin_atc_set(body: AtcSettingsBody, request: Request):
    _require_admin(request)
    if body.api_key is not None:
        set_app_setting("atc_api_key", body.api_key.strip()[:100])
    if body.api_secret is not None:      # 空串 = 清除；前端不回传打码值
        set_app_setting("atc_api_secret", body.api_secret.strip()[:100])
    if body.base_url is not None:
        base = body.base_url.strip().rstrip("/")
        if base and not base.startswith("https://"):
            raise ApiError(400, "Base URL 必须是 https:// 地址")
        set_app_setting("atc_base_url", base[:200])
    if body.enabled is not None:
        set_app_setting("atc_enabled", "1" if body.enabled else "0")
    if body.play_enhance is not None:
        set_app_setting("atc_play_enhance", "1" if body.play_enhance else "0")
    if body.transcript_enabled is not None:
        set_app_setting("atc_transcript_enabled", "1" if body.transcript_enabled else "0")
    if body.transcript_daily is not None:
        set_app_setting("atc_transcript_daily",
                        str(max(0, min(100, int(body.transcript_daily)))))
    if body.url_ttl is not None:
        set_app_setting("atc_url_ttl", str(max(600, min(86400, int(body.url_ttl)))))
    if body.play_priority is not None:
        if not (isinstance(body.play_priority, list)
                and sorted(body.play_priority) == sorted(SHARE_PLAY_SOURCES)):
            raise ApiError(400, "播放优先级必须且只能包含 dy1 / dy2 / atc / proxy 四项")
        set_app_setting("share_play_priority", json.dumps(body.play_priority))
    return _atc_status()


class AtcTestBody(BaseModel):
    work_url: str = ""


@app.post("/api/admin/atc/test")
def admin_atc_test(body: AtcTestBody, request: Request):
    """测试连接：真实提交一个提取任务验证密钥可用，结果异步到。
    状态写 app_settings.atc_test_state（task_id + 提交耗时），由 GET 轮询推进。"""
    _require_admin(request)
    cfg = _atc_cfg()
    if not (cfg["key"] and cfg["secret"]):
        raise ApiError(400, "请先保存 API Key 与 Secret")
    work_url = body.work_url.strip() or "https://v.douyin.com/uc_Eukb0zUM/"
    t0 = time.time()
    try:
        resp = _atc_request("POST", "/video/extract",
                            {"workUrl": work_url, "taskType": "TEXT"}, cfg)
    except Exception as e:
        set_app_setting("atc_test_state", json.dumps(
            {"state": "failed", "error": f"{type(e).__name__}: {e}"}))
        raise ApiError(502, f"连接失败：{type(e).__name__}: {e}")
    ms = int((time.time() - t0) * 1000)
    if resp.get("code") != 200 or not resp.get("data"):
        set_app_setting("atc_test_state", json.dumps(
            {"state": "failed", "error": str(resp.get("msg") or resp)[:200]}))
        raise ApiError(502, f"任务提交被拒：{resp.get('msg') or '未知错误'}")
    set_app_setting("atc_test_state", json.dumps(
        {"state": "submitted", "task_id": str(resp["data"]), "ms": ms}))
    return {"ok": True, "task_id": str(resp["data"]), "ms": ms}


# ---- 分享页管理（含侵权下架）----

@app.get("/api/admin/shares")
def admin_shares(request: Request, limit: int = 100, q: str = ""):
    _require_admin(request)
    like = f"%{q}%"
    rows = db_exec(
        "SELECT id,item_id,kind,title,author,status,views,plays,downloads,cta_clicks,"
        "expires_at,created,owner_user_id,owner_ip FROM shares "
        "WHERE (?='' OR title LIKE ? OR author LIKE ? OR id=?) "
        "ORDER BY created DESC LIMIT ?",
        (q, like, like, q, max(1, min(500, limit))), "all")
    tot = db_exec("SELECT COUNT(*) c, COALESCE(SUM(views),0) v, COALESCE(SUM(plays),0) p, "
                  "COALESCE(SUM(cta_clicks),0) k FROM shares", (), "one")
    return {"shares": [dict(r) for r in rows],
            "total": tot["c"], "views": tot["v"], "plays": tot["p"], "cta": tot["k"]}


@app.post("/api/admin/shares/{sid}/takedown")
def admin_takedown(sid: str, request: Request):
    _require_admin(request)
    n = db_exec("UPDATE shares SET status='takedown' WHERE id=?", (sid,), "rowcount")
    if not n:
        raise ApiError(404, "分享页不存在")
    return {"ok": True}


@app.delete("/api/admin/shares/{sid}")
def admin_del_share(sid: str, request: Request):
    _require_admin(request)
    db_exec("DELETE FROM shares WHERE id=?", (sid,))
    return {"ok": True}


@app.get("/api/admin/play-stats")
def admin_play_stats(request: Request, hours: int = 24, limit: int = 60):
    """播放诊断看板：看清「微信内哪些视频能播、走的哪条线路、失败在哪一步」。

    所有环境（含微信）播放链路均为 dy1 → dy2 → proxy，抖音直连优先、同源代理兜底。
    每条线路的 try/ok/fail 都会上报，带宽分析需要按微信内/外分别观察。"""
    _require_admin(request)
    hours = max(1, min(24 * 30, hours))
    limit = max(1, min(300, limit))
    since = int(time.time()) - hours * 3600

    # 按 微信内/外 × 线路 汇总成功率。排除 giveup —— 它是「三条都挂了」的汇总事件，
    # 没有 source，混进来会多出一行空线路并把失败数重复计一遍。
    rows = db_exec(
        "SELECT wechat, COALESCE(source,'') source, kind, COUNT(*) n, "
        "CAST(AVG(ms) AS INTEGER) avg_ms FROM share_events "
        "WHERE ts>=? AND kind IN ('play_ok','play_fail') AND COALESCE(stage,'')<>'giveup' "
        "GROUP BY wechat, source, kind", (since,), "all")
    agg: dict = {}
    for r in rows:
        k = (r["wechat"], r["source"])
        cur = agg.setdefault(k, {"wechat": r["wechat"], "source": r["source"],
                                 "ok": 0, "fail": 0, "ok_ms": 0})
        if r["kind"] == "play_ok":
            cur["ok"] = r["n"]
            cur["ok_ms"] = r["avg_ms"] or 0
        else:
            cur["fail"] = r["n"]
    lines = sorted(agg.values(), key=lambda x: (-x["wechat"], x["source"]))
    for x in lines:
        t = x["ok"] + x["fail"]
        x["total"] = t
        x["rate"] = round(x["ok"] * 100.0 / t, 1) if t else 0.0

    # 彻底放弃（三条线路全挂）的次数，微信内外分开
    gv = db_exec("SELECT wechat, COUNT(*) n FROM share_events "
                 "WHERE ts>=? AND kind='play_fail' AND stage='giveup' "
                 "GROUP BY wechat", (since,), "all")
    giveup = {("wechat" if r["wechat"] else "other"): r["n"] for r in gv}

    # 按作品维度：微信内失败最多的分享页（就是「哪些视频播不了」）。
    # 同样排除 giveup 汇总事件，并单列 giveup 次数 —— 那才是「彻底放不出来」的次数。
    bad = db_exec(
        "SELECT e.sid, COALESCE(s.title,'(已删除)') title, COALESCE(s.kind,'') vkind, "
        "SUM(CASE WHEN e.kind='play_ok' THEN 1 ELSE 0 END) ok, "
        "SUM(CASE WHEN e.kind='play_fail' AND COALESCE(e.stage,'')<>'giveup' THEN 1 ELSE 0 END) fail, "
        "SUM(CASE WHEN COALESCE(e.stage,'')='giveup' THEN 1 ELSE 0 END) giveup "
        "FROM share_events e LEFT JOIN shares s ON s.id=e.sid "
        "WHERE e.ts>=? AND e.wechat=1 AND e.kind IN ('play_ok','play_fail') "
        "GROUP BY e.sid HAVING fail>0 ORDER BY giveup DESC, fail DESC, ok ASC LIMIT ?",
        (since, limit), "all")

    # 播放尝试全量日志：每条线路的 尝试(play_try)/成功/失败 明细，带粗粒度 UA
    # 定位是哪个机型/内核播不了。play_try 为 v1.11 新增，旧数据只有 ok/fail。
    log = db_exec(
        "SELECT ts,sid,kind,COALESCE(source,'') source,COALESCE(stage,'') stage,"
        "COALESCE(detail,'') detail,ms,wechat,ua FROM share_events "
        "WHERE ts>=? AND kind IN ('play_try','play_ok','play_fail') "
        "ORDER BY ts DESC LIMIT ?",
        (since, limit), "all")

    return {"hours": hours, "lines": lines, "giveup": giveup,
            "bad": [dict(r) for r in bad], "log": [dict(r) for r in log]}


@app.get("/api/admin/play-logs")
def admin_play_logs(request: Request, page: int = 1, size: int = 20,
                    result: str = "", wechat: str = "", sid: str = ""):
    """播放请求日志（服务端分页）：记录每次播放的浏览器环境、线路、成败，
    以及失败后将重试的下一条线路（next_src）。支撑「微信内是否播放成功」的逐条核查。
    只存粗粒度环境与线路名，不存完整媒体签名地址（隐私红线）。"""
    _require_admin(request)
    page = max(1, int(page))
    size = max(1, min(100, int(size)))
    where = ["e.kind IN ('play_try','play_ok','play_fail')"]
    params: list = []
    if result in ("try", "ok", "fail"):
        where.append("e.kind=?")
        params.append("play_" + result)
    if wechat in ("0", "1"):
        where.append("e.wechat=?")
        params.append(int(wechat))
    if sid.strip():
        where.append("e.sid=?")
        params.append(sid.strip()[:20])
    cond = " AND ".join(where)
    total = db_exec(f"SELECT COUNT(*) FROM share_events e WHERE {cond}",
                    tuple(params), "one")[0]
    rows = db_exec(
        "SELECT e.ts,e.sid,COALESCE(s.title,'(已删除)') title,e.kind,"
        "COALESCE(e.source,'') source,COALESCE(e.stage,'') stage,"
        "COALESCE(e.detail,'') detail,COALESCE(e.next_src,'') next_src,"
        "e.ms,e.wechat,e.ua FROM share_events e "
        "LEFT JOIN shares s ON s.id=e.sid "
        f"WHERE {cond} ORDER BY e.ts DESC, e.id DESC LIMIT ? OFFSET ?",
        (*params, size, (page - 1) * size), "all")
    return {"rows": [dict(r) for r in rows], "total": total,
            "page": page, "size": size,
            "pages": max(1, (total + size - 1) // size)}


@app.get("/api/admin/traffic-stats")
def admin_traffic_stats(request: Request, days: int = 30):
    """转发流量统计：/api/video 同源流式转发的按天字节/次数聚合。

    scope=play 是播放兜底线路、scope=download 是视频下载；
    浏览器直连抖音 CDN 的流量不经过本服务器，无法也不需要统计。"""
    _require_admin(request)
    days = max(1, min(365, days))
    _flush_media_traffic()          # 先把内存里的增量落库，保证读到最新
    since = _today() - days + 1
    rows = db_exec("SELECT day,scope,requests,bytes FROM media_traffic "
                   "WHERE day>=? ORDER BY day DESC", (since,), "all")
    daily: dict = {}
    for r in rows:
        d = daily.setdefault(r["day"], {
            "day": r["day"],
            "date": time.strftime("%Y-%m-%d", time.gmtime(r["day"] * 86400)),
            "play_requests": 0, "play_bytes": 0,
            "download_requests": 0, "download_bytes": 0})
        if r["scope"] == "download":
            d["download_requests"] += r["requests"] or 0
            d["download_bytes"] += r["bytes"] or 0
        else:
            d["play_requests"] += r["requests"] or 0
            d["play_bytes"] += r["bytes"] or 0
    out = sorted(daily.values(), key=lambda x: -x["day"])
    totals = {k: sum(d[k] for d in out) for k in
              ("play_requests", "play_bytes", "download_requests", "download_bytes")}
    return {"days": days, "daily": out, "totals": totals}


@app.get("/api/admin/reports")
def admin_reports(request: Request, limit: int = 100):
    _require_admin(request)
    rows = db_exec("SELECT * FROM reports ORDER BY ts DESC LIMIT ?",
                   (max(1, min(500, limit)),), "all")
    return {"reports": [dict(r) for r in rows]}


@app.post("/api/admin/reports/{rid}/handle")
def admin_handle_report(rid: int, request: Request):
    _require_admin(request)
    db_exec("UPDATE reports SET handled=1 WHERE id=?", (rid,))
    return {"ok": True}


@app.get("/api/admin/share-config")
def admin_share_config(request: Request):
    _require_admin(request)
    off = _domains_off()
    return {
        "primary_domain": app_setting("share_primary_domain", ""),
        "domains": [{"url": d, "enabled": d not in off} for d in SHARE_DOMAINS],
        "env_hint": "主分享域名在此填写即时生效（无需重启）；额外的备用域名池仍通过环境变量 "
                    "SHARE_DOMAINS 配置（逗号分隔，重启生效）。",
        "wx": {"appid": app_setting("wx_appid"),
               "configured": bool(app_setting("wx_appid") and app_setting("wx_secret"))},
    }


class ShareConfigBody(BaseModel):
    wx_appid: Optional[str] = None
    wx_secret: Optional[str] = None
    toggle_domain: Optional[str] = None
    primary_domain: Optional[str] = None


@app.post("/api/admin/share-config")
def admin_set_share_config(body: ShareConfigBody, request: Request):
    _require_admin(request)
    if body.primary_domain is not None:
        d = body.primary_domain.strip().rstrip("/")
        if d and not d.startswith(("http://", "https://")):
            d = "https://" + d                 # 容错：只填了域名就补 https
        set_app_setting("share_primary_domain", d)
    if body.wx_appid is not None:
        set_app_setting("wx_appid", body.wx_appid.strip())
    if body.wx_secret is not None and body.wx_secret.strip():
        set_app_setting("wx_secret", body.wx_secret.strip())
        set_app_setting("wx_ticket", "")          # 换了密钥，缓存的 ticket 立刻作废
        set_app_setting("wx_ticket_exp", 0)
    if body.toggle_domain:
        off = _domains_off()
        off.symmetric_difference_update({body.toggle_domain})
        set_app_setting("share_domains_off", json.dumps(sorted(off)))
    return admin_share_config(request)


# ---- 内置 mihomo（机场加速）----

@app.get("/api/admin/mihomo")
def admin_mihomo(request: Request):
    _require_admin(request)
    return mihomo_mgr.status()


class MihomoBody(BaseModel):
    sub_url: Optional[str] = None        # 填/改订阅 URL（空串=停用并清除）
    action: Optional[str] = None         # start / stop / restart


@app.post("/api/admin/mihomo")
def admin_set_mihomo(body: MihomoBody, request: Request):
    _require_admin(request)
    if body.sub_url is not None:
        set_app_setting("mihomo_sub_url", body.sub_url.strip())
        mihomo_mgr.reload()               # 重建配置并按有无订阅启停
    elif body.action == "stop":
        set_app_setting("mihomo_sub_url", "")
        mihomo_mgr.reload()
    elif body.action in ("start", "restart"):
        mihomo_mgr.reload()
    return mihomo_mgr.status()


# ---- 用户自助 API 密钥（登录后）----

@app.get("/api/keys")
def user_list_keys(request: Request):
    u = current_user(request)
    if not u:
        raise ApiError(401, "请先登录")
    return {"keys": list_api_keys(u["id"]), "price_cents": api_price_cents()}


@app.post("/api/keys")
def user_create_key(body: NewKeyBody, request: Request):
    u = current_user(request)
    if not u:
        raise ApiError(401, "请先登录")
    if len(list_api_keys(u["id"])) >= 10:
        raise ApiError(400, "每个账号最多 10 个密钥")
    return create_api_key(u["id"], body.name)


@app.post("/api/keys/revoke")
def user_revoke_key(body: KeyBody, request: Request):
    u = current_user(request)
    if not u:
        raise ApiError(401, "请先登录")
    if not revoke_api_key(body.key, u["id"]):
        raise ApiError(404, "密钥不存在或无权删除")
    return {"ok": True}


# ---------------------------------------------------------------- 后台健康检查

def _health_loop():
    """守护线程：按间隔并发测试启用中（及被自动禁用）的代理，自动禁用/自愈。"""
    while True:
        interval = max(1, int(proxy_mgr.settings.get("health_interval_min", 10)))
        for _ in range(interval * 60):
            time.sleep(1)
        if not proxy_mgr.settings.get("auto_health", True):
            continue
        targets = [p for p in list(proxy_mgr.proxies) if p["enabled"] or p.get("auto_off")]
        if targets:
            _probe_all(targets, proxy_mgr.settings.get("test_reach_douyin", True))


@app.on_event("startup")
def _start_health():
    # 崩溃遗留的网页配额先退款；API 作业先恢复/对账，再清理过期明细。
    # cleanup 放在 legacy recovery 后，避免提前删掉旧 api_logs 导致无法重建已结算项。
    _refund_stale_quota_reservations()
    _prepare_api_jobs()
    _cleanup_retained_data(force=True)
    # 所有可能阻断 startup 的迁移/清理完成后才启动非 daemon worker，避免半启动悬挂。
    _start_api_job_workers(prepared=True)
    threading.Thread(target=_health_loop, daemon=True).start()
    threading.Thread(target=mihomo_mgr.supervise, daemon=True).start()
    threading.Thread(target=_atc_worker, daemon=True).start()


@app.on_event("shutdown")
def _stop_mihomo():
    _stop_api_job_workers()
    # 内存里的转发流量计数落库，重启不丢
    try:
        _flush_media_traffic()
    except Exception:
        pass
    # 关服务时杀掉内置 mihomo 子进程，避免留下孤儿进程占用端口
    try:
        mihomo_mgr.stop()
    except Exception:
        pass


# ---------------------------------------------------------------- 页面 + SEO

def _origin(request: Request) -> str:
    """反代下取真实站点 origin，用于 canonical / og:url / sitemap。"""
    proto = request.headers.get("x-forwarded-proto") or request.url.scheme
    host = request.headers.get("x-forwarded-host") or request.headers.get("host") or "localhost"
    return f"{proto}://{host}"


SUPPORTED_LANGS = {"zh": "zh-CN", "en": "en"}


def _pick_lang(request: Request) -> str:
    q = (request.query_params.get("lang") or "").lower()
    if q in SUPPORTED_LANGS:
        return q
    c = (request.cookies.get("lang") or "").lower()
    if c in SUPPORTED_LANGS:
        return c
    al = (request.headers.get("accept-language") or "").lower()
    return "zh" if al.startswith("zh") or not al else ("en" if al[:2] not in ("zh",) else "zh")


# 落地页 SEO 覆盖：键为路径，内容与该页可见文案保持一致（FAQ 与页面对应）
_LANDING_SEO = {
    "/transcript": {
        "meta": {
            "zh": {
                "title": "抖音文案提取 · 视频语音转文字 — 抖音无水印下载器",
                "desc": "把抖音视频里的语音自动转成文字：注册用户每天免费提取，标题、正文、口播全文一次拿全，适合素材收集与内容分析。开源可审查、不保存媒体文件。",
                "kw": "抖音文案提取,视频转文字,抖音语音转文字,口播文案提取,视频文案提取,douyin transcript,抖音字幕提取",
                "site": "抖音无水印下载器",
                "ogt": "抖音文案提取 · 视频语音一键转文字",
                "ogd": "粘贴抖音链接，自动把视频语音转成完整文字。注册用户每天免费提取，不保存媒体文件。",
                "locale": "zh_CN",
            },
            "en": {
                "title": "Douyin Transcript Extractor — Speech to Text, Free Daily",
                "desc": "Turn the speech in any Douyin video into text: title, caption and full transcript in one go. Free daily quota for signed-in users. Open source, no media-file storage.",
                "kw": "douyin transcript,video to text,douyin speech to text,extract video caption,douyin subtitle extractor",
                "site": "Douyin Downloader",
                "ogt": "Douyin Transcript Extractor — Speech to Text",
                "ogd": "Paste a Douyin link and get the full transcript of its speech. Free daily quota, no media-file storage.",
                "locale": "en_US",
            },
        },
        "ld": {
            "zh": {
                "app_desc": "抖音视频文案提取工具：粘贴链接即可把视频语音转成文字，同时获得标题与正文。注册用户每天免费提取，同一视频只计一次；本站不保存媒体文件。",
                "features": ["抖音视频语音转文字", "标题与正文提取", "音频试听", "注册用户每日免费", "结果可复制", "不保存媒体文件"],
                "faq": [
                    ("什么是抖音文案提取？", "把视频里的语音自动转成文字，同时保留作品的标题与正文，适合收集口播文案、做内容分析。该功能由可选的增强线路处理，异步任务通常 1–3 分钟完成。"),
                    ("文案提取收费吗？", "注册用户每天有免费提取次数（默认 5 次，以页面显示为准）。同一视频全站只提取一次，再次打开命中缓存不重复扣次。"),
                    ("我的链接会发给第三方吗？", "仅在管理员开启增强线路后，你请求提取的作品链接会提交给第三方服务（AnyToCopy）处理，用于获取语音转文字结果与播放地址；未开启则不发送。本站只保存处理结果元数据，不保存视频文件。"),
                    ("提取要等多久？", "通常 1–3 分钟，取决于视频时长，短视频更快。提交后可以离开页面，回来后重新打开开关即可查看结果。"),
                ],
                "howto": ("如何提取抖音视频文案", [
                    ("粘贴并解析", "把抖音分享链接粘贴到本站输入框，点击解析。"),
                    ("打开「获取文案」开关", "登录后在解析结果卡上打开「获取文案（语音转文字）」开关。未登录时点开关会引导你先登录。"),
                    ("等待并复制", "提取通常 1–3 分钟，完成后可一键复制全文或试听音频。")]),
            },
            "en": {
                "app_desc": "A Douyin transcript extractor: paste a link to turn a video's speech into text, together with its title and caption. Signed-in users get a free daily quota; only one extraction per video site-wide. No media files are stored.",
                "features": ["Douyin speech to text", "Title & caption extraction", "Audio preview", "Free daily quota", "One-click copy", "No media-file storage"],
                "faq": [
                    ("What is Douyin transcript extraction?", "It turns a video's speech into text and keeps the post's title and caption — built for collecting scripts and content analysis. Powered by an optional enhancement route as an async job, usually done in 1–3 minutes."),
                    ("Is transcript extraction free?", "Signed-in users get a free daily quota (5/day by default, as shown on the page). Each video is extracted only once site-wide; reopening a cached result costs nothing."),
                    ("Is my link sent to a third party?", "Only when the enhancement route is enabled by the site operator: the link you ask to transcribe is submitted to a third-party service (AnyToCopy) to produce the transcript and a playback URL; otherwise nothing is sent. Only result metadata is kept — never the video file."),
                    ("How long does it take?", "Usually 1–3 minutes depending on video length; short clips are faster. You can leave after submitting — reopen the toggle later to see the result."),
                ],
                "howto": ("How to extract the transcript of a Douyin video", [
                    ("Paste and parse", "Paste the Douyin share link into the input box and click Parse."),
                    ("Turn on the transcript toggle", "After signing in, switch on “Transcript (speech to text)” on the result card. Signed-out users are guided to sign in first."),
                    ("Wait and copy", "Extraction usually takes 1–3 minutes. When done, copy the full text or listen to the audio.")]),
            },
        },
    },
}


def _seo_head(lang: str, origin: str, path = "/") -> str:
    """按语言生成整段 SEO 头（title/description/OG/Twitter/hreflang/JSON-LD）。"""
    zh = lang == "zh"
    base = f"{origin}{path}"
    canon = base if zh else f"{base}?lang=en"
    meta = {
        "zh": {
            "title": "抖音无水印下载器 · 粘贴链接即下 · 一键分享给微信好友",
            "desc": "免费的抖音无水印下载与分享工具：粘贴分享链接，即可在线预览、下载抖音视频与图集的无水印原片，或一键生成分享页发给微信好友——对方点开就能看，无需安装抖音 App。开源可审查、不保存媒体文件、站内账号可选、永不接广告。",
            "kw": "抖音下载,抖音无水印下载,抖音视频下载,抖音去水印,douyin downloader,抖音图集下载,抖音解析,无水印下载器,抖音下载器在线,抖音API,抖音视频分享,抖音怎么分享到微信,抖音视频发微信,抖音分享链接生成,抖音视频免App观看,微信打开抖音视频",
            "site": "抖音无水印下载器",
            "ogt": "抖音无水印下载器 · 下载原片 + 一键分享给微信好友",
            "ogd": "粘贴抖音分享链接，可靠下载无水印原片；还能一键生成分享页发到微信，好友点开即看、无需装 App。开源可审查、不保存媒体文件、账号可选、永不接广告。",
            "locale": "zh_CN",
        },
        "en": {
            "title": "Douyin Downloader & Share — No Watermark, Free & Open Source",
            "desc": "Free Douyin (Chinese TikTok) no-watermark downloader and share tool. Paste a link to download original videos and galleries, or create a WeChat-friendly share page. Open source and auditable, no media-file storage, optional site account, and no ads.",
            "kw": "douyin downloader,douyin video download,no watermark,tiktok downloader,save douyin video,douyin photo download,douyin api,open source downloader,share douyin video,send douyin video to wechat,douyin share link,watch douyin without app",
            "site": "Douyin Downloader",
            "ogt": "Douyin Downloader — Download No-Watermark Originals & Share to WeChat",
            "ogd": "Download a no-watermark Douyin original or create a share page friends can watch on WeChat. Open source, no media-file storage, optional account, and no ads.",
            "locale": "en_US",
        },
    }[lang]

    ld = {
        "zh": {
            "app_desc": "无需登录抖音账号的抖音视频与图集无水印下载、分享工具：粘贴链接即可预览与下载，也能生成微信友好的分享页。播放直连、视频签名流式下载，开源可审查，不保存媒体文件，站内账号可选，并提供开发者 API。",
            "features": ["抖音视频无水印下载", "抖音图集下载", "一键生成分享页", "分享到微信生成卡片", "免 App 观看抖音视频", "分享海报生成", "在线预览播放", "批量解析", "播放直连且媒体不落地", "开发者 API"],
            "faq": [
                ("这个抖音下载器会处理和保留哪些数据？", "前端代码开源可审查，本站不保存视频或图片文件。浏览器使用 30 天随机第一方匿名 ID；免费额度、防滥用和播放诊断会处理用途化网络/匿名 ID 摘要、粗粒度浏览器信息及事件。相关明细及 API 任务结果的保留期最多设为 30 天，到期后由每 5 分钟运行的任务删除。站内账号可选，注册会保存邮箱与加盐密码哈希；浏览器直连播放或图片时抖音会收到请求方网络与浏览器信息，视频下载则由本站代理流式转发。仅当管理员开启「增强线路」且用户主动使用文案提取时，对应作品链接会提交给第三方服务（AnyToCopy）处理；未开启或不使用则不发送。"),
                ("怎么把抖音视频分享到微信？发出去是卡片还是链接？", "解析后点「生成分享页」得到一条链接。想让好友收到带封面标题的卡片，要在微信里打开这个页面，再点右上角 ··· →「发送给朋友」，这样转发出去才是卡片。若只是复制链接粘贴到聊天窗口，微信不会把网址展开成卡片，会显示为一条普通网址（这是微信的机制，对任何网站都一样）。两种方式好友点开都能直接观看无水印原片，无需安装抖音 App、不用复制口令跳转。"),
                ("分享给朋友后，对方需要装抖音 App 吗？链接会过期吗？", "不需要装任何 App，用微信内置浏览器点开就能看。分享页匿名有效期 7 天、登录后 30 天；页面只保存作品的标题封面等信息，不存储任何视频文件，版权仍归原作者。你也可以生成带二维码的分享海报，长按保存后发朋友圈。"),
                ("需要登录或安装软件吗？", "无需登录抖音账号或安装软件。基础解析无需注册本站账号；API 控制台等账号功能需要登录。"),
                ("下载的抖音视频有水印吗？", "没有水印。下载的是无水印原片，也不会加入本站自己的二次水印。"),
                ("支持图集（图片作品）下载吗？", "支持。图集作品会自动识别，可逐张下载原图，也可批量下载。"),
                ("有没有 API 可以批量调用？", "有。登录后可在 API 控制台生成密钥，通过异步接口批量提交链接并查询结果，按次计费。"),
                ("怎么提取抖音视频的文案（语音转文字）？", "解析后打开结果卡上的「获取文案（语音转文字）」开关即可自动提取，通常 1–3 分钟完成，可复制全文或试听音频。该功能需要登录，注册用户每天有免费提取次数；同一视频全站只提取一次，命中缓存不重复扣次。"),
            ],
            "howto": ("如何下载抖音无水印视频并分享给微信好友", [
                ("复制分享链接", "在抖音 App 里点分享，复制作品链接或整段分享文案。"),
                ("粘贴并解析", "把链接粘贴到本站输入框，点击解析。"),
                ("下载或生成分享页", "一键下载无水印原片；或生成分享页发到微信。直接粘贴显示普通网址；在微信内打开页面后从右上角转发，才会显示带封面标题的卡片。")]),
        },
        "en": {
            "app_desc": "A no-watermark Douyin video and gallery downloader and sharing tool that requires no Douyin login. Playback is browser-direct while video downloads use a signed streaming route. It is open source and auditable, stores no media files, offers an optional site account, and includes a developer API.",
            "features": ["Douyin no-watermark video download", "Photo gallery download", "One-click share page", "WeChat share card", "Watch Douyin without the app", "Share poster generator", "In-browser preview", "Batch parsing", "Direct playback with no media-file storage", "Developer API"],
            "faq": [
                ("What data does this downloader process and retain?", "The front end is open source and auditable, and the service stores no video or image files. A random first-party anonymous ID lasts 30 days. Purpose-specific network and anonymous-ID digests, coarse browser details, quota or diagnostic events, API jobs and results have a configurable 1–30 day retention period; expired records are removed by a cleanup task that runs every five minutes. A site account is optional and stores an email address and salted password hash. Direct playback and image requests disclose browser network information to Douyin; video downloads are streamed through the site's proxy. Only when the operator enables the enhancement route and the user actively uses transcript extraction is the corresponding video link submitted to a third-party service (AnyToCopy); otherwise nothing is sent."),
                ("How do I share a Douyin video to WeChat? Does it show as a card or a plain link?", "Create a share page after parsing. Pasting its URL into a chat produces a plain link. To send a card with a cover and title, open the page inside WeChat and forward it from the top-right menu. Either form opens without the Douyin app."),
                ("Do my friends need the Douyin app? Do share links expire?", "No app is needed — the page opens right in WeChat's built-in browser. Share pages last 7 days anonymously and 30 days when signed in. The page only stores the post's title and cover; no video files are stored and copyright stays with the original creator. You can also generate a poster with a QR code to save and post to Moments."),
                ("Do I need to log in or install anything?", "No Douyin login, app, or extension is required. Basic parsing needs no site account; account features such as the API console require sign-in."),
                ("Do downloaded videos have a watermark?", "No. You get the original video with no watermark, and we never add our own."),
                ("Can I download photo galleries (image posts)?", "Yes. Image posts are detected automatically; download each original image or batch-download them."),
                ("Is there an API for bulk use?", "Yes. After signing in you can create an API key in the console, submit links in bulk via the async API and poll for results, billed per request."),
                ("How do I extract the transcript of a Douyin video?", "After parsing, switch on the transcript toggle on the result card — extraction usually takes 1–3 minutes and the full text can be copied. Sign-in is required, with a free daily quota; each video is extracted only once site-wide."),
            ],
            "howto": ("How to download a Douyin video without watermark and share it on WeChat", [
                ("Copy the share link", "In the Douyin app tap Share and copy the link or the whole share text."),
                ("Paste and parse", "Paste the link into the input box and click Parse."),
                ("Download or create a share page", "Download the original file, or create a share page for WeChat. A pasted URL remains a plain link; open the page in WeChat and forward it from the top-right menu to send a card.")]),
        },
    }[lang]

    # 落地页（如 /transcript）用独立文案覆盖默认首页 SEO
    landing = _LANDING_SEO.get(path)
    if landing:
        meta = landing["meta"][lang]
        ld = landing["ld"][lang]

    org_id = f"{origin}/#org"
    site_id = f"{origin}/#website"
    graph = [
        {"@type": "Organization", "@id": org_id, "name": meta["site"],
         "url": f"{origin}/", "logo": f"{origin}/og.png"},
        {"@type": "WebSite", "@id": site_id, "name": meta["site"],
         "url": f"{origin}/", "publisher": {"@id": org_id},
         "inLanguage": ["zh-CN", "en"]},
        {"@type": "WebApplication", "name": meta["site"], "url": f"{origin}/",
         "applicationCategory": "MultimediaApplication", "operatingSystem": "All",
         "isPartOf": {"@id": site_id}, "publisher": {"@id": org_id},
         "offers": {"@type": "Offer", "price": "0", "priceCurrency": "CNY"},
         "description": ld["app_desc"], "featureList": ld["features"]},
        {"@type": "FAQPage", "mainEntity": [
            {"@type": "Question", "name": q,
             "acceptedAnswer": {"@type": "Answer", "text": a}} for q, a in ld["faq"]]},
        {"@type": "HowTo", "name": ld["howto"][0], "step": [
            {"@type": "HowToStep", "position": i + 1, "name": n, "text": t}
            for i, (n, t) in enumerate(ld["howto"][1])]},
    ]
    jsonld = json.dumps({"@context": "https://schema.org", "@graph": graph}, ensure_ascii=False)

    def esc(s):
        return s.replace("&", "&amp;").replace('"', "&quot;").replace("<", "&lt;")

    return f'''<title>{esc(meta["title"])}</title>
<meta name="app-version" content="{APP_VERSION}">
<meta name="description" content="{esc(meta["desc"])}">
<meta name="keywords" content="{esc(meta["kw"])}">
<meta name="robots" content="index,follow,max-image-preview:large">
<meta name="theme-color" content="#0E1013" media="(prefers-color-scheme:dark)">
<meta name="theme-color" content="#FFFFFF" media="(prefers-color-scheme:light)">
<meta name="author" content="{esc(meta["site"])}">
<meta name="application-name" content="{esc(meta["site"])}">
<meta name="apple-mobile-web-app-title" content="{esc(meta["site"])}">
<meta name="apple-mobile-web-app-capable" content="yes">
<link rel="preconnect" href="https://aweme.snssdk.com" crossorigin>
<link rel="dns-prefetch" href="https://aweme.snssdk.com">
<link rel="canonical" href="{canon}">
<link rel="alternate" hreflang="zh-CN" href="{base}">
<link rel="alternate" hreflang="en" href="{base}?lang=en">
<link rel="alternate" hreflang="x-default" href="{base}">
<meta property="og:type" content="website">
<meta property="og:site_name" content="{esc(meta["site"])}">
<meta property="og:title" content="{esc(meta["ogt"])}">
<meta property="og:description" content="{esc(meta["ogd"])}">
<meta property="og:url" content="{canon}">
<meta property="og:image" content="{origin}/og.png">
<meta property="og:image:type" content="image/png">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:image:alt" content="{esc(meta["site"])}">
<meta property="og:locale" content="{meta["locale"]}">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="{esc(meta["ogt"])}">
<meta name="twitter:description" content="{esc(meta["ogd"])}">
<meta name="twitter:image" content="{origin}/og.png">
<script type="application/ld+json">{jsonld}</script>
<script>window.__LANG={lang!r};window.__ORIGIN={origin!r};</script>'''


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    log_pageview(request)
    lang = _pick_lang(request)
    origin = _origin(request)
    html = Path("static/index.html").read_text("utf-8")
    html = (html.replace("{{HTMLLANG}}", SUPPORTED_LANGS[lang])
                .replace("{{SEO_HEAD}}", _seo_head(lang, origin))
                .replace("{{ORIGIN}}", origin))
    resp = HTMLResponse(html)
    resp.headers["Cache-Control"] = "private, no-store"
    resp.headers["Pragma"] = "no-cache"
    resp.set_cookie("lang", lang, max_age=31536000, samesite="lax")
    return resp


def _share_head(view: Optional[dict], origin: str) -> str:
    """分享页的 per-share 头信息。**一律 noindex** —— 不收录他人作品内容。"""
    def esc(s):
        return (str(s or "").replace("&", "&amp;").replace('"', "&quot;")
                .replace("<", "&lt;").replace(">", "&gt;"))

    if not view or view["state"] != "ok":
        return ('<title>内容不可用 · 抖音分享</title>\n'
                '<meta name="robots" content="noindex,nofollow">\n'
                '<meta name="theme-color" content="#0E1013">')
    # 卡片大标题 = 抖音文案原文（与抖音里一模一样）；作者放进描述行。
    # 微信抓取网页 meta 生成卡片：title/og:title→标题，og:image→缩略图，description→摘要。
    # 卡片底部的"来源/抬头"由微信按域名自动填（域名或其绑定的公众号名称），网页无法自定义。
    title = (view["title"] or "抖音作品")[:60]
    author = view["author"] or "抖音创作者"
    desc = f"@{author} 的抖音作品 · 点开即可观看，无需安装 App"
    # 卡片图：抖音封面转成无签名 JPEG（webp 微信缩略图支持不稳定、签名 14 天过期）；
    # 兜底用 og.png 而非 og.svg —— 微信不渲染 SVG，会退化成无图的纯链接。
    cover = _card_cover(view["cover"]) or f"{origin}/og.png"
    return f'''<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}">
<meta name="robots" content="noindex,nofollow">
<meta name="theme-color" content="#0E1013">
<meta property="og:type" content="video.other">
<meta property="og:site_name" content="@{esc(author)}">
<meta property="og:title" content="{esc(title)}">
<meta property="og:description" content="{esc(desc)}">
<meta property="og:image" content="{esc(cover)}">
<meta property="og:image:type" content="image/jpeg">
<meta property="og:image:alt" content="{esc(title)}">
<meta property="og:url" content="{esc(view['url'])}">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="{esc(title)}">
<meta name="twitter:description" content="{esc(desc)}">
<meta name="twitter:image" content="{esc(cover)}">'''


@app.get("/s/{sid}", response_class=HTMLResponse)
def share_page(sid: str, request: Request):
    """分享页：服务端渲染，微信内可直接打开与播放。"""
    origin = _origin(request)
    row = db_exec("SELECT * FROM shares WHERE id=?", (sid,), "one")
    view = None
    if row:
        row = dict(row)
        # 图集直链会过期：超过刷新窗口就按 item_id 重新解析（视频靠 vid 重拼，无需刷新）
        should_refresh_note = (row["kind"] == "note"
                               and time.time() - (row["refreshed_at"] or 0)
                               > SHARE_REFRESH_TTL)
        should_repair_video = row["kind"] != "note" and not row["vid"]
        if (_share_state(row) == "ok"
                and (should_refresh_note or should_repair_video)):
            row = _refresh_share(row)
        view = _share_view(row, origin)
        _share_event(request, sid, "view")

    html = Path("static/share.html").read_text("utf-8")
    # 注入 <script> 前把 < 转义成 <，防止标题里的 </script> 打断脚本
    payload = json.dumps(view or {"state": "notfound", "sid": sid},
                         ensure_ascii=False).replace("<", "\\u003c")
    html = (html.replace("{{HTMLLANG}}", "zh-CN")
                .replace("{{SHARE_HEAD}}", _share_head(view, origin))
                .replace("{{ORIGIN}}", origin)
                .replace("{{WECHAT}}", "true" if _is_wechat(request) else "false")
                .replace("{{SHARE_DATA}}", payload))
    resp = HTMLResponse(html, status_code=200 if view else 404)
    resp.headers["Cache-Control"] = "private, no-store"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Vary"] = "User-Agent"
    return resp


@app.get("/api-docs", response_class=HTMLResponse)
def api_docs(request: Request):
    log_pageview(request)
    lang = _pick_lang(request)
    origin = _origin(request)
    html = Path("static/api-docs.html").read_text("utf-8")
    html = (html.replace("{{HTMLLANG}}", SUPPORTED_LANGS[lang])
                .replace("{{SEO_HEAD}}", _seo_head(lang, origin, "/api-docs"))
                .replace("{{ORIGIN}}", origin))
    resp = HTMLResponse(html)
    # no-cache = 每次请求都回源校验（配合 ETag 未变返回 304），部署后用户无需强刷
    resp.headers["Cache-Control"] = "no-cache"
    resp.set_cookie("lang", lang, max_age=31536000, samesite="lax")
    return resp


@app.get("/transcript", response_class=HTMLResponse)
def transcript_page(request: Request):
    """文案提取落地页（SEO）：功能本体在首页结果卡，本页负责被搜到。"""
    log_pageview(request)
    lang = _pick_lang(request)
    origin = _origin(request)
    html = Path("static/transcript.html").read_text("utf-8")
    html = (html.replace("{{HTMLLANG}}", SUPPORTED_LANGS[lang])
                .replace("{{SEO_HEAD}}", _seo_head(lang, origin, "/transcript"))
                .replace("{{ORIGIN}}", origin))
    resp = HTMLResponse(html)
    resp.headers["Cache-Control"] = "no-cache"
    resp.set_cookie("lang", lang, max_age=31536000, samesite="lax")
    return resp


@app.get("/api/quota")
def api_quota(request: Request):
    """前端查询今日剩余免费次数（含文案提取额度，供结果卡开关展示）。"""
    limit, used, remaining = quota_status(request)
    u = current_user(request)
    cfg = _atc_cfg()
    atc_on = cfg["enabled"] and cfg["transcript_enabled"]
    if u and atc_on:
        atc_limit, atc_used, atc_remaining = _atc_transcript_status(u["id"])
    else:
        # 匿名也返回真实每日上限（前端提示"登录后每天 N 次"要用），剩余恒 0
        atc_limit, atc_remaining = (cfg["transcript_daily"] if atc_on else 0), 0
    return {"limit": limit, "used": used, "remaining": remaining,
            "user_daily": FREE_USER_DAILY,
            "user": {"email": u["email"]} if u else None,
            "atc": {"enabled": atc_on,
                    "daily": atc_limit, "remaining": atc_remaining}}


# ---------------------------------------------------------------- 用户鉴权 API

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# 滑块漏斗观测（内存计数，重启清零；只看失败率趋势，不落库、不含个人标识）
_captcha_stats = {"load": 0, "ok": 0, "fail": 0}


@app.get("/api/auth/captcha")
def auth_captcha(request: Request):
    if not _captcha_rate_ok(_client_ip(request)):        # 防验证码 CPU-DoS
        raise ApiError(429, "操作过于频繁，请稍后再试")
    _captcha_stats["load"] += 1
    return make_captcha(request)


class CaptchaBody(BaseModel):
    cid: str = ""
    x: float = -1
    trajectory: list = []
    nonce: str = ""


@app.post("/api/auth/captcha/verify")
def auth_captcha_verify(body: CaptchaBody, request: Request):
    """滑块校验独立成步。通过后返回一次性通行令牌，注册/登录必须携带它。"""
    ok, err = verify_captcha(body.cid, body.x, body.trajectory, body.nonce, request)
    _captcha_stats["ok" if ok else "fail"] += 1
    if not ok:
        raise ApiError(400, err)
    return {"ok": True, "pass_token": issue_pass(request)}


@app.get("/api/auth/me")
def auth_me(request: Request):
    u = current_user(request)
    if not u:
        return {"user": None}
    return {"user": {"email": u["email"], "id": u["id"], "created_at": u["created_at"]}}


class RegisterBody(BaseModel):
    email: str
    password: str
    pass_token: str = ""      # 滑块通过后签发的一次性令牌（缺它必拒）
    hp: str = ""              # 蜜罐字段，正常用户为空


def _do_auth_guard(request: Request, body: RegisterBody):
    if body.hp:                                     # 蜜罐命中 → 机器人
        raise ApiError(400, "验证失败")
    if not _auth_rate_ok(_client_ip(request)):
        raise ApiError(429, "操作过于频繁，请一小时后再试")
    if not consume_pass(body.pass_token, request):  # 必须先过滑块拿到令牌
        raise ApiError(400, "请先完成滑块验证（验证已失效，请重试）")


def _issue_session(uid: int) -> JSONResponse:
    tok = _new_user_session(uid)
    resp = JSONResponse({"ok": True})
    resp.set_cookie("sess", tok, httponly=True, samesite="lax",
                    secure=COOKIE_SECURE, max_age=USER_SESSION_TTL)
    return resp


@app.post("/api/auth/register")
def auth_register(body: RegisterBody, request: Request):
    email = (body.email or "").strip().lower()
    if not _EMAIL_RE.match(email):
        raise ApiError(400, "邮箱格式不正确")
    if len(body.password or "") < 6:
        raise ApiError(400, "密码至少 6 位")
    _do_auth_guard(request, body)
    if db_exec("SELECT id FROM users WHERE email=?", (email,), "one"):
        raise ApiError(409, "该邮箱已注册，请直接登录")
    salt, h = hash_pw(body.password)
    uid = db_exec("INSERT INTO users(email,pw_salt,pw_hash,created_at,last_login,reg_ip) "
                  "VALUES(?,?,?,?,?,?)",
                  (email, salt, h, int(time.time()), int(time.time()), ""))
    return _issue_session(uid)


@app.post("/api/auth/login")
def auth_login(body: RegisterBody, request: Request):
    email = (body.email or "").strip().lower()
    _do_auth_guard(request, body)
    row = db_exec("SELECT * FROM users WHERE email=?", (email,), "one")
    if not row or not verify_pw(body.password or "", row["pw_salt"], row["pw_hash"]):
        raise ApiError(403, "邮箱或密码错误")
    if row["disabled"]:
        raise ApiError(403, "该账号已被停用")
    db_exec("UPDATE users SET last_login=? WHERE id=?", (int(time.time()), row["id"]))
    return _issue_session(row["id"])


@app.post("/api/auth/logout")
def auth_logout(request: Request):
    tok = request.cookies.get("sess", "")
    _user_sessions.pop(tok, None)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("sess")
    return resp


@app.get("/api-console")
def api_console():
    # FileResponse 自带 ETag/Last-Modified；no-cache 强制每次回源校验，部署即生效
    return FileResponse("static/api-console.html",
                        headers={"Cache-Control": "no-cache"})


@app.get("/admin_d")
def admin_page():
    return FileResponse("static/admin.html",
                        headers={"Cache-Control": "no-cache"})


@app.get("/robots.txt", response_class=PlainTextResponse)
def robots(request: Request):
    # /s/ 是用户生成的他人作品分享页，一律不收录（详见 docs/分享页功能规划.md §3.1）
    o = _origin(request)
    rules = "Allow: /\nDisallow: /admin_d\nDisallow: /api/\nDisallow: /s/\n"
    # 生成式引擎（GEO）：显式放行主流 AI 抓取器，规则同普通爬虫
    ai_bots = ["GPTBot", "OAI-SearchBot", "ChatGPT-User", "PerplexityBot",
               "ClaudeBot", "Claude-Web", "Google-Extended", "Applebot-Extended", "CCBot"]
    blocks = [f"User-agent: *\n{rules}"] + [f"User-agent: {b}\n{rules}" for b in ai_bots]
    return ("\n".join(blocks) + f"\nLLM: {o}/llms.txt\nSitemap: {o}/sitemap.xml\n")


@app.get("/llms.txt", response_class=PlainTextResponse)
def llms_txt(request: Request):
    # 面向大模型/生成式引擎的站点说明（llmstxt.org 约定），帮助其准确引用本站
    o = _origin(request)
    return f"""# 抖音无水印下载器（Douyin Downloader）

> 免费、开源的抖音视频与图集**下载 + 分享**工具。粘贴抖音分享链接即可在线预览并下载无水印原片；也可生成分享页，好友无需安装抖音 App 即可观看。在微信聊天中直接粘贴会显示普通网址；在微信内打开页面后从右上角转发，才会显示带封面标题的卡片。普通浏览器播放与图片优先直连，视频下载和微信兼容播放使用带有效期授权的同源流式转发；本站不落地保存媒体文件。基础解析无需账号，站内账号与开发者 API 可选，永不接广告。

## 核心特性
- 抖音视频无水印下载：解析分享链接，去除水印，下载原始清晰度视频。
- 图集（图片作品）下载：自动识别多图作品，可逐张或批量下载原图。
- **一键生成分享页**：把抖音作品变成一个网页，发给朋友点开即看。
- **分享到微信显示为卡片**：在微信内打开分享页，点右上角 ··· 转发，好友收到带封面标题的卡片（直接粘贴网址则是纯链接，这是微信机制）。
- **免 App 观看**：接收方无需安装抖音、无需登录，微信内置浏览器直接播放。
- **分享海报**：前端合成带二维码的海报图，长按保存后可发朋友圈；链接被拦截时的传播兜底。
- 在线预览：下载前可直接在网页中预览播放。
- **文案提取（语音转文字）**：解析后打开「获取文案」开关，自动把视频语音转成完整文字；注册用户每天免费提取，同一视频全站只计一次。仅在管理员开启增强线路时，链接才会提交给第三方服务（AnyToCopy）处理；未开启则不发送。
- 可靠媒体链路：播放与图片直连优先，视频下载走同源签名流式转发，本站不缓存、不留存。
- 开源可审查、数据最小化：不保存媒体文件；免费额度和诊断只处理必要的用途化摘要、粗粒度环境与事件，保留期最多设为 30 天，到期后由每 5 分钟运行的任务删除；站内账号可选。
- 开发者 API：登录后于控制台生成密钥，异步批量提交链接、轮询结果，按次计费。

## 使用方式
1. 在抖音 App 点「分享 → 复制链接」，得到分享文案或短链。
2. 打开 {o}/ ，把链接粘贴进输入框并解析。
3. 在线预览后一键下载无水印原片；**或点「生成分享页」，把链接发给微信好友，对方点开即可观看**。

## 常见问答
- 会处理和保留哪些数据？——不保存视频或图片文件。浏览器使用 30 天随机匿名 ID；免费额度、防滥用和播放诊断会处理用途化网络/匿名 ID 摘要、粗粒度浏览器信息与事件。相关明细及 API 任务结果的保留期最多设为 30 天，到期后由每 5 分钟运行的任务删除。站内账号可选并保存邮箱与加盐密码哈希；媒体直连抖音时，抖音会收到请求方网络与浏览器信息。
- 怎么把抖音视频分享到微信？——解析后生成分享页。想发出带封面标题的卡片，需在微信里打开该页面，点右上角 ··· →「发送给朋友」；直接复制链接粘贴进聊天窗口不会展开成卡片，只显示为一条网址（微信机制，对所有网站一致）。两种方式好友点开都能直接观看无水印原片，无需装抖音 App。
- 对方需要装抖音 App 吗？会过期吗？——不需要装 App，微信内直接看；分享页匿名 7 天、登录后 30 天有效，仅保存标题封面等元数据，不存储视频文件。
- 需要登录或装软件吗？——无需登录抖音或安装软件；基础解析无需本站账号，API 控制台等账号功能需要登录。
- 下载的视频有水印吗？——没有，是无水印原片，也不加本站二次水印。
- 支持图集吗？——支持，自动识别并可批量下载原图。
- 有批量 API 吗？——有，登录后在 API 控制台生成密钥调用。

## 相关链接
- 首页（下载 + 生成分享页）：{o}/
- 文案提取（语音转文字）：{o}/transcript
- API 文档：{o}/api-docs
- API 控制台：{o}/api-console
"""


@app.get("/sitemap.xml")
def sitemap(request: Request):
    o = _origin(request)

    def entry(path, pri):
        return (f'  <url><loc>{o}{path}</loc>'
                f'<xhtml:link rel="alternate" hreflang="zh-CN" href="{o}{path}"/>'
                f'<xhtml:link rel="alternate" hreflang="en" href="{o}{path}?lang=en"/>'
                f'<lastmod>{_BUILD_DATE}</lastmod>'
                f'<changefreq>daily</changefreq><priority>{pri}</priority></url>\n')

    xml = ('<?xml version="1.0" encoding="UTF-8"?>\n'
           '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" '
           'xmlns:xhtml="http://www.w3.org/1999/xhtml">\n'
           + entry("/", "1.0") + entry("/api-docs", "0.7")
           + entry("/transcript", "0.8")
           + '</urlset>\n')
    return Response(xml, media_type="application/xml")


@app.get("/og.svg")
def og_image():
    # SVG 是可维护源；static/og.png 由 tools/render_og.swift 从同一文件生成。
    return FileResponse("static/og.svg", media_type="image/svg+xml",
                        headers={"Cache-Control": "public, max-age=86400"})


@app.get("/og.png")
def og_png():
    # 社交/微信卡片首选位图：微信、多数抓取器不渲染 SVG，PNG 才能出图（og.svg 保留兜底）
    return FileResponse("static/og.png", media_type="image/png",
                        headers={"Cache-Control": "public, max-age=86400"})


@app.get("/healthz")
def healthz():
    return {"ok": True, "version": APP_VERSION, "proxies": len(proxy_mgr.proxies),
            "enabled": sum(p["enabled"] for p in proxy_mgr.proxies)}
